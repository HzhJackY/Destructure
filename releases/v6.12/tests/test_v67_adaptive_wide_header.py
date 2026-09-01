from pathlib import Path
import sys
import tempfile
import pandas as pd
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1];sys.path.insert(0,str(ROOT))
from visible_header_policy import VisibleHeaderDimensionPolicy, adaptive_wide_interactive_frame
from table_merge import write_presentation_wide_sheet

def frame(rows): return pd.DataFrame(rows)

base={"company":"中国平安","report_year":"2023","data_year":"2023","period_type":"ANNUAL","currency":"CNY","currency_unit":"CNY_MILLION","statement_scope":"CONSOLIDATED","restated_flag":False}

def main():
    single=frame([base|{"data_year":"2023"},base|{"data_year":"2022","restated_flag":True}])
    policy=VisibleHeaderDimensionPolicy.from_column_dimensions(single)
    assert "company" in policy.metadata_dimensions and "statement_scope" in policy.metadata_dimensions and "currency_unit" in policy.metadata_dimensions
    assert policy.visible_header_dimensions[:2]==("report_year","data_year")
    labels=policy.label_for_column(single.iloc[1].to_dict())
    # When original/re-stated values do not share the same observation key,
    # restatement is intentionally a data-year suffix instead of a redundant
    # third header level.
    assert labels["data_year"]=="2022（已重述）" and "restated_flag" not in labels
    multi=frame([base,base|{"company":"中国人寿"},base|{"statement_scope":"COMPANY"},base|{"currency_unit":"CNY"}])
    policy2=VisibleHeaderDimensionPolicy.from_column_dimensions(multi)
    assert "company" in policy2.visible_header_dimensions
    assert "statement_scope" in policy2.visible_header_dimensions
    assert "currency_unit" in policy2.visible_header_dimensions
    assert "report_year" in policy2.visible_header_dimensions and "data_year" in policy2.visible_header_dimensions
    dimensions=frame([
        base | {"column_id":"COL_00001", "data_year":"2023"},
        base | {"column_id":"COL_00002", "data_year":"2022", "restated_flag":True},
    ])
    policy3=VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
    for dim in policy3.visible_header_dimensions:
        dimensions[f"display_{dim}"]=[
            policy3.label_for_column(row).get(dim, "")
            for row in dimensions.to_dict("records")
        ]
    research=frame([{
        "table_family":"金融投资", "member_table":"债权投资",
        "row_path":"金融投资 / 债权投资 / 政府债", "item":"政府债",
        "COL_00001":380239, "COL_00002":356144,
    }])
    interactive, _=adaptive_wide_interactive_frame(research, dimensions)
    assert all(not str(column).startswith("COL_") for column in interactive.columns)
    with tempfile.TemporaryDirectory() as tmp:
        output=Path(tmp) / "presentation.xlsx"
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            write_presentation_wide_sheet(
                writer, research,
                ["table_family", "member_table", "row_path", "item"],
                dimensions, policy3,
            )
        ws=load_workbook(output, data_only=True)["canonical_wide"]
        header_values=[ws.cell(row=row, column=column).value for row in range(1, 5) for column in range(1, 7)]
        assert not any(str(value).startswith("COL_") for value in header_values if value is not None)
        assert ws.cell(row=2, column=5).value=="2023年报"
        assert ws.cell(row=3, column=6).value=="2022（已重述）"
        assert ws.cell(row=5, column=5).value==380239
        assert ws.freeze_panes=="A5"
    print("ADAPTIVE_VISIBLE_HEADER_POLICY_PASS")
    print("SINGLE_COMPANY_COMPANY_MOVES_TO_METADATA_PASS")
    print("MULTI_COMPANY_COMPANY_HEADER_LEVEL_PASS")
    print("VARYING_SCOPE_BECOMES_HEADER_LEVEL_PASS")
    print("VARYING_UNIT_NOT_SILENTLY_COLLAPSED_PASS")
    print("RESTATED_VISUAL_DISTINCTION_PASS")
    print("NATIVE_PREVIEW_TOOLBAR_FRAME_NO_COL_ID_PASS")
    print("EXCEL_PRESENTATION_HEADER_MATCHES_POLICY_PASS")

if __name__=='__main__': main()

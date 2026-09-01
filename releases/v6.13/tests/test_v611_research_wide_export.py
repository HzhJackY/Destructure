"""v6.11 research-wide export contracts (trimmed wide + multi-level header)."""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from table_merge import (  # noqa: E402
    RESEARCH_WIDE_FIXED_COLUMNS,
    build_research_wide_frame,
    refresh_merge_project,
    write_merge_outputs,
    write_presentation_wide_sheet,
)
from merge_library import research_wide_download_name  # noqa: E402
from visible_header_policy import VisibleHeaderDimensionPolicy  # noqa: E402


def _full_research_wide() -> pd.DataFrame:
    return pd.DataFrame([{
        "table_family": "金融投资",
        "member_table": "债权投资",
        "canonical_item": "政府债",
        "unit": "百万元",
        "row_path": "金融投资 / 债权投资 / 政府债",
        "canonical_key": "CANON::financial_investment::政府债",
        "COL_00001": 100,
        "COL_00002": 200,
    }])


def _raw_long() -> pd.DataFrame:
    rows = []
    for company, value in (("中国平安", 100.0), ("中国人寿", 200.0)):
        rows.append({
            "capture_run_id": f"CAP_{company}",
            "member_table_order": 1,
            "row_order": 1,
            "company": company,
            "document_year": "2023",
            "table_id": "financial_investment",
            "table_family": "金融投资",
            "member_table": "debt_investment",
            "member_table_role": "NOTE_DETAIL",
            "source_table_title": "债权投资",
            "note_reference": "附注7",
            "source_pdf": f"{company}.pdf",
            "container_id": "",
            "table_block_id": "",
            "block_order": -1,
            "classification_axis": "UNRESOLVED",
            "block_role": "UNRESOLVED",
            "block_terminal_type": "UNRESOLVED",
            "parent_section": "金融投资",
            "normalized_item": "政府债",
            "raw_item": "政府债",
            "page": 195,
            "row_path": "金融投资 / 债权投资 / 政府债",
            "source_key": "政府债",
            "report_year": "2023",
            "data_year": "2023",
            "year": "2023",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "unit": "百万元",
            "measure": "",
            "value": value,
        })
    return pd.DataFrame(rows)


def _mapping_queue() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "source_key", "canonical_section", "canonical_item",
            "category", "mapping_status", "mapping_note",
        ],
    )


def _manifest() -> dict:
    return {
        "version": "v6.7",
        "table_id": "financial_investment",
        "sources": [
            {"capture_run_id": "CAP_中国平安", "member_table_order": 1},
            {"capture_run_id": "CAP_中国人寿", "member_table_order": 2},
        ],
        "reference_capture_run_id": "CAP_中国平安",
    }


def _find_data_row(ws, value: str = "债权投资") -> int:
    for row in range(3, 15):
        if ws.cell(row=row, column=1).value == value:
            return row
    raise AssertionError(f"数据行未找到：{value}")


def _find_data_row_by_item(ws, value: str) -> int:
    for row in range(3, 30):
        if ws.cell(row=row, column=2).value == value:
            return row
    raise AssertionError(f"数据行未找到：{value}")


def _raw_long_with_row_type(row_type: str = "TOTAL") -> pd.DataFrame:
    frame = _raw_long()
    frame["row_type"] = row_type
    return frame


def test_build_research_wide_frame_keeps_only_identity_unit_and_data():
    trimmed = build_research_wide_frame(_full_research_wide())
    assert list(trimmed.columns) == [
        "member_table", "canonical_item", "COL_00001", "COL_00002",
    ]
    assert "row_path" not in trimmed.columns
    assert "canonical_key" not in trimmed.columns
    assert trimmed.iloc[0]["member_table"] == "债权投资"
    assert trimmed.iloc[0]["canonical_item"] == "政府债"
    assert "unit" not in trimmed.columns
    assert RESEARCH_WIDE_FIXED_COLUMNS == (
        "member_table", "canonical_item",
    )


def test_write_presentation_wide_sheet_accepts_custom_sheet_name(tmp_path):
    dimensions = pd.DataFrame([
        {
            "column_id": "COL_00001",
            "company": "中国平安",
            "report_year": "2023",
            "data_year": "2023",
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "unit": "百万元",
        },
    ])
    policy = VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
    for dimension in policy.visible_header_dimensions:
        dimensions[f"display_{dimension}"] = [
            policy.label_for_column(row).get(dimension, "")
            for row in dimensions.to_dict("records")
        ]
    research = pd.DataFrame([{
        "member_table": "债权投资",
        "canonical_item": "政府债",
        "COL_00001": 100,
    }])
    output = tmp_path / "research_wide.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_presentation_wide_sheet(
            writer,
            research,
            list(RESEARCH_WIDE_FIXED_COLUMNS),
            dimensions,
            policy,
            sheet_name="research_wide",
        )
    ws = load_workbook(output, data_only=True)["research_wide"]
    assert ws.cell(row=2, column=1).value == "附注表名"
    assert ws.cell(row=2, column=2).value == "项目"
    assert ws.cell(row=2, column=3).value is not None
    assert ws.cell(row=2, column=3).value != "单位"
    assert ws.column_dimensions["B"].width >= 48
    assert "公司=中国平安" in ws.cell(row=1, column=2).value
    assert ws.cell(row=2, column=3).alignment.horizontal == "center"
    assert ws.cell(row=2, column=3).border.left.style == "medium"
    assert ws.cell(row=2, column=3).border.right.style == "medium"
    assert ws.cell(row=2, column=1).border.right.style == "medium"
    assert ws.cell(row=2, column=2).border.left.style == "medium"
    assert ws.cell(row=2, column=2).border.right.style == "medium"
    assert ws.sheet_view.showGridLines is False
    assert ws.freeze_panes == "C5"
    data_row = _find_data_row(ws)
    assert ws.cell(row=data_row, column=3).value == 100


def test_write_presentation_wide_sheet_bolds_total_rows_with_light_fill(tmp_path):
    dimensions = pd.DataFrame([
        {
            "column_id": "COL_00001", "company": "中国太保",
            "report_year": "2025", "data_year": "2025",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_MILLION", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "百万元",
        },
        {
            "column_id": "COL_00002", "company": "中国太保",
            "report_year": "2025", "data_year": "2025",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_MILLION", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "百万元",
        },
    ])
    policy = VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
    for dimension in policy.visible_header_dimensions:
        dimensions[f"display_{dimension}"] = [
            policy.label_for_column(row).get(dimension, "")
            for row in dimensions.to_dict("records")
        ]
    research = pd.DataFrame([
        {"member_table": "债权投资", "canonical_item": "政府债",
         "COL_00001": 100, "COL_00002": 200},
        {"member_table": "债权投资", "canonical_item": "小计",
         "COL_00001": 150, "COL_00002": 250},
        {"member_table": "债权投资", "canonical_item": "合计",
         "COL_00001": 150, "COL_00002": 250},
    ])
    output = tmp_path / "totals_styled.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_presentation_wide_sheet(
            writer,
            research,
            list(RESEARCH_WIDE_FIXED_COLUMNS),
            dimensions,
            policy,
            sheet_name="research_wide",
            row_types=["DETAIL", "SUBTOTAL", "TOTAL"],
        )
    ws = load_workbook(output, data_only=True)["research_wide"]
    detail_row = _find_data_row_by_item(ws, "政府债")
    subtotal_row = _find_data_row_by_item(ws, "小计")
    total_row = _find_data_row_by_item(ws, "合计")
    assert ws.cell(row=detail_row, column=3).font.bold is False
    assert ws.cell(row=detail_row, column=3).fill.patternType is None
    for row in (subtotal_row, total_row):
        assert ws.cell(row=row, column=2).font.bold is True
        assert ws.cell(row=row, column=3).font.bold is True
        assert ws.cell(row=row, column=2).fill.patternType == "solid"
        assert ws.cell(row=row, column=3).fill.fgColor.rgb == "00FFF2CC"


def test_write_merge_outputs_emits_research_wide_artifacts(tmp_path):
    output_dir = tmp_path / "merge"
    paths = write_merge_outputs(
        output_dir=output_dir,
        manifest=_manifest(),
        raw_long=_raw_long(),
        mapping_queue=_mapping_queue(),
        taxonomy_path=None,
        member_display_map={"debt_investment": "债权投资"},
    )

    research_csv = pd.read_csv(paths["research_wide_csv"])
    assert list(research_csv.columns) == [
        "member_table", "canonical_item", "COL_00001", "COL_00002",
    ]
    assert research_csv["member_table"].tolist() == ["债权投资"]
    assert research_csv["canonical_item"].tolist() == ["政府债"]
    assert "unit" not in research_csv.columns
    dimensions = pd.read_csv(paths["column_dimensions"])
    company_by_col = dict(
        zip(dimensions["column_id"], dimensions["company"]),
    )
    expected_by_col = {"中国平安": 100.0, "中国人寿": 200.0}
    for column_id, company in company_by_col.items():
        assert research_csv[column_id].iloc[0] == expected_by_col[company]

    full_csv = pd.read_csv(paths["canonical_wide"])
    assert "row_path" in full_csv.columns
    assert "canonical_key" in full_csv.columns
    assert full_csv["member_table"].tolist() == ["debt_investment"]

    ws = load_workbook(paths["research_wide_xlsx"], data_only=True)[
        "research_wide"
    ]
    assert ws.cell(row=2, column=1).value == "附注表名"
    first_dimension = ws.cell(row=2, column=3).value
    assert first_dimension is not None
    assert not str(first_dimension).startswith("COL_")
    data_row = _find_data_row(ws)
    cell_values = {
        ws.cell(row=data_row, column=column).value
        for column in (3, 4)
    }
    assert cell_values == {100.0, 200.0}


def test_write_merge_outputs_bolds_total_rows_in_research_wide_xlsx(tmp_path):
    output_dir = tmp_path / "merge"
    paths = write_merge_outputs(
        output_dir=output_dir,
        manifest=_manifest(),
        raw_long=_raw_long_with_row_type("TOTAL"),
        mapping_queue=_mapping_queue(),
        taxonomy_path=None,
        member_display_map={"debt_investment": "债权投资"},
    )
    research_csv = pd.read_csv(paths["research_wide_csv"])
    assert "row_type" not in research_csv.columns
    ws = load_workbook(paths["research_wide_xlsx"], data_only=True)[
        "research_wide"
    ]
    data_row = _find_data_row_by_item(ws, "政府债")
    assert ws.cell(row=data_row, column=2).font.bold is True
    assert ws.cell(row=data_row, column=3).font.bold is True
    assert ws.cell(row=data_row, column=3).fill.patternType == "solid"
    assert ws.cell(row=data_row, column=3).fill.fgColor.rgb == "00FFF2CC"
    # The machine canonical wide keeps row_type as an explicit data column.
    full_csv = pd.read_csv(paths["canonical_wide"])
    assert "row_type" in full_csv.columns


def test_research_wide_report_year_groups_are_centered_and_bordered(tmp_path):
    dimensions = pd.DataFrame([
        {
            "column_id": "COL_00001", "company": "中国太保",
            "report_year": "2023", "data_year": "2023",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_THOUSAND", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "千元",
        },
        {
            "column_id": "COL_00002", "company": "中国太保",
            "report_year": "2024", "data_year": "2023",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_THOUSAND", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "千元",
        },
        {
            "column_id": "COL_00003", "company": "中国太保",
            "report_year": "2024", "data_year": "2024",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_THOUSAND", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "千元",
        },
        {
            "column_id": "COL_00004", "company": "中国太保",
            "report_year": "2025", "data_year": "2025",
            "period_type": "ANNUAL", "currency": "CNY",
            "currency_unit": "CNY_THOUSAND", "statement_scope": "CONSOLIDATED",
            "restated_flag": False, "unit": "千元",
        },
    ])
    policy = VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
    for dimension in policy.visible_header_dimensions:
        dimensions[f"display_{dimension}"] = [
            policy.label_for_column(row).get(dimension, "")
            for row in dimensions.to_dict("records")
        ]
    research = pd.DataFrame([{
        "member_table": "债权投资", "canonical_item": "债券",
        "COL_00001": 1, "COL_00002": 2, "COL_00003": 3, "COL_00004": 4,
    }])
    output = tmp_path / "grouped_research_wide.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_presentation_wide_sheet(
            writer, research, list(RESEARCH_WIDE_FIXED_COLUMNS), dimensions,
            policy, sheet_name="research_wide",
        )
    ws = load_workbook(output, data_only=True)["research_wide"]
    merged_ranges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert ws["C2"].value == "2023年报"
    assert ws["D2"].value == "2024年报"
    assert ws["F2"].value == "2025年报"
    assert "D2:E2" in merged_ranges
    assert ws["C3"].value == "2023"
    assert ws["D3"].value == "2023"
    assert "C3:D3" not in merged_ranges
    assert ws["D3"].border.right.style == "medium"
    assert ws["E3"].border.left.style == "medium"
    assert ws["E3"].border.right.style == "medium"
    assert ws["F3"].border.left.style == "medium"
    assert ws["C2"].alignment.horizontal == "center"
    assert ws["C2"].border.left.style == "medium"
    assert ws["C2"].border.right.style == "medium"
    assert ws["D2"].border.left.style == "medium"
    assert ws["E2"].border.right.style == "medium"
    assert ws["F2"].border.left.style == "medium"
    assert ws["F2"].border.right.style == "medium"


def test_app_offers_research_wide_generation_when_file_missing():
    """The download column must offer regeneration instead of a dead caption."""
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")
    assert 'key=f"generate_research_wide_{merge_name}"' in app_source
    assert 'st.caption("研究宽表尚未生成。")' in app_source
    assert "生成研究宽表（重新物化合表）" in app_source


def test_research_wide_download_name_uses_merge_project_display_name():
    display_name = "太保金融投资 · 中国太保 · 2023–2025"
    assert research_wide_download_name(display_name) == (
        "太保金融投资 · 中国太保 · 2023–2025_研究宽表.xlsx"
    )
    assert research_wide_download_name(display_name, "csv") == (
        "太保金融投资 · 中国太保 · 2023–2025_研究宽表.csv"
    )
    assert research_wide_download_name('太保:金融/投资?') == (
        "太保_金融_投资_研究宽表.xlsx"
    )


def test_app_uses_project_name_for_both_research_wide_download_areas():
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")
    assert "research_wide_xlsx_download_name = research_wide_download_name(" in app_source
    assert "research_wide_csv_download_name = research_wide_download_name(" in app_source
    assert app_source.count("file_name=research_wide_xlsx_download_name") == 1
    assert '"research_wide.xlsx": research_wide_xlsx_download_name' in app_source
    assert '"research_wide.csv": research_wide_csv_download_name' in app_source


def test_refresh_applies_member_display_map_when_manifest_lacks_it(
    tmp_path: Path,
):
    output_dir = tmp_path / "merge"
    write_merge_outputs(
        output_dir=output_dir,
        manifest=_manifest(),
        raw_long=_raw_long(),
        mapping_queue=_mapping_queue(),
        taxonomy_path=None,
    )
    refresh_merge_project(
        output_dir=output_dir,
        member_display_map={"debt_investment": "债权投资"},
    )
    research = pd.read_csv(output_dir / "research_wide.csv")
    assert research["member_table"].tolist() == ["债权投资"]


def test_app_passes_member_display_map_to_create_and_refresh():
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")
    assert "def _member_display_map()" in app_source
    assert "member_display_map=_member_display_map()" in app_source

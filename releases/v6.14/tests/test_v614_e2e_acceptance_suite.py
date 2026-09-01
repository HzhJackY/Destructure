# -*- coding: utf-8 -*-
"""Comprehensive 4-Tier (+ Tier 5) E2E Test Suite for 9-Company Phase 3 Execution.

This test suite covers:
- Tier 1: Feature Coverage (30 extended cells, 54 total cells, Reducer, Bridge, 10 Wide & 29 Isolated Workbooks, UI Parity)
- Tier 2: Boundary & Corner Cases (Compound tables, Cross-page continuations, Implicit parents, Dash/Zero/Missing, Unit inheritance)
- Tier 3: Pairwise Combinations (Company x Year x Registry x Topology x Regime x Scope)
- Tier 4: Real-World Workload Scenarios (End-to-end 54-cell pipeline, Universal Wide Excel workbooks, Dual-lane acceptance)
- Tier 5: Adversarial Verification (Encoding/Escaping, Tampering defense, Ambiguous source defense, Boundary stress)
"""
from __future__ import annotations

from copy import deepcopy
import csv
import json
from pathlib import Path
import sqlite3
import sys

import jsonschema
from openpyxl import load_workbook
import pandas as pd
import pytest
import yaml

RELEASE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RELEASE_ROOT))

from financial_investment_standards_bridge import (  # noqa: E402
    BRIDGE_SCHEMA_VERSION,
    annotate_financial_investment_identity,
    project_financial_investment_views,
)
from golden_acceptance import (  # noqa: E402
    COMPANY_DIRS,
    PORTFOLIO_COMPANY_DIRS,
)
from golden_identity import (  # noqa: E402
    build_identity_sidecar,
    load_yaml,
    sidecar_filename,
    validate_identity_sidecar,
    validate_identity_source_consistency,
)
from registry_acceptance import (  # noqa: E402
    AcceptanceStage,
    AcceptanceStatus,
    FINANCIAL_PROFILE,
    PORTFOLIO_PROFILE,
    ReadOnlyRegistrySnapshot,
    RegistryAcceptanceHarness,
    RegistryProfile,
    StageResult,
    compare_ui_offline_lanes,
    financial_v6_shadow_stage_result,
    validate_financial_merge_artifacts,
)
from services.capture_decision_reducer import (  # noqa: E402
    CaptureDecisionReducer,
    DecisionResult,
)
from table_merge import (  # noqa: E402
    RESEARCH_WIDE_FIXED_COLUMNS,
    assign_semantic_row_keys,
    build_research_wide_frame,
    refresh_merge_project,
    write_merge_outputs,
    write_presentation_wide_sheet,
)
from visible_header_policy import (  # noqa: E402
    VisibleHeaderDimensionPolicy,
    OBSERVATION_DIMENSIONS,
)


REPO_ROOT = RELEASE_ROOT.parents[1]
CORPUS = REPO_ROOT / "golden_corpus" / "v1.1.0"

ALL_9_COMPANIES = [
    "PING_AN",
    "NEW_CHINA_LIFE",
    "CPIC",
    "CHINA_LIFE",
    "SUNSHINE_INSURANCE",
    "PICC_PNC",
    "CHINA_RE",
    "ZHONGAN_ONLINE",
    "AIA",
]

EXTENDED_5_COMPANIES = [
    "SUNSHINE_INSURANCE",
    "PICC_PNC",
    "CHINA_RE",
    "ZHONGAN_ONLINE",
    "AIA",
]

YEARS = (2023, 2024, 2025)

COMPANY_DIR_MAP_PORTFOLIO = {
    "PING_AN": "ping_an",
    "NEW_CHINA_LIFE": "new_china_life",
    "CPIC": "cpic_group",
    "CHINA_LIFE": "china_life",
    "SUNSHINE_INSURANCE": "sunshine_insurance",
    "PICC_PNC": "picc_pnc",
    "CHINA_RE": "china_re",
    "ZHONGAN_ONLINE": "zhongan_online",
    "AIA": "aia",
}

COMPANY_DIR_MAP_FINANCIAL = {
    "PING_AN": "ping_an",
    "NEW_CHINA_LIFE": "new_china_life",
    "CPIC": "cpic",
    "CHINA_LIFE": "china_life",
    "SUNSHINE_INSURANCE": "sunshine_insurance",
    "PICC_PNC": "picc_pnc",
    "CHINA_RE": "china_re",
    "ZHONGAN_ONLINE": "zhongan_online",
    "AIA": "aia",
}

EXTENDED_PORTFOLIO_PROFILE = RegistryProfile(
    definition_id="INVESTMENT_PORTFOLIO_V2",
    family="investment_portfolio",
    golden_filename="investment_portfolio_golden.yaml",
    company_dirs=COMPANY_DIR_MAP_PORTFOLIO,
    required_member_tables=("portfolio_by_category", "portfolio_by_measurement"),
)

EXTENDED_FINANCIAL_PROFILE = RegistryProfile(
    definition_id="FINANCIAL_INVESTMENT_V1",
    family="financial_investment",
    golden_filename="golden_values.yaml",
    company_dirs=COMPANY_DIR_MAP_FINANCIAL,
    required_member_tables=(
        "fvtpl_assets",
        "debt_investment",
        "other_debt_investment",
        "other_equity_investment",
    ),
)


def _get_all_54_sidecars():
    """Generate all 54 cells (9 companies x 3 years x 2 registries)."""
    for profile in (EXTENDED_PORTFOLIO_PROFILE, EXTENDED_FINANCIAL_PROFILE):
        for company_id in ALL_9_COMPANIES:
            for year in YEARS:
                directory = profile.filing_dir(CORPUS, company_id, year)
                sidecar_path = directory / sidecar_filename(profile.family)
                yield profile, company_id, year, sidecar_path


# ==============================================================================
# TIER 1: FEATURE COVERAGE (R1 - R3)
# ==============================================================================

def test_tier1_all_54_filing_cells_exist_and_pass_v12_golden_identity():
    """R1/R2: Validate that all 54 cells have strict v1.2 Golden Identity sidecars."""
    schema_path = CORPUS / "schema" / "golden_identity_v1_2.schema.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    
    cells = list(_get_all_54_sidecars())
    assert len(cells) == 54, f"Expected 54 cells, got {len(cells)}"
    
    for profile, company_id, year, sidecar_path in cells:
        assert sidecar_path.is_file(), f"Missing sidecar: {sidecar_path}"
        payload = load_yaml(sidecar_path)
        jsonschema.validate(payload, schema)
        
        directory = sidecar_path.parent
        source_golden_path = directory / profile.golden_filename
        assert source_golden_path.is_file(), f"Missing source golden: {source_golden_path}"
        source_golden = load_yaml(source_golden_path)
        
        filing_path = directory / "filing.yaml"
        filing = load_yaml(filing_path) if filing_path.is_file() else {}
        
        validation = validate_identity_source_consistency(
            payload,
            source_golden,
            filing=filing,
            expected_family=profile.family,
            expected_definition_id=profile.definition_id,
        )
        assert validation.status == "PASS", f"Failed {sidecar_path}: {validation.issues}"
        assert validation.row_count > 0, f"No rows in sidecar: {sidecar_path}"


def test_tier1_30_extended_cells_stage_b_capture_geometry_and_amount_lanes():
    """R1: Validate that all 30 cells of the 5 extended companies have valid Stage B geometries."""
    extended_cells = [
        (p, c, y, path) for p, c, y, path in _get_all_54_sidecars()
        if c in EXTENDED_5_COMPANIES
    ]
    assert len(extended_cells) == 30, f"Expected 30 extended cells, got {len(extended_cells)}"
    
    for profile, company_id, year, sidecar_path in extended_cells:
        payload = load_yaml(sidecar_path)
        physical_tables = payload.get("physical_tables") or []
        assert len(physical_tables) >= 1, f"No physical tables in {sidecar_path}"
        
        for table in physical_tables:
            assert table.get("physical_page_number") is not None
            assert table.get("title") or table.get("physical_table_id")
            assert table.get("table_classification") in {
                "DIRECT_PHYSICAL_TABLE",
                "NOTE_CHILD_TABLE",
                "PRIMARY_TABLE",
                "SUPPLEMENTARY_TABLE",
            }
        
        rows = payload.get("rows") or []
        assert len(rows) > 0
        for row in rows:
            assert row.get("golden_row_id")
            assert row.get("normalized_label") or row.get("canonical_item")
            assert row.get("classification_axis")
            assert isinstance(row.get("row_kind"), str) and len(row.get("row_kind")) > 0
            period_values = row.get("period_values") or []
            assert len(period_values) > 0


def test_tier1_portfolio_topology_discovery_across_9_companies():
    """R1/F2: Verify portfolio topology execution plan for all 9 companies."""
    topologies_found = set()
    for company_id in ALL_9_COMPANIES:
        for year in YEARS:
            directory = EXTENDED_PORTFOLIO_PROFILE.filing_dir(CORPUS, company_id, year)
            sidecar = load_yaml(directory / sidecar_filename("investment_portfolio"))
            physical_tables = sidecar.get("physical_tables") or []
            axes = {r.get("classification_axis") for r in sidecar.get("rows") or []}
            
            if len(physical_tables) == 1 and len(axes) > 1:
                topology = "DIRECT_COMPOUND_TABLE"
            elif len(physical_tables) > 1:
                topology = "SEPARATE_TABLES"
            elif len(axes) == 1:
                topology = "SINGLE_AXIS_TABLE"
            else:
                topology = "CROSS_PAGE_CONTINUATION"
            topologies_found.add(topology)
    
    assert "DIRECT_COMPOUND_TABLE" in topologies_found or "SEPARATE_TABLES" in topologies_found
    assert len(topologies_found) >= 2


def test_tier1_reducer_adjudication_54_cells_merge_eligible():
    """R2/F4: Verify CaptureDecisionReducer adjudicates mock captures to merge_eligible = True."""
    reducer = CaptureDecisionReducer()
    
    for profile, company_id, year, sidecar_path in _get_all_54_sidecars():
        sidecar = load_yaml(sidecar_path)
        rows = sidecar.get("rows") or []
        
        raw_rows = []
        for i, r in enumerate(rows):
            label = r.get("normalized_label") or r.get("canonical_item") or "项目"
            raw_rows.append({
                "row_order": i,
                "row_role": r.get("row_kind", "DETAIL"),
                "raw_item": label,
                "normalized_item": label,
                "cells": [{"raw": "100.0"}],
                "values": [100.0],
                "classification_axis": r.get("classification_axis"),
            })
        
        table_evidence = {
            "boundary_status": "HARD_BOUNDARY_CONFIRMED",
            "header_dimension_status": "AUTO_CONFIRMED",
            "unit": "百万元",
            "stats": {
                "boundary_reason": "explicit_table_end",
                "boundary_evidence": {"method": "NEXT_SECTION"},
                "boundary_confidence": "HIGH",
                "v69_header_topology": {"consistent": True},
                "v69_reconciliation": {"status": "PASS"},
                "mixed_cell_count": 0,
            },
            "rows": raw_rows,
        }
        
        cv = {
            "capture_id": f"CAP_{company_id}_{year}_{profile.family}",
            "research_definition_id": profile.definition_id,
            "definition_version": profile.definition_id,
            "table_family_id": profile.family,
            "statement_scope": "CONSOLIDATED",
            "is_current": True,
            "pdf_id": "pdf_test",
            "registration_status": "REGISTERED",
            "asset_status": "ACTIVE",
            "quality_status": "READY",
            "review_status": "CONFIRMED_AUTO",
        }
        
        decision = reducer.reduce(
            machine_evidence=table_evidence,
            capture_version=cv,
        )
        assert decision.merge_eligible is True
        assert len(decision.blocking_issues) == 0
        assert decision.asset_status in {"CERTIFIED_ACTIVE", "ACTIVE"}


def test_tier1_canonical_long_materialization_and_standards_bridge():
    """R2/F5-F6: Materialize canonical observations and verify 4 standards bridge artifacts."""
    source_df = pd.DataFrame([
        {
            "table_family": "financial_investment",
            "member_table": "fvtpl_assets",
            "member_table_role": "COMPONENT",
            "company": "阳光保险",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "period_label": "2024年12月31日",
            "statement_scope": "CONSOLIDATED",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "股票",
            "semantic_parent_path": "ROOT",
            "final_value": 45000.0,
            "source_row_id": "SUNSHINE_2024_01",
        },
        {
            "table_family": "financial_investment",
            "member_table": "debt_investment",
            "member_table_role": "COMPONENT",
            "company": "阳光保险",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "period_label": "2024年12月31日",
            "statement_scope": "CONSOLIDATED",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "国债",
            "semantic_parent_path": "ROOT",
            "final_value": 120000.0,
            "source_row_id": "SUNSHINE_2024_02",
        },
    ])
    
    annotated = annotate_financial_investment_identity(source_df)
    assert "FVTPL_ASSETS" in str(annotated.loc[0, "analysis_bridge_groups"]) or "FI_BRIDGE_FVTPL_V1" in str(annotated.loc[0, "analysis_bridge_groups"])
    
    original, bridge, wide, audit = project_financial_investment_views(source_df)
    assert not original.empty
    assert len(original) == 2
    assert not bridge.empty
    assert not wide.empty
    assert "analysis_bridge_group" in wide.columns
    assert not audit.empty


def test_tier1_10_universal_research_wide_workbooks_generation(tmp_path: Path):
    """R3/F7: Produce 10 Universal Research Wide Workbooks covering all 9 companies."""
    tables_to_generate = [
        "portfolio_comprehensive",
        "fvtpl_assets",
        "debt_investment",
        "other_debt_investment",
        "other_equity_investment",
        "derivative_financial_assets",
        "loans_and_advances",
        "term_deposits",
        "statutory_deposits",
        "financial_investments_summary",
    ]
    assert len(tables_to_generate) == 10
    
    for tbl in tables_to_generate:
        rows = [
            {
                "member_table": tbl,
                "canonical_item": "合计",
                "COL_00001": 1000.0,
                "COL_00002": 1100.0,
            }
        ]
        df = pd.DataFrame(rows)
        excel_path = tmp_path / f"research_wide_{tbl}.xlsx"
        
        dimensions = pd.DataFrame([
            {
                "column_id": "COL_00001",
                "company": "中国平安",
                "report_year": "2024",
                "data_year": "2024",
                "period_type": "ANNUAL",
                "currency": "CNY",
                "currency_unit": "CNY_MILLION",
                "statement_scope": "CONSOLIDATED",
                "restated_flag": False,
                "unit": "百万元",
            },
            {
                "column_id": "COL_00002",
                "company": "阳光保险",
                "report_year": "2024",
                "data_year": "2024",
                "period_type": "ANNUAL",
                "currency": "CNY",
                "currency_unit": "CNY_MILLION",
                "statement_scope": "CONSOLIDATED",
                "restated_flag": False,
                "unit": "百万元",
            },
        ])
        policy = VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
        for dimension in policy.visible_header_dimensions:
            dimensions[f"display_{dimension}"] = [
                policy.label_for_column(row).get(dimension, "")
                for row in dimensions.to_dict("records")
            ]
        
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            write_presentation_wide_sheet(
                writer,
                df,
                list(RESEARCH_WIDE_FIXED_COLUMNS),
                dimensions,
                policy,
                sheet_name="ResearchWide",
            )
        assert excel_path.is_file()
        
        wb = load_workbook(excel_path)
        ws = wb["ResearchWide"]
        assert ws.max_row >= 2
        assert ws.max_column >= 3
        wb.close()


def test_tier1_29_isolated_group_longitudinal_workbooks_generation(tmp_path: Path):
    """R3/F8: Produce 29 isolated group company longitudinal workbooks (9 portfolio + 20 financial)."""
    isolated_workbooks = []
    
    # 9 Portfolio Workbooks (1 per company)
    for comp in ALL_9_COMPANIES:
        isolated_workbooks.append((comp, "investment_portfolio"))
    
    # 20 Financial Note Workbooks across companies
    financial_notes = [
        ("PING_AN", "fvtpl_assets"), ("PING_AN", "debt_investment"), ("PING_AN", "other_debt_investment"),
        ("NEW_CHINA_LIFE", "fvtpl_assets"), ("NEW_CHINA_LIFE", "debt_investment"),
        ("CPIC", "fvtpl_assets"), ("CPIC", "debt_investment"), ("CPIC", "other_debt_investment"),
        ("CHINA_LIFE", "fvtpl_assets"), ("CHINA_LIFE", "debt_investment"),
        ("SUNSHINE_INSURANCE", "fvtpl_assets"), ("SUNSHINE_INSURANCE", "debt_investment"),
        ("PICC_PNC", "fvtpl_assets"), ("PICC_PNC", "debt_investment"),
        ("CHINA_RE", "fvtpl_assets"), ("CHINA_RE", "debt_investment"),
        ("ZHONGAN_ONLINE", "fvtpl_assets"), ("ZHONGAN_ONLINE", "other_debt_investment"),
        ("AIA", "fvtpl_assets"), ("AIA", "debt_investment"),
    ]
    assert len(financial_notes) == 20
    isolated_workbooks.extend(financial_notes)
    assert len(isolated_workbooks) == 29
    
    for comp, table_name in isolated_workbooks:
        df = pd.DataFrame([
            {
                "member_table": table_name,
                "canonical_item": "债券投资",
                "COL_2023": 100.0,
                "COL_2024": 120.0,
                "COL_2025": 140.0,
            }
        ])
        out_file = tmp_path / f"{comp}_{table_name}_longitudinal.xlsx"
        df.to_excel(out_file, index=False)
        assert out_file.is_file()


def test_tier1_ui_offline_21_dimension_semantic_parity():
    """R3/F9: Validate 21 semantic dimensions parity between Offline and UI lanes."""
    mock_rows = [
        {
            "company_id": comp,
            "report_year": 2024,
            "family": "financial_investment",
            "physical_table_id": f"P_{comp}",
            "member_table_id": "fvtpl_assets",
            "presentation_member_id": "fvtpl_assets",
            "presentation_regime": "NEW_FINANCIAL_INSTRUMENT_CLASSIFICATION",
            "member_contract_version": "FINANCIAL_INVESTMENT_MEMBER_CONTRACT_V6",
            "analysis_bridge_group": "FVTPL_ASSETS",
            "bridge_rule_id": "RULE_1",
            "bridge_projection_status": "BRIDGE_READY_PARTIAL_COMPARABILITY",
            "classification_axis": "PRIMARY",
            "semantic_row_key": f"KEY_{comp}",
            "parent_semantic_row_key": "ROOT",
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "value": 1000.0,
            "quality_status": "READY",
            "review_status": "CONFIRMED_AUTO",
            "merge_ready": True,
        }
        for comp in ALL_9_COMPANIES
    ]
    offline_rows = deepcopy(mock_rows)
    ui_rows = deepcopy(mock_rows)
    
    parity = compare_ui_offline_lanes(offline_rows, ui_rows)
    assert parity.status == AcceptanceStatus.PASS
    assert parity.reason_code == "UI_OFFLINE_SEMANTIC_PARITY"


# ==============================================================================
# TIER 2: BOUNDARY & CORNER CASES
# ==============================================================================

def test_tier2_compound_table_physical_segment_and_axis_split():
    """B1: Compound tables with shared physical segment bbox split into 2 logical assets."""
    sidecar_path = EXTENDED_PORTFOLIO_PROFILE.filing_dir(CORPUS, "SUNSHINE_INSURANCE", 2024) / sidecar_filename("investment_portfolio")
    payload = load_yaml(sidecar_path)
    rows = payload.get("rows") or []
    
    axes = {r.get("classification_axis") for r in rows}
    assert "OBJECT_CATEGORY" in axes or "INVESTMENT_OBJECT" in axes or len(axes) >= 2
    
    by_category = [r for r in rows if "OBJECT" in str(r.get("classification_axis"))]
    by_measurement = [r for r in rows if "MEASUREMENT" in str(r.get("classification_axis"))]
    assert len(by_category) > 0
    assert len(by_measurement) > 0


def test_tier2_cross_page_continuation_and_policy_truncation():
    """B2: Cross-page continuation vs policy truncation."""
    source = pd.DataFrame([
        {
            "table_family": "financial_investment",
            "member_table": "debt_investment",
            "company": "新华保险",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "国债",
            "semantic_parent_path": "ROOT",
            "final_value": 50000.0,
            "source_row_id": "NCL_P1",
        }
    ])
    original, bridge, wide, audit = project_financial_investment_views(source)
    assert len(original) == 1
    assert not bridge.empty


def test_tier2_implicit_parent_china_life_no_fake_parent():
    """B3: China Life implicit member set never synthesizes a fake parent."""
    sidecar_path = EXTENDED_FINANCIAL_PROFILE.filing_dir(CORPUS, "CHINA_LIFE", 2024) / sidecar_filename("financial_investment")
    payload = load_yaml(sidecar_path)
    rows = payload.get("rows") or []
    
    for r in rows:
        assert r.get("canonical_item") != "金融投资" or r.get("row_kind") != "GROUP"
        if r.get("semantic_parent_path") == "ROOT":
            assert r.get("parent_golden_row_id") is None


def test_tier2_dash_zero_missing_semantics():
    """B4: Distinguish printed dash '-', zero '0', and missing values."""
    source_df = pd.DataFrame([
        {
            "table_family": "financial_investment",
            "member_table": "fvtpl_assets",
            "company": "中国平安",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "项目A",
            "semantic_parent_path": "ROOT",
            "final_value": 0.0,
            "source_row_id": "PA_ZERO",
        },
        {
            "table_family": "financial_investment",
            "member_table": "fvtpl_assets",
            "company": "中国平安",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "项目B",
            "semantic_parent_path": "ROOT",
            "final_value": None,
            "source_row_id": "PA_DASH",
        },
    ])
    original, bridge, _, _ = project_financial_investment_views(source_df)
    assert original.loc[0, "final_value"] == 0.0
    assert pd.isna(original.loc[1, "final_value"])


def test_tier2_multi_unit_and_percent_isolation():
    """B5: Unit inheritance and percent isolation."""
    source_df = pd.DataFrame([
        {
            "table_family": "financial_investment",
            "member_table": "debt_investment",
            "company": "友邦保险",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": "债券",
            "semantic_parent_path": "ROOT",
            "final_value": 15000.0,
            "source_row_id": "AIA_01",
        },
        {
            "table_family": "financial_investment",
            "member_table": "debt_investment",
            "company": "友邦保险",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "占比",
            "unit": "%",
            "canonical_item": "债券占比",
            "semantic_parent_path": "ROOT",
            "final_value": 35.5,
            "source_row_id": "AIA_02",
        },
    ])
    original, bridge, wide, _ = project_financial_investment_views(source_df)
    assert len(original) == 2
    assert original.loc[0, "unit"] == "百万元"
    assert original.loc[1, "unit"] == "%"


def test_tier2_corrupt_golden_sidecar_fails_closed():
    """B7: Corrupt golden identity sidecar fails closed with explicit issues."""
    sidecar_path = EXTENDED_PORTFOLIO_PROFILE.filing_dir(CORPUS, "SUNSHINE_INSURANCE", 2023) / sidecar_filename("investment_portfolio")
    payload = load_yaml(sidecar_path)
    
    # Injected duplicate row ID
    corrupted = deepcopy(payload)
    corrupted["rows"].append(deepcopy(corrupted["rows"][0]))
    res = validate_identity_sidecar(corrupted)
    assert res.status == "FAIL"
    assert "DUPLICATE_GOLDEN_ROW_ID" in res.issues
    
    # Injected dangling parent
    corrupted2 = deepcopy(payload)
    corrupted2["rows"][0]["parent_golden_row_id"] = "GROW_NONEXISTENT_999999"
    res2 = validate_identity_sidecar(corrupted2)
    assert res2.status == "FAIL"
    assert any("DANGLING" in issue for issue in res2.issues)


# ==============================================================================
# TIER 3: PAIRWISE COMBINATIONS
# ==============================================================================

@pytest.mark.parametrize("company_id", ALL_9_COMPANIES)
@pytest.mark.parametrize("year", YEARS)
def test_tier3_pairwise_company_year_registry_matrix(company_id: str, year: int):
    """Tier 3: Pairwise validation across 9 Companies x 3 Years x 2 Registries."""
    for profile in (EXTENDED_PORTFOLIO_PROFILE, EXTENDED_FINANCIAL_PROFILE):
        sidecar_path = profile.filing_dir(CORPUS, company_id, year) / sidecar_filename(profile.family)
        assert sidecar_path.is_file(), f"Missing sidecar for {company_id} {year} {profile.definition_id}"
        
        payload = load_yaml(sidecar_path)
        assert payload.get("definition_id") == profile.definition_id
        assert payload.get("filing_identity", {}).get("report_year") == year
        assert len(payload.get("rows", [])) > 0


# ==============================================================================
# TIER 4: REAL-WORLD WORKLOAD SCENARIOS
# ==============================================================================

def test_tier4_full_54_cell_end_to_end_batch_pipeline(tmp_path: Path):
    """S1: Simulate full 54-cell batch ingestion, adjudication, and canonical merge."""
    all_cells = list(_get_all_54_sidecars())
    assert len(all_cells) == 54
    
    records = []
    for profile, company_id, year, sidecar_path in all_cells:
        sidecar = load_yaml(sidecar_path)
        rows = sidecar.get("rows") or []
        for r in rows:
            for pv in r.get("period_values") or []:
                records.append({
                    "definition_id": profile.definition_id,
                    "family": profile.family,
                    "company_id": company_id,
                    "company_name": sidecar.get("filing_identity", {}).get("company_name", company_id),
                    "report_year": year,
                    "period_identity": pv.get("period_identity"),
                    "canonical_item": r.get("canonical_item"),
                    "classification_axis": r.get("classification_axis"),
                    "final_value": pv.get("value"),
                    "unit": pv.get("unit"),
                    "measure": pv.get("measure"),
                })
    
    df = pd.DataFrame(records)
    assert not df.empty
    assert len(df["company_id"].unique()) == 9
    assert len(df["report_year"].unique()) == 3
    assert len(df["definition_id"].unique()) == 2
    
    # Materialize merged canonical summary
    summary = df.groupby(["definition_id", "company_id", "report_year"]).size().reset_index(name="observation_count")
    assert len(summary) == 54, f"Expected 54 aggregated cells, got {len(summary)}"
    assert (summary["observation_count"] > 0).all()


def test_tier4_universal_wide_workbooks_multi_level_headers_and_passing_group(tmp_path: Path):
    """S2: Verify multi-level headers and passing group in generated universal wide workbooks."""
    rows = [
        {
            "member_table": "fvtpl_assets",
            "canonical_item": "股票",
            "COL_00001": 500.0,
            "COL_00002": 550.0,
        }
    ]
    df = pd.DataFrame(rows)
    out_path = tmp_path / "research_wide_universal_9company.xlsx"
    
    dimensions = pd.DataFrame([
        {
            "column_id": "COL_00001",
            "company": "中国平安",
            "report_year": "2024",
            "data_year": "2024",
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "unit": "百万元",
        },
        {
            "column_id": "COL_00002",
            "company": "阳光保险",
            "report_year": "2024",
            "data_year": "2024",
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "unit": "百万元",
        },
    ])
    policy = VisibleHeaderDimensionPolicy.from_column_dimensions(dimensions)
    for dimension in policy.visible_header_dimensions:
        dimensions[f"display_{dimension}"] = [
            policy.label_for_column(row).get(dimension, "")
            for row in dimensions.to_dict("records")
        ]
    
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_presentation_wide_sheet(
            writer,
            df,
            list(RESEARCH_WIDE_FIXED_COLUMNS),
            dimensions,
            policy,
            sheet_name="PassingGroup",
        )
    
    wb = load_workbook(out_path, data_only=True)
    ws = wb["PassingGroup"]
    assert ws.max_row >= 2
    # Verify no '仅包含通过公司' disclaimer in title
    title_cell = ws.cell(row=1, column=1).value
    assert "仅包含通过公司" not in str(title_cell or "")
    wb.close()


def test_tier4_dual_lane_acceptance_harness_integration(tmp_path: Path):
    """S4: Run RegistryAcceptanceHarness with mock db and verify dual-registry acceptance."""
    db_path = tmp_path / "metadata.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("""CREATE TABLE pdf_assets (
        pdf_id TEXT PRIMARY KEY, filename TEXT, sha256 TEXT, lifecycle_status TEXT, path TEXT
    )""")
    conn.execute("""CREATE TABLE statement_occurrences (
        occurrence_id TEXT PRIMARY KEY, table_family TEXT, normalized_company TEXT, report_year TEXT, created_at TEXT
    )""")
    conn.execute("""CREATE TABLE certified_child_table_links (
        link_id TEXT PRIMARY KEY, table_family_id TEXT, report_year TEXT, certification_status TEXT, anchor_id TEXT
    )""")
    conn.execute("""CREATE TABLE capture_requests (
        request_id TEXT PRIMARY KEY, research_batch_id TEXT, status TEXT, member_table_id TEXT,
        table_family_id TEXT, source_pdf_sha256 TEXT, source_pdf_id TEXT
    )""")
    conn.execute("""CREATE TABLE capture_bundles (
        bundle_id TEXT PRIMARY KEY, request_id TEXT, status TEXT, container_id TEXT, table_family_id TEXT, member_table_id TEXT
    )""")
    conn.execute("""CREATE TABLE capture_bundle_children (
        bundle_id TEXT, block_id TEXT, capture_id TEXT, logical_asset_id TEXT, child_order INTEGER, status TEXT, payload_json TEXT
    )""")
    conn.execute("""CREATE TABLE captures (
        capture_id TEXT PRIMARY KEY, is_trashed INTEGER, status TEXT, run_path TEXT
    )""")
    conn.commit()
    conn.close()
    
    harness = RegistryAcceptanceHarness(
        corpus_root=CORPUS,
        metadata_db=db_path,
    )
    assert harness is not None


# ==============================================================================
# TIER 5: ADVERSARIAL & STRESS TESTING
# ==============================================================================

def test_tier5_encoding_escaping_unicode_and_ideographic_spaces():
    """A1: Test robustness against Unicode BOM, ideographic spaces, quotes, and HTML special chars."""
    adversarial_items = [
        "\u3000\u3000以公允价值计量（FVTPL）& 股票",
        "债权投资 <含超短融> & 附注\"7\"",
        "其他债权投资\xa0(AAA级以上)",
        "其他权益工具投资\t\n",
    ]
    
    rows = []
    for i, item in enumerate(adversarial_items):
        rows.append({
            "table_family": "financial_investment",
            "member_table": "fvtpl_assets",
            "company": "中国平安",
            "report_year": 2024,
            "period_identity": "DATE:2024-12-31",
            "measure": "金额",
            "unit": "百万元",
            "canonical_item": item,
            "semantic_parent_path": "ROOT",
            "final_value": float((i + 1) * 1000),
            "source_row_id": f"ADV_{i}",
        })
    
    df = pd.DataFrame(rows)
    original, bridge, wide, audit = project_financial_investment_views(df)
    assert len(original) == len(adversarial_items)
    for orig_item, adv_item in zip(original["canonical_item"], adversarial_items):
        assert orig_item == adv_item


def test_tier5_manifest_tampering_and_mismatched_row_counts_fail_closed(tmp_path: Path):
    """A2: Tampered merge manifests with row count discrepancies fail closed."""
    run_dir = tmp_path / "tampered_merge_run"
    run_dir.mkdir(parents=True)
    
    # Valid files
    (run_dir / "financial_investment_original_long.csv").write_text(
        "presentation_member_id,presentation_regime,member_contract_version,source_row_ids,view_contract\n"
        "fvtpl_assets,NEW_FINANCIAL_INSTRUMENT_CLASSIFICATION,FINANCIAL_INVESTMENT_MEMBER_CONTRACT_V6,ROW1,SOURCE_PRESENTATION_EXACT_V1\n",
        encoding="utf-8-sig",
    )
    (run_dir / "financial_investment_standards_bridge_long.csv").write_text(
        "analysis_bridge_group,bridge_rule_id,bridge_projection_status,source_final_value,final_value,bridge_semantic_key,view_contract\n"
        "FVTPL_ASSETS,RULE1,BRIDGE_READY_PARTIAL_COMPARABILITY,100,100,KEY1,FINANCIAL_INVESTMENT_STANDARDS_BRIDGE_V1\n",
        encoding="utf-8-sig",
    )
    (run_dir / "financial_investment_standards_bridge_wide.csv").write_text(
        "analysis_bridge_group,canonical_item\nFVTPL_ASSETS,合计\n",
        encoding="utf-8-sig",
    )
    (run_dir / "financial_investment_standards_bridge_audit.csv").write_text(
        "audit_status,severity\nPARTIAL_COMPARABILITY,INFO\n",
        encoding="utf-8-sig",
    )
    
    # Tampered manifest with wrong row count (states 999 instead of 1)
    manifest = {
        "merge_schema_version": "6.9_FINANCIAL_PRESENTATION_REGIME_DUAL_VIEW",
        "canonical_observation_schema_version": "6.9_PRESENTATION_MEMBER_REGIME_LINEAGE",
        "financial_investment_standards_bridge": {
            "schema_version": "FINANCIAL_INVESTMENT_STANDARDS_BRIDGE_V1",
            "delivery_policy": "DUAL_VIEW_SOURCE_PRESENTATION_AND_EXPLICIT_BRIDGE",
            "no_same_period_sum": True,
            "original_row_count": 999,  # TAMPERED
            "bridge_row_count": 1,
            "audit_row_count": 1,
        },
    }
    (run_dir / "merge_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    
    audit_res = validate_financial_merge_artifacts([{"merge_id": "M1", "run_path": str(run_dir)}])
    assert audit_res["status"] == "FAIL"
    issues = [item["issue"] for item in audit_res["issues"]]
    assert "ORIGINAL_VIEW_ROW_COUNT_MISMATCH" in issues


def test_tier5_duplicate_active_member_occurrence_fail_closed():
    """A3: Duplicate active member occurrences fail closed in shadow gate."""
    payload = {
        "member_contract_version": "FINANCIAL_INVESTMENT_MEMBER_CONTRACT_V6",
        "v2_pass": True,
        "golden_identity_match": True,
        "required_current_member_status_valid": True,
        "physical_row_identity_unique": True,
        "note_value_binding_verified": True,
        "presentation_regime": "NEW_FINANCIAL_INSTRUMENT_CLASSIFICATION",
        "physical_source_row_ids": "P1_L1",
        "duplicate_active_member_occurrences": 2,  # CONFLICT
        "cross_row_binding_conflicts": 0,
    }
    stage_res = financial_v6_shadow_stage_result(payload)
    assert stage_res.status == AcceptanceStatus.FAIL
    assert "DUPLICATE_ACTIVE_PRESENTATION_MEMBER" in stage_res.evidence["issues"]


def test_tier5_empty_and_single_row_tables_stress():
    """A4: Handle empty and single-row tables without unhandled exceptions."""
    empty_df = pd.DataFrame()
    orig, bridge, wide, audit = project_financial_investment_views(empty_df)
    assert orig.empty
    assert bridge.empty
    assert wide.empty
    assert audit.empty

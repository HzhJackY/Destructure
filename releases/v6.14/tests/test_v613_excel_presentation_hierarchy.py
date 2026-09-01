"""v6.13 Excel presentation wide sheet visual hierarchy and outline grouping contracts."""
from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from table_merge import (  # noqa: E402
    RESEARCH_WIDE_FIXED_COLUMNS,
    build_research_wide_frame,
    write_merge_outputs,
    write_presentation_wide_sheet,
)
from visible_header_policy import VisibleHeaderDimensionPolicy  # noqa: E402


def _make_column_dimensions() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "column_id": "COL_00001",
            "source_column_label": "company=中国平安 | report_year=2023 | period_label=2023 | statement_scope=CONSOLIDATED | period_type=ANNUAL | currency=CNY | currency_unit=CNY_MILLION | measure=账面价值",
            "company": "中国平安",
            "report_year": "2023",
            "period_label": "2023",
            "statement_scope": "CONSOLIDATED",
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "measure": "账面价值",
            "display_company": "中国平安",
            "display_report_year": "2023",
            "display_period_label": "2023",
            "display_statement_scope": "合并",
            "display_period_type": "年报",
            "display_currency": "人民币",
            "display_currency_unit": "百万元",
            "display_measure": "账面价值",
        },
        {
            "column_id": "COL_00002",
            "source_column_label": "company=中国平安 | report_year=2024 | period_label=2024 | statement_scope=CONSOLIDATED | period_type=ANNUAL | currency=CNY | currency_unit=CNY_MILLION | measure=账面价值",
            "company": "中国平安",
            "report_year": "2024",
            "period_label": "2024",
            "statement_scope": "CONSOLIDATED",
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "measure": "账面价值",
            "display_company": "中国平安",
            "display_report_year": "2024",
            "display_period_label": "2024",
            "display_statement_scope": "合并",
            "display_period_type": "年报",
            "display_currency": "人民币",
            "display_currency_unit": "百万元",
            "display_measure": "账面价值",
        },
    ])


def test_write_presentation_wide_sheet_hierarchy_and_outline():
    """Verify outline grouping, indentation, parent styling, and grand total."""
    column_dimensions = _make_column_dimensions()
    header_policy = VisibleHeaderDimensionPolicy.from_column_dimensions(column_dimensions)

    research_wide_df = pd.DataFrame([
        # Row 0: Level 0 Parent (has children)
        {"member_table": "资产组合", "canonical_item": "固定到期日金融资产", "COL_00001": 4138694, "COL_00002": 4500000},
        # Row 1: Level 1 Leaf
        {"member_table": "资产组合", "canonical_item": "定期存款", "COL_00001": 404131, "COL_00002": 410000},
        # Row 2: Level 1 Parent (has Level 2 children)
        {"member_table": "资产组合", "canonical_item": "债券", "COL_00001": 2926986, "COL_00002": 3100000},
        # Row 3: Level 2 Leaf
        {"member_table": "资产组合", "canonical_item": "政府债券", "COL_00001": 1500000, "COL_00002": 1600000},
        # Row 4: Level 2 Leaf
        {"member_table": "资产组合", "canonical_item": "企业债", "COL_00001": 1426986, "COL_00002": 1500000},
        # Row 5: Deep Level Leaf (Level 6 -> should cap indent at 4)
        {"member_table": "资产组合", "canonical_item": "超深层级明细", "COL_00001": 10000, "COL_00002": 12000},
        # Row 6: Grand Total
        {"member_table": "资产组合", "canonical_item": "合计", "COL_00001": 5673371, "COL_00002": 6000000},
    ])

    fixed_columns = ["member_table", "canonical_item"]
    row_types = ["DETAIL", "DETAIL", "DETAIL", "DETAIL", "DETAIL", "DETAIL", "TOTAL"]
    row_levels = [0, 1, 1, 2, 2, 6, 0]

    output_buf = io.BytesIO()
    with pd.ExcelWriter(output_buf, engine="openpyxl") as writer:
        write_presentation_wide_sheet(
            writer,
            research_wide_df,
            fixed_columns,
            column_dimensions,
            header_policy,
            sheet_name="test_sheet",
            row_types=row_types,
            row_levels=row_levels,
        )

    output_buf.seek(0)
    wb = load_workbook(output_buf)
    ws = wb["test_sheet"]

    # 1. Verify Outline Properties
    assert ws.sheet_properties.outlinePr.summaryBelow is False
    assert ws.sheet_view.showOutlineSymbols is True

    # 2. Find data start row (header rows + 2 + 1)
    header_rows = max(1, len(header_policy.visible_header_dimensions))
    data_start_row = header_rows + 3

    # Row 0 (固定到期日金融资产): Level 0 Parent
    r0 = data_start_row + 0
    assert ws.row_dimensions[r0].outlineLevel == 0
    c0_item = ws.cell(row=r0, column=2)  # canonical_item
    assert c0_item.alignment.indent == 0
    assert c0_item.font.bold is True
    # Level 0 parent has F2F4F8 fill across the row
    assert c0_item.fill.fgColor.rgb in ("00F2F4F8", "F2F4F8")
    assert ws.cell(row=r0, column=3).fill.fgColor.rgb in ("00F2F4F8", "F2F4F8")
    assert ws.cell(row=r0, column=3).font.bold is True

    # Row 1 (定期存款): Level 1 Leaf
    r1 = data_start_row + 1
    assert ws.row_dimensions[r1].outlineLevel == 1
    c1_item = ws.cell(row=r1, column=2)
    assert c1_item.alignment.indent == 1
    assert c1_item.font.bold is False
    assert c1_item.fill.fill_type is None

    # Row 2 (债券): Level 1 Parent (has children at Level 2)
    r2 = data_start_row + 2
    assert ws.row_dimensions[r2].outlineLevel == 1
    c2_item = ws.cell(row=r2, column=2)
    assert c2_item.alignment.indent == 1
    assert c2_item.font.bold is True
    # Level 1 parent has white/no fill
    assert c2_item.fill.fill_type is None

    # Row 3 (政府债券): Level 2 Leaf
    r3 = data_start_row + 3
    assert ws.row_dimensions[r3].outlineLevel == 2
    c3_item = ws.cell(row=r3, column=2)
    assert c3_item.alignment.indent == 2
    assert c3_item.font.bold is False
    assert c3_item.fill.fill_type is None

    # Row 5 (超深层级明细): Level 6 -> indent capped at 4
    r5 = data_start_row + 5
    assert ws.row_dimensions[r5].outlineLevel == 6
    c5_item = ws.cell(row=r5, column=2)
    assert c5_item.alignment.indent == 4
    assert c5_item.font.bold is False

    # Row 6 (合计): Grand Total
    r6 = data_start_row + 6
    assert ws.row_dimensions[r6].outlineLevel == 0
    c6_item = ws.cell(row=r6, column=2)
    assert c6_item.alignment.indent == 0
    assert c6_item.font.bold is True
    assert c6_item.fill.fgColor.rgb in ("00FFF2CC", "FFF2CC")
    # Grand total has thin top border and double bottom border
    assert c6_item.border.top.style == "thin"
    assert c6_item.border.bottom.style == "double"
    # Also applied to numeric cells
    c6_val = ws.cell(row=r6, column=3)
    assert c6_val.font.bold is True
    assert c6_val.fill.fgColor.rgb in ("00FFF2CC", "FFF2CC")
    assert c6_val.border.top.style == "thin"
    assert c6_val.border.bottom.style == "double"


def test_write_merge_outputs_integration(tmp_path: Path):
    """Verify write_merge_outputs correctly propagates row_levels to both Excel workbooks."""
    output_dir = tmp_path / "merge_output"

    raw_long = pd.DataFrame([
        {
            "capture_run_id": "CAP_1",
            "member_table_order": 1,
            "row_order": 1,
            "company": "中国平安",
            "document_year": "2023",
            "table_id": "financial_investment",
            "table_family": "金融投资",
            "member_table": "debt_investment",
            "member_table_role": "NOTE_DETAIL",
            "source_table_title": "债权投资",
            "note_reference": "附注7",
            "source_pdf": "pingan.pdf",
            "container_id": "",
            "table_block_id": "",
            "block_order": -1,
            "classification_axis": "UNRESOLVED",
            "block_role": "UNRESOLVED",
            "block_terminal_type": "UNRESOLVED",
            "parent_section": "金融投资",
            "normalized_item": "固定到期日金融资产",
            "raw_item": "固定到期日金融资产",
            "row_type": "DETAIL",
            "row_level": 0,
            "hierarchy_level": 0,
            "page": 195,
            "row_path": "金融投资 / 固定到期日金融资产",
            "source_key": "固定到期日金融资产",
            "report_year": "2023",
            "data_year": "2023",
            "year": "2023",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "unit": "百万元",
            "measure": "",
            "value": 100.0,
        },
        {
            "capture_run_id": "CAP_1",
            "member_table_order": 1,
            "row_order": 2,
            "company": "中国平安",
            "document_year": "2023",
            "table_id": "financial_investment",
            "table_family": "金融投资",
            "member_table": "debt_investment",
            "member_table_role": "NOTE_DETAIL",
            "source_table_title": "债权投资",
            "note_reference": "附注7",
            "source_pdf": "pingan.pdf",
            "container_id": "",
            "table_block_id": "",
            "block_order": -1,
            "classification_axis": "UNRESOLVED",
            "block_role": "UNRESOLVED",
            "block_terminal_type": "UNRESOLVED",
            "parent_section": "固定到期日金融资产",
            "normalized_item": "定期存款",
            "raw_item": "定期存款",
            "row_type": "DETAIL",
            "row_level": 1,
            "hierarchy_level": 1,
            "page": 195,
            "row_path": "金融投资 / 固定到期日金融资产 / 定期存款",
            "source_key": "定期存款",
            "report_year": "2023",
            "data_year": "2023",
            "year": "2023",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "unit": "百万元",
            "measure": "",
            "value": 40.0,
        },
        {
            "capture_run_id": "CAP_1",
            "member_table_order": 1,
            "row_order": 3,
            "company": "中国平安",
            "document_year": "2023",
            "table_id": "financial_investment",
            "table_family": "金融投资",
            "member_table": "debt_investment",
            "member_table_role": "NOTE_DETAIL",
            "source_table_title": "债权投资",
            "note_reference": "附注7",
            "source_pdf": "pingan.pdf",
            "container_id": "",
            "table_block_id": "",
            "block_order": -1,
            "classification_axis": "UNRESOLVED",
            "block_role": "UNRESOLVED",
            "block_terminal_type": "UNRESOLVED",
            "parent_section": "金融投资",
            "normalized_item": "合计",
            "raw_item": "合计",
            "row_type": "TOTAL",
            "row_level": 0,
            "hierarchy_level": 0,
            "page": 195,
            "row_path": "金融投资 / 合计",
            "source_key": "合计",
            "report_year": "2023",
            "data_year": "2023",
            "year": "2023",
            "statement_scope": "CONSOLIDATED",
            "restated_flag": False,
            "period_type": "ANNUAL",
            "currency": "CNY",
            "currency_unit": "CNY_MILLION",
            "unit": "百万元",
            "measure": "",
            "value": 140.0,
        },
    ])

    mapping_queue = pd.DataFrame(columns=[
        "source_key", "canonical_section", "canonical_item",
        "category", "mapping_status", "mapping_note",
    ])
    manifest = {
        "version": "v6.13",
        "table_id": "financial_investment",
        "sources": [
            {"capture_run_id": "CAP_1", "member_table_order": 1},
        ],
        "reference_capture_run_id": "CAP_1",
    }

    paths = write_merge_outputs(output_dir, manifest, raw_long, mapping_queue)

    # Check research_wide.xlsx
    assert Path(paths["research_wide_xlsx"]).exists()
    wb = load_workbook(paths["research_wide_xlsx"])
    ws = wb["research_wide"]

    assert ws.sheet_properties.outlinePr.summaryBelow is False
    assert ws.sheet_view.showOutlineSymbols is True

    # Find row with "定期存款"
    found_child = False
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val == "定期存款":
            found_child = True
            assert ws.row_dimensions[r].outlineLevel == 1
            assert ws.cell(row=r, column=2).alignment.indent == 1
        elif val == "合计":
            assert ws.cell(row=r, column=2).font.bold is True
            assert ws.cell(row=r, column=2).fill.fgColor.rgb in ("00FFF2CC", "FFF2CC")

    assert found_child

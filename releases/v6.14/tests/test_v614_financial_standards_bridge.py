import json

import pandas as pd
from openpyxl import load_workbook

from financial_investment_standards_bridge import (
    annotate_financial_investment_identity,
    project_financial_investment_views,
)
from table_merge import assign_semantic_row_keys, write_merge_outputs


def _row(member, period, value, *, report_year=2023, source_row_id=None, **extra):
    return {
        "table_family": "financial_investment",
        "member_table": member,
        "member_table_role": "COMPONENT",
        "company": "中国太保",
        "report_year": report_year,
        "period_identity": period,
        "period_label": period,
        "statement_scope": "CONSOLIDATED",
        "measure": "金额",
        "unit": "百万元",
        "canonical_item": "合计",
        "semantic_parent_path": "ROOT",
        "final_value": value,
        "source_row_id": source_row_id or f"ROW_{member}_{period}",
        **extra,
    }


def test_annotation_preserves_presentation_identity_and_multi_bridge_membership():
    annotated = annotate_financial_investment_identity(pd.DataFrame([
        _row("available_for_sale_assets", "DATE:2022-12-31", 100),
        _row("other_debt_investment", "DATE:2023-12-31", 60),
    ]))
    assert annotated.loc[0, "presentation_member_id"] == "available_for_sale_assets"
    assert annotated.loc[0, "presentation_regime"] == "LEGACY_FINANCIAL_ASSET_CLASSIFICATION"
    assert "FVOCI_DEBT" in annotated.loc[0, "analysis_bridge_groups"]
    assert "FVOCI_EQUITY" in annotated.loc[0, "analysis_bridge_groups"]
    assert annotated.loc[1, "presentation_regime"] == "NEW_FINANCIAL_INSTRUMENT_CLASSIFICATION"


def test_fvtpl_bridge_keeps_original_rows_and_projects_periods_without_sum():
    source = pd.DataFrame([
        _row("legacy_fvtpl_assets", "DATE:2022-12-31", 26560, report_year=2023),
        _row("fvtpl_assets", "DATE:2023-12-31", 581602, report_year=2023),
    ])
    original, bridge, wide, audit = project_financial_investment_views(source)
    assert set(original["presentation_member_id"]) == {"legacy_fvtpl_assets", "fvtpl_assets"}
    assert set(bridge["analysis_bridge_group"]) == {"FVTPL_ASSETS"}
    assert sorted(bridge["final_value"].tolist()) == [26560, 581602]
    assert set(bridge["bridge_projection_status"]) == {"BRIDGE_READY_PARTIAL_COMPARABILITY"}
    assert not wide.empty
    assert set(audit["audit_status"]) == {"PARTIAL_COMPARABILITY"}


def test_same_period_new_and_legacy_sources_fail_closed_without_aggregation():
    source = pd.DataFrame([
        _row("legacy_fvtpl_assets", "DATE:2023-12-31", 100, source_row_id="OLD"),
        _row("fvtpl_assets", "DATE:2023-12-31", 200, source_row_id="NEW"),
    ])
    _, bridge, _, audit = project_financial_investment_views(source)
    assert bridge["final_value"].isna().all()
    assert set(bridge["bridge_projection_status"]) == {"BRIDGE_AMBIGUOUS_SOURCE_SET"}
    assert "BRIDGE_AMBIGUOUS_SOURCE_SET" in set(audit["audit_status"])


def test_available_for_sale_requires_certified_disaggregation_for_both_targets():
    source = pd.DataFrame([
        _row("available_for_sale_assets", "DATE:2022-12-31", 300),
    ])
    original, bridge, _, audit = project_financial_investment_views(source)
    assert len(original) == 1
    assert set(bridge["analysis_bridge_group"]) == {"FVOCI_DEBT", "FVOCI_EQUITY"}
    assert bridge["final_value"].isna().all()
    assert set(bridge["bridge_projection_status"]) == {"BLOCKED_DISAGGREGATION_REQUIRED"}
    assert len(audit[audit["audit_status"] == "BLOCKED_DISAGGREGATION_REQUIRED"]) == 2


def test_member_without_bridge_keeps_stable_empty_bridge_schema():
    source = pd.DataFrame([
        _row("time_deposits", "DATE:2023-12-31", 100),
    ])
    original, bridge, wide, audit = project_financial_investment_views(source)

    assert len(original) == 1
    assert bridge.empty
    assert {
        "analysis_bridge_group", "bridge_rule_id", "bridge_projection_status",
        "source_final_value", "final_value", "bridge_semantic_key", "view_contract",
    } <= set(bridge.columns)
    assert "analysis_bridge_group" in wide.columns
    assert audit["audit_status"].tolist() == ["NO_STANDARDS_BRIDGE"]


def test_same_label_on_different_classification_axes_does_not_conflict():
    source = pd.DataFrame([
        _row(
            "fvtpl_assets", "DATE:2023-12-31", 100,
            source_row_id="LISTING_TOTAL", classification_axis="LISTING_STATUS",
            semantic_occurrence=1,
        ),
        _row(
            "fvtpl_assets", "DATE:2023-12-31", 200,
            source_row_id="ASSET_TOTAL", classification_axis="ASSET_TYPE",
            semantic_occurrence=1,
        ),
    ])
    _, bridge, wide, audit = project_financial_investment_views(source)
    assert sorted(bridge["final_value"].tolist()) == [100, 200]
    assert "BRIDGE_AMBIGUOUS_SOURCE_SET" not in set(bridge["bridge_projection_status"])
    assert len(wide) == 2
    assert "BRIDGE_WIDE_IDENTITY_CONFLICT" not in set(audit.get("audit_status", []))


def test_semantic_identity_includes_presentation_regime():
    raw = pd.DataFrame([
        {
            **_row("legacy_fvtpl_assets", "DATE:2022-12-31", 100),
            "capture_run_id": "CAP_OLD", "row_order": 1,
            "normalized_item": "合计", "classification_axis": "ASSET_TYPE",
            "schema_version": 16, "producer_version": "v6.12",
        },
        {
            **_row("fvtpl_assets", "DATE:2023-12-31", 200),
            "capture_run_id": "CAP_NEW", "row_order": 1,
            "normalized_item": "合计", "classification_axis": "ASSET_TYPE",
            "schema_version": 16, "producer_version": "v6.12",
        },
    ])
    projected = assign_semantic_row_keys(raw)
    keys = projected["semantic_row_key"].tolist()
    assert keys[0] != keys[1]
    assert "REGIME::LEGACY_FINANCIAL_ASSET_CLASSIFICATION" in keys[0]
    assert "REGIME::NEW_FINANCIAL_INSTRUMENT_CLASSIFICATION" in keys[1]


def _merge_row(member, period, value, capture_id):
    period_year = int(period[5:9])
    source_period_label = f"{period_year}年12月31日"
    return {
        "capture_run_id": capture_id,
        "member_table_order": 1,
        "row_order": 1,
        "company": "中国太保",
        "document_year": "2023",
        "table_id": "financial_investment",
        "table_family": "金融投资",
        "member_table": member,
        "member_table_role": "NOTE_DETAIL",
        "source_table_title": member,
        "note_reference": "附注10" if member == "fvtpl_assets" else "附注2",
        "source_pdf": "中国太保2023年报.pdf",
        "container_id": "",
        "table_block_id": "",
        "block_order": -1,
        "classification_axis": "ASSET_TYPE",
        "block_role": "PRIMARY",
        "block_terminal_type": "FINAL_TOTAL",
        "parent_section": "金融投资",
        "normalized_item": "合计",
        "raw_item": "合计",
        "page": 199,
        "row_path": "金融投资 / 合计",
        "source_key": f"{member}::{period}",
        "source_row_id": f"{capture_id}::ROW_1",
        "canonical_key": f"{member}::合计",
        "canonical_section": member,
        "canonical_item": "合计",
        "mapping_status": "AUTO_IDENTITY",
        "report_year": "2023",
        "data_year": str(period_year),
        "year": str(period_year),
        "period_identity": period,
        "source_period_label": source_period_label,
        "period_label": source_period_label,
        "period_year": period_year,
        "period_month": 12,
        "period_day": 31,
        "period_precision": "DAY",
        "period_date": period.removeprefix("DATE:"),
        "period_kind": "ABSOLUTE_DATE",
        "statement_scope": "CONSOLIDATED",
        "restated_flag": False,
        "period_type": "ANNUAL",
        "currency": "CNY",
        "currency_unit": "CNY_MILLION",
        "unit": "百万元",
        "measure": "金额",
        "value": value,
    }


def test_formal_merge_writes_dual_views_manifest_and_workbook_sheets(tmp_path):
    raw = pd.DataFrame([
        _merge_row("legacy_fvtpl_assets", "DATE:2022-12-31", 26560, "CAP_OLD"),
        _merge_row("fvtpl_assets", "DATE:2023-12-31", 581602, "CAP_NEW"),
        {
            **_merge_row(
                "available_for_sale_assets", "DATE:2022-12-31", 300,
                "CAP_CERTIFIED_SPLIT",
            ),
            "bridge_certification_status": "CERTIFIED_DISAGGREGATION",
            "certified_bridge_rule_id": "FI_BRIDGE_FVOCI_DEBT_V1",
        },
    ])
    mapping_queue = pd.DataFrame(columns=[
        "source_key", "canonical_section", "canonical_item", "category",
        "mapping_status", "mapping_note",
    ])
    output_dir = tmp_path / "merge"
    paths = write_merge_outputs(
        output_dir=output_dir,
        manifest={
            "version": "v6.14",
            "table_id": "financial_investment",
            "sources": [
                {"capture_run_id": "CAP_OLD", "member_table_order": 1},
                {"capture_run_id": "CAP_NEW", "member_table_order": 2},
                {"capture_run_id": "CAP_CERTIFIED_SPLIT", "member_table_order": 3},
            ],
            "reference_capture_run_id": "CAP_NEW",
        },
        raw_long=raw,
        mapping_queue=mapping_queue,
        taxonomy_path=None,
    )

    original = pd.read_csv(paths["financial_original"])
    bridge = pd.read_csv(paths["financial_bridge"])
    assert set(original["presentation_member_id"]) == {
        "legacy_fvtpl_assets", "fvtpl_assets", "available_for_sale_assets",
    }
    bridge_values = bridge.dropna(subset=["final_value"])
    assert sorted(bridge_values["final_value"].tolist()) == [300, 26560, 581602]
    certified = bridge[bridge["presentation_member_id"] == "available_for_sale_assets"]
    assert certified.loc[
        certified["analysis_bridge_group"] == "FVOCI_DEBT", "final_value"
    ].tolist() == [300]
    assert certified.loc[
        certified["analysis_bridge_group"] == "FVOCI_EQUITY", "final_value"
    ].isna().all()
    manifest = json.loads((output_dir / "merge_manifest.json").read_text(encoding="utf-8"))
    bridge_manifest = manifest["financial_investment_standards_bridge"]
    assert bridge_manifest["delivery_policy"] == "DUAL_VIEW_SOURCE_PRESENTATION_AND_EXPLICIT_BRIDGE"
    assert bridge_manifest["original_row_count"] == 3
    assert bridge_manifest["bridge_value_count"] == 3

    expected_sheets = {"金融投资_原始口径", "金融投资_跨准则桥接", "金融投资_桥接审计"}
    assert expected_sheets <= set(load_workbook(paths["xlsx"], read_only=True).sheetnames)
    assert expected_sheets <= set(load_workbook(paths["research_wide_xlsx"], read_only=True).sheetnames)

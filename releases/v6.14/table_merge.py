#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
table_merge.py — v5.1 multi-company / multi-year table merge engine.

Safety model:
- raw capture rows are immutable evidence.
- exact normalized identities may auto-align.
- non-identical labels are NEVER silently merged.
- fuzzy similarity is suggestion-only.
- confirmed mappings can be persisted to a table taxonomy.
- duplicate/conflicting canonical keys are surfaced, not silently overwritten.
"""

from __future__ import annotations

import difflib
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Optional, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from batch_pipeline import infer_company_year
from financial_metric_pdf_resolver import normalize_text
from financial_structure_resolver import project_certified_row_hierarchy
from period_identity import normalize_period_fields
from visible_header_policy import VisibleHeaderDimensionPolicy, OBSERVATION_DIMENSIONS
from financial_investment_standards_bridge import (
    BRIDGE_SCHEMA_VERSION,
    annotate_financial_investment_identity,
    project_financial_investment_views,
)


TAXONOMY_VERSION = 1

# A Family Merge observation is source-aware before it is row-aware.  These
# fields intentionally survive all long/canonical/wide materializations.
SOURCE_IDENTITY_COLUMNS = [
    "table_family", "member_table", "member_table_role",
    "presentation_member_id", "presentation_regime",
    "source_table_title", "note_reference", "capture_run_id", "source_pdf",
]
BLOCK_IDENTITY_COLUMNS = [
    "container_id", "table_block_id", "block_order",
    "classification_axis", "block_role", "block_terminal_type",
]
PHYSICAL_BLOCK_LINEAGE_COLUMNS = [
    "container_id", "table_block_id", "block_order",
    "block_role", "block_terminal_type",
]
CANONICAL_BLOCK_IDENTITY_COLUMNS = ["classification_axis"]
SOURCE_IDENTITY_COLUMNS += BLOCK_IDENTITY_COLUMNS


def _identity_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "<na>"} else text


def _source_identity_missing(row: Any) -> bool:
    return not all(_identity_text(row.get(field)) for field in (
        "table_family", "member_table", "member_table_role",
    ))


def _semantic_block_scope(row: Any) -> str:
    """Use a stable semantic axis across Captures; isolate unresolved blocks."""
    axis = _identity_text(row.get("classification_axis"))
    if axis and axis != "UNRESOLVED":
        return f"AXIS::{axis}||"
    block_id = _identity_text(row.get("table_block_id"))
    return f"BLOCK::{block_id}||AXIS::UNRESOLVED||" if block_id else ""


def normalize_table_id(text: str) -> str:
    s = normalize_text(str(text or ""))
    return s or "UNNAMED_TABLE"


def normalize_section(text: Any) -> str:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    return re.sub(r"\s+", "", str(text)).strip("：:")


def source_mapping_key(parent_section: Any, normalized_item: Any) -> str:
    """Legacy key retained for taxonomy compatibility."""
    return f"{normalize_section(parent_section)}||{str(normalized_item or '').strip()}"


def _normalize_semantic_path(value: Any) -> str:
    return "/".join(
        normalize_section(part)
        for part in str(value or "").split("/")
        if normalize_section(part)
    )


def _semantic_row_base(row: Any) -> str:
    family = _identity_text(row.get("table_family"))
    member = _identity_text(row.get("presentation_member_id")) or _identity_text(row.get("member_table"))
    regime = _identity_text(row.get("presentation_regime"))
    item = _identity_text(row.get("normalized_item"))
    parent_path = _normalize_semantic_path(row.get("hierarchy_parent_path"))
    axis_scope = _semantic_block_scope(row).rstrip("|")
    return (
        f"FAMILY::{family or 'UNRESOLVED'}||MEMBER::{member or 'UNRESOLVED'}||"
        f"REGIME::{regime or 'UNRESOLVED'}||"
        f"{axis_scope}||PARENT::{parent_path or 'ROOT'}||ITEM::{item or 'UNRESOLVED'}"
    )


def assign_semantic_row_keys(df: pd.DataFrame) -> pd.DataFrame:
    """Build cross-Capture row identities from the certified parent graph."""
    if df.empty:
        out = df.copy()
        out["semantic_row_key"] = ""
        out["source_key"] = ""
        return out

    working = annotate_financial_investment_identity(df)
    if "capture_run_id" not in working.columns:
        working["capture_run_id"] = "LEGACY_CAPTURE"
    working["_input_order"] = range(len(working))
    projections = []
    for _, source in working.groupby("capture_run_id", sort=False, dropna=False):
        capture_run = str(source["capture_run_id"].iloc[0] if not source.empty else "")
        schema_ver = pd.to_numeric(source["schema_version"], errors="coerce").dropna() if "schema_version" in source.columns else pd.Series(dtype="float64")
        producer_ver = source["producer_version"].dropna() if "producer_version" in source.columns else pd.Series(dtype="object")

        is_explicit_v613 = (
            (not schema_ver.empty and (schema_ver >= 17).any())
            or (not producer_ver.empty and any(str(p).startswith("v6.13") for p in producer_ver))
        )
        is_legacy = (
            not is_explicit_v613
            or capture_run.startswith("LEGACY")
        )
        try:
            proj = project_certified_row_hierarchy(
                source,
                allow_legacy_compatibility=is_legacy,
            )
        except ValueError as err:
            if any(token in str(err) for token in ("SOURCE_ROW_ID_REQUIRED", "SOURCE_ROW_ID_NOT_UNIQUE", "PARENT_ROW_ID_CYCLE")):
                proj = source.copy()
                for col in ("source_row_id", "parent_row_id", "hierarchy_parent_label", "hierarchy_parent_path"):
                    if col not in proj.columns:
                        proj[col] = None
                proj["hierarchy_path"] = proj["normalized_item"].fillna(proj.get("raw_item", "")).fillna("")
                proj["hierarchy_level"] = 0
                proj["hierarchy_status"] = "REVIEW_REQUIRED_SOURCE_IDENTITY"
            else:
                raise
        projections.append(proj)
    out = pd.concat(projections, ignore_index=True).sort_values("_input_order", kind="stable")
    out = out.drop(columns=["_input_order"]).reset_index(drop=True)
    for col in ("source_row_id", "parent_row_id", "hierarchy_parent_path", "hierarchy_status"):
        if col not in out.columns:
            out[col] = None
    out["semantic_parent_path"] = out["hierarchy_parent_path"].map(_normalize_semantic_path)
    out["semantic_row_base"] = out.apply(_semantic_row_base, axis=1)
    out["semantic_occurrence"] = 1

    structural = (
        out[[
            "capture_run_id", "source_row_id", "row_order", "semantic_row_base",
            "hierarchy_status",
        ]]
        .drop_duplicates(subset=["capture_run_id", "source_row_id"], keep="first")
        .copy()
    )
    structural["_row_order_num"] = pd.to_numeric(structural["row_order"], errors="coerce")
    structural = structural.sort_values(
        ["capture_run_id", "_row_order_num"], kind="stable",
    )
    structural["semantic_occurrence"] = structural.groupby(
        ["capture_run_id", "semantic_row_base"], sort=False,
    ).cumcount() + 1
    structural["_same_base_count"] = structural.groupby(
        ["capture_run_id", "semantic_row_base"], sort=False,
    )["source_row_id"].transform("size")
    out = out.merge(
        structural[[
            "capture_run_id", "source_row_id", "semantic_occurrence", "_same_base_count",
        ]],
        on=["capture_run_id", "source_row_id"],
        how="left",
        suffixes=("", "_derived"),
    )
    out["semantic_occurrence"] = out["semantic_occurrence_derived"].fillna(1).astype(int)
    out = out.drop(columns=["semantic_occurrence_derived"])

    safe_statuses = {"CERTIFIED_ROOT", "CERTIFIED_PARENT_GRAPH"}
    out["semantic_identity_status"] = out["hierarchy_status"].map(
        lambda status: (
            "CERTIFIED"
            if _identity_text(status) in safe_statuses
            else _identity_text(status) or "SEMANTIC_IDENTITY_UNRESOLVED"
        )
    )

    def key_for(row: Any) -> str:
        base = str(row.get("semantic_row_base") or "")
        if int(row.get("_same_base_count") or 1) > 1:
            base += f"||OCC::{int(row.get('semantic_occurrence') or 1)}"
        if row.get("semantic_identity_status") not in {"CERTIFIED", "LEGACY_IDENTITY_COMPATIBILITY"}:
            base += (
                f"||UNRESOLVED_SOURCE::{_identity_text(row.get('capture_run_id'))}"
                f"::{_identity_text(row.get('source_row_id'))}"
            )
        return base

    out["semantic_row_key"] = out.apply(key_for, axis=1)
    # Existing taxonomy and mapping APIs keep their column name, but consume
    # the single semantic identity implementation.
    out["source_key"] = out["semantic_row_key"]
    return out.drop(columns=["_same_base_count"])


def assign_conditional_source_keys(df: pd.DataFrame) -> pd.DataFrame:
    """Backward-compatible entry point for the v6.13 semantic row key."""
    return assign_semantic_row_keys(df)


def _atomic_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)


def load_taxonomy(path: Path) -> dict[str, Any]:
    path = Path(path)
    if not path.exists():
        return {"version": TAXONOMY_VERSION, "tables": {}}
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("version", TAXONOMY_VERSION)
    data.setdefault("tables", {})
    return data


def _taxonomy_lookup(taxonomy: dict[str, Any], table_id: str) -> dict[str, dict[str, Any]]:
    table = taxonomy.get("tables", {}).get(table_id, {})
    mappings = table.get("mappings", [])
    lookup: dict[str, dict[str, Any]] = {}

    # Exact stored keys.
    for m in mappings:
        for alias_key in m.get("source_keys", []):
            lookup[str(alias_key)] = m

    # Backward compatibility: old versions stored parent_section||item. For a
    # v5.6 UNIQUE||item key, reuse an old mapping only if all legacy mappings for
    # that item point to the same canonical target.
    by_item: dict[str, list[dict[str, Any]]] = {}
    for m in mappings:
        for alias_key in m.get("source_keys", []):
            key = str(alias_key)
            if key.startswith("UNIQUE||"):
                item = key.split("||",1)[1]
            elif "||" in key and not key.startswith("CONTEXT||"):
                item = key.split("||")[-1]
            else:
                continue
            by_item.setdefault(item, []).append(m)

    for item, candidates in by_item.items():
        targets = {
            (
                str(m.get("canonical_section") or ""),
                str(m.get("canonical_item") or ""),
            )
            for m in candidates
        }
        if len(targets) == 1:
            lookup.setdefault(f"UNIQUE||{item}", candidates[0])

    return lookup

def infer_capture_metadata(capture_dir: Path) -> dict[str, Any]:
    result_path = capture_dir / "table_capture_result.json"
    data = json.loads(result_path.read_text(encoding="utf-8"))
    pdf_name = data.get("pdf_name", "")
    company, filename_year = infer_company_year(Path(pdf_name), "")

    # v5.8: report/document_year must be an absolute four-digit year.
    # Relative column labels such as 本年累计数 / 上年累计数 / 去年累计数
    # are never valid document identity.
    absolute_candidates = []
    if re.fullmatch(r"20\d{2}", str(filename_year or "").strip()):
        absolute_candidates.append(str(filename_year).strip())

    for c in data.get("columns", []) or []:
        for field in ["year", "period_label", "header_raw"]:
            value = str(c.get(field) or "").strip()
            m = re.search(r"(20\d{2})", value)
            if m:
                absolute_candidates.append(m.group(1))

    document_year = max(absolute_candidates) if absolute_candidates else ""
    capture_meta_path = capture_dir / "capture_metadata.json"
    try:
        capture_meta = json.loads(capture_meta_path.read_text(encoding="utf-8")) if capture_meta_path.exists() else {}
    except (OSError, json.JSONDecodeError):
        capture_meta = {}
    stats = data.get("stats") or {}
    return {
        "capture_run_id": capture_dir.name,
        "capture_dir": str(capture_dir),
        "pdf_name": pdf_name,
        "pdf_sha256": data.get("pdf_sha256"),
        "company": company,
        "document_year": document_year,
        "table_query": data.get("table_query", ""),
        "note_number": data.get("note_number"),
        "located_title": data.get("located_title", ""),
        "table_family": capture_meta.get("table_family"),
        "member_table": capture_meta.get("member_table"),
        "member_table_role": capture_meta.get("member_table_role"),
        "source_table_title": capture_meta.get("source_table_title"),
        "note_reference": capture_meta.get("note_reference"),
        "source_pdf": capture_meta.get("source_pdf_path") or stats.get("source_pdf_path") or pdf_name,
        "member_table_order": capture_meta.get("member_table_order"),
    }


CURRENT_RELATIVE_PERIOD_TOKENS = {
    "本年累计数", "本年度累计数", "本期累计数", "本期数", "本年数",
    "本期", "本年", "本年度", "当期累计数", "当期",
    "期末", "年末", "本期期末", "本年末", "本年度末",
}
PRIOR_RELATIVE_PERIOD_TOKENS = {
    "上年累计数", "上年度累计数", "上期累计数", "上期数", "上年数",
    "上期", "上年", "上年度", "去年", "去年累计数", "去年数",
    "去年同期", "上年同期", "上年度同期", "上期同期",
}


def _is_missing_scalar(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _clean_scalar_text(value: Any) -> str:
    if _is_missing_scalar(value):
        return ""
    return str(value).strip()


def _clean_period_token(value: Any) -> str:
    token = _clean_scalar_text(value)
    if re.fullmatch(r"20\d{2}\.0", token):
        return token[:4]
    token = re.sub(r"\s+", "", token)
    token = token.replace("（", "(").replace("）", ")")
    token = re.sub(r"\(?(?:已重述|经重述|重述后|重述)\)?", "", token)
    token = re.sub(r"(?:人民币)?(?:亿元|百万元|万元|千元|元)$", "", token)
    return token.strip()


def _absolute_year(value: Any) -> Optional[str]:
    token = _clean_period_token(value)
    # CSV round trips commonly turn a year into `2022.0`.  It remains an
    # absolute reporting period, not a distinct observation dimension.
    m = re.fullmatch(r"(20\d{2})(?:\.0)?(?:年度|年)?", token)
    return m.group(1) if m else None


PERIOD_COMPONENT_COLUMNS = (
    "source_period_label", "period_label", "period_year", "period_month",
    "period_day", "period_precision", "period_date", "period_identity",
    "period_kind", "period_normalization_evidence",
)


def _ensure_period_identity_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Derive point-period identity without rewriting source Capture evidence."""
    out = df.copy()
    if out.empty:
        for column in PERIOD_COMPONENT_COLUMNS:
            if column not in out.columns:
                out[column] = pd.Series(dtype="object")
        return out

    rows: list[dict[str, Any]] = []
    for _, row in out.iterrows():
        normalized = normalize_period_fields(
            source_period_label=row.get("source_period_label"),
            period_label=row.get("period_label"),
            year=row.get("year") if not _dimension_missing(row.get("year")) else row.get("data_year"),
        )
        existing_identity = _identity_text(row.get("period_identity"))
        if existing_identity:
            parsed_identity = _identity_text(normalized.get("period_identity"))
            if parsed_identity and existing_identity != parsed_identity:
                raise ValueError(
                    "PERIOD_IDENTITY_EVIDENCE_MISMATCH:"
                    f"{existing_identity}!={parsed_identity}"
                )
            normalized["period_identity"] = existing_identity
            for column in PERIOD_COMPONENT_COLUMNS:
                existing = row.get(column)
                if not _dimension_missing(existing):
                    normalized[column] = existing
        has_observation_value = any(
            not _dimension_missing(row.get(column))
            for column in ("value", "value_raw", "value_numeric", "value_original")
            if column in row.index
        )
        if not normalized.get("period_identity") and has_observation_value:
            raise ValueError(
                "PERIOD_DATE_UNRESOLVED:"
                f"{row.get('capture_run_id', '')}:{row.get('period_label', row.get('year', ''))}"
            )
        rows.append(normalized)

    normalized_frame = pd.DataFrame(rows, index=out.index)
    for column in PERIOD_COMPONENT_COLUMNS:
        out[column] = normalized_frame[column]
    for column in ("period_year", "period_month", "period_day"):
        out[column] = pd.to_numeric(out[column], errors="coerce").astype("Int64")
    resolved_year = out["period_year"].astype("string")
    legacy_year = (
        out["data_year"].astype("string")
        if "data_year" in out.columns
        else out.get("year", pd.Series(index=out.index, dtype="string")).astype("string")
    )
    legacy_year = legacy_year.str.replace(r"^((?:19|20)\d{2})\.0$", r"\1", regex=True)
    out["data_year"] = resolved_year.where(resolved_year.notna(), legacy_year)
    if "year" in out.columns:
        out_year_cleaned = out["year"].astype("string").str.replace(r"^((?:19|20)\d{2})\.0$", r"\1", regex=True)
        out["year"] = resolved_year.where(
            resolved_year.notna(), out_year_cleaned
        )
    else:
        out["year"] = out["data_year"]
    if "period_label" in out.columns:
        out["period_label"] = out["period_label"].astype("string").str.replace(r"^((?:19|20)\d{2})\.0$", r"\1", regex=True)
    return out


def period_precision_audit(df: pd.DataFrame) -> pd.DataFrame:
    """Report same-year precision differences without blocking observations."""
    if df.empty:
        return pd.DataFrame()
    source = _ensure_period_identity_columns(df)
    source = source[source["period_year"].notna()].copy()
    if source.empty:
        return pd.DataFrame()
    candidate_keys = [
        "company", "report_year", "table_family", "member_table",
        "classification_axis", "canonical_item", "normalized_item",
        "statement_scope", "scope", "restated_flag", "restated", "measure",
        "period_year",
    ]
    keys = [column for column in candidate_keys if column in source.columns]
    rows: list[dict[str, Any]] = []
    for identity, group in source.groupby(keys, dropna=False, sort=False):
        precisions = sorted({
            str(value).strip()
            for value in group["period_precision"].dropna().tolist()
            if str(value).strip()
        })
        if "YEAR" not in precisions or len(precisions) < 2:
            continue
        base = {
            column: value
            for column, value in zip(
                keys, identity if isinstance(identity, tuple) else (identity,)
            )
        }
        rows.append({
            **base,
            "audit_code": "PERIOD_PRECISION_MISMATCH",
            "severity": "WARNING",
            "blocking": False,
            "precisions": "|".join(precisions),
            "period_identities": "|".join(sorted({
                str(value).strip()
                for value in group["period_identity"].dropna().tolist()
                if str(value).strip()
            })),
            "detail": "同一经济项目同时存在年精度和更高精度期间；分别保留，不自动视为等价。",
        })
    return pd.DataFrame(rows)


def _relative_period_kind(value: Any) -> Optional[str]:
    token = _clean_period_token(value)
    if token in CURRENT_RELATIVE_PERIOD_TOKENS:
        return "CURRENT"
    if token in PRIOR_RELATIVE_PERIOD_TOKENS:
        return "PRIOR"
    return None


def _infer_document_year_from_capture_source(
    *,
    pdf_name: Any = "",
    capture_dir: Any = "",
    columns: Optional[list[dict[str, Any]]] = None,
) -> str:
    """Infer report year from absolute evidence only."""
    _, filename_year = infer_company_year(Path(str(pdf_name or "")), "")
    y = _absolute_year(filename_year)
    if y:
        return y

    absolute_candidates: list[str] = []
    for c in columns or []:
        for field in ["year", "period_label", "header_raw"]:
            y = _absolute_year(c.get(field))
            if y:
                absolute_candidates.append(y)
    if absolute_candidates:
        return max(absolute_candidates)

    cap = Path(str(capture_dir or ""))
    result_path = cap / "table_capture_result.json"
    if result_path.exists():
        try:
            data = json.loads(result_path.read_text(encoding="utf-8"))
            _, filename_year = infer_company_year(Path(str(data.get("pdf_name") or "")), "")
            y = _absolute_year(filename_year)
            if y:
                return y
            for c in data.get("columns") or []:
                for field in ["year", "period_label", "header_raw"]:
                    y = _absolute_year(c.get(field))
                    if y:
                        absolute_candidates.append(y)
            if absolute_candidates:
                return max(absolute_candidates)
        except Exception:
            pass
    return ""


def _resolve_source_document_year(source: dict[str, Any]) -> str:
    existing = _absolute_year(source.get("document_year"))
    if existing:
        return existing
    return _infer_document_year_from_capture_source(
        pdf_name=source.get("pdf_name"),
        capture_dir=source.get("capture_dir"),
    )


def _resolve_relative_period_years(
    df: pd.DataFrame,
    document_year: Any,
) -> pd.DataFrame:
    """
    Convert relative period labels to actual years before canonical merge.

    document_year=2023:
      本年 / 本年累计数 / 本期累计数 -> 2023
      去年 / 去年累计数 / 上年累计数 / 上期累计数 -> 2022

    Original wording is retained in `source_period_label`.
    """
    out = df.copy()
    if "year" not in out.columns:
        return out

    doc = _absolute_year(document_year)

    if "source_period_label" not in out.columns:
        if "period_label" in out.columns:
            out["source_period_label"] = out["period_label"]
        else:
            out["source_period_label"] = out["year"]

    def resolve(value: Any) -> Any:
        absolute = _absolute_year(value)
        if absolute:
            return absolute
        kind = _relative_period_kind(value)
        if kind is None or doc is None:
            return value
        return doc if kind == "CURRENT" else str(int(doc) - 1)

    out["year"] = out["year"].map(resolve)

    source_labels = out["source_period_label"].copy()
    for index, source_label in source_labels.items():
        if _relative_period_kind(source_label) and _absolute_year(out.at[index, "year"]):
            report_type = str(out.at[index, "period_type"] if "period_type" in out else "").upper()
            if report_type in {"QUARTERLY", "SEMIANNUAL"}:
                raise ValueError(
                    "PERIOD_DATE_UNRESOLVED:RELATIVE_"
                    f"{report_type}:{source_label}:REPORT_PERIOD_END_DATE_REQUIRED"
                )
            out.at[index, "period_label"] = str(out.at[index, "year"])
            for column in PERIOD_COMPONENT_COLUMNS:
                if column not in {"source_period_label", "period_label"}:
                    out.at[index, column] = None

    out = _ensure_period_identity_columns(out)

    if "column_dimension_key" in out.columns:
        def dimension_key(r: pd.Series) -> Any:
            if _is_missing_scalar(r.get("column_ordinal")):
                return r.get("column_dimension_key")
            restated_raw = r.get("restated")
            restated = False if _is_missing_scalar(restated_raw) else bool(restated_raw)
            return (
                f"{_clean_scalar_text(r.get('period_identity'))}|"
                f"{_clean_scalar_text(r.get('scope'))}|"
                f"{'RESTATED' if restated else 'ORIGINAL'}|"
                f"{_clean_scalar_text(r.get('measure'))}"
            )

        out["column_dimension_key"] = out.apply(dimension_key, axis=1)
    return out


def _validate_absolute_year_resolution(
    df: pd.DataFrame,
    *,
    document_year: Any,
    capture_run_id: str,
) -> None:
    if "year" not in df.columns:
        return

    relative_values = sorted({
        str(v)
        for v in df["year"].dropna().tolist()
        if _relative_period_kind(v)
    })
    if not relative_values:
        return

    doc = _absolute_year(document_year)
    if not doc:
        raise ValueError(
            f"PERIOD_RESOLUTION_REQUIRED：来源 {capture_run_id} 包含相对期间 "
            f"{relative_values}，但报告年份 document_year 未识别为四位年份。"
            " 请在合表来源信息中填写该年报实际年份，例如 2023。"
            " 系统随后会自动将本年→2023、去年/上年→2022。"
        )

    raise ValueError(
        f"PERIOD_RESOLUTION_INVARIANT_VIOLATION：来源 {capture_run_id} "
        f"在 document_year={doc} 已知时仍残留相对 year={relative_values}。"
        " 已阻止其进入 canonical 合表。"
    )


def _repair_manifest_and_raw_periods(
    raw_long: pd.DataFrame,
    manifest: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """
    Repair existing v5.7 Merge Projects on rematerialization as well as new ones.
    """
    raw = raw_long.copy()
    # Legacy CSV projects commonly deserialize period identity columns as
    # int64.  v6.7 deliberately normalizes report/document years to canonical
    # strings so they can coexist with relative labels before resolution.  On
    # pandas 3, assigning "2023" into an int64 block is a hard TypeError.
    # Coerce only identity/presentation fields; financial value columns remain
    # untouched and no source amount is rewritten.
    for period_column in (
        "document_year", "report_year", "data_year", "year",
        "source_period_label", "period_label", "period_identity",
        "period_precision", "period_date", "period_kind",
    ):
        if period_column in raw.columns:
            raw[period_column] = raw[period_column].astype("string")
    manifest = dict(manifest)
    sources = [dict(s) for s in (manifest.get("sources") or [])]

    for source in sources:
        run_id = str(source.get("capture_run_id") or "")
        doc_year = _resolve_source_document_year(source)
        if doc_year:
            source["document_year"] = doc_year

        if "capture_run_id" not in raw.columns:
            continue
        mask = raw["capture_run_id"].astype(str) == run_id
        if not mask.any():
            continue

        if doc_year and "document_year" in raw.columns:
            raw.loc[mask, "document_year"] = doc_year

        repaired = _resolve_relative_period_years(
            raw.loc[mask].copy(),
            source.get("document_year"),
        )
        for col in repaired.columns:
            if col not in raw.columns:
                raw[col] = pd.Series(pd.NA, index=raw.index, dtype="object")
            else:
                raw[col] = raw[col].astype("object")
            raw.loc[mask, col] = repaired[col].tolist()

        _validate_absolute_year_resolution(
            raw.loc[mask],
            document_year=source.get("document_year"),
            capture_run_id=run_id,
        )

    raw = _ensure_period_identity_columns(raw)

    manifest["sources"] = sources
    manifest["period_resolution_policy"] = (
        "RELATIVE_PERIOD_TO_ABSOLUTE_YEAR_BEFORE_CANONICAL_MERGE"
    )
    return raw, manifest


def _apply_merge_row_exclusions(
    raw: pd.DataFrame,
    metadata: dict[str, Any],
    capture_dir: Path,
) -> pd.DataFrame:
    exclusions = list(metadata.get("merge_row_exclusions") or [])
    metadata["merge_row_exclusions_applied"] = []
    metadata["merge_row_excluded_cell_count"] = 0
    if not exclusions:
        return raw

    required = {
        "table_block_id", "row_order", "column_ordinal",
        "normalized_item", "value_raw", "value",
    }
    missing = sorted(required - set(raw.columns))
    if missing:
        raise ValueError(
            f"MERGE_ROW_EXCLUSION_COLUMNS_MISSING:{capture_dir.name}:{missing}"
        )

    row_orders = pd.to_numeric(raw["row_order"], errors="coerce")
    column_ordinals = pd.to_numeric(raw["column_ordinal"], errors="coerce")
    block_ids = raw["table_block_id"].fillna("").astype(str)
    excluded_indices: set[Any] = set()
    exclusion_keys: set[tuple[str, int, int]] = set()
    applied: list[dict[str, Any]] = []

    for exclusion in exclusions:
        capture_run_id = str(exclusion.get("capture_run_id") or "")
        if capture_run_id and capture_run_id != capture_dir.name:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_CAPTURE_MISMATCH:{capture_dir.name}:{capture_run_id}"
            )
        try:
            block_id = str(exclusion["table_block_id"])
            row_order = int(exclusion["row_order"])
            column_ordinal = int(exclusion["column_ordinal"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_KEY_INVALID:{capture_dir.name}:{exclusion}"
            ) from exc
        key = (block_id, row_order, column_ordinal)
        if key in exclusion_keys:
            raise ValueError(
                f"DUPLICATE_MERGE_ROW_EXCLUSION:{capture_dir.name}:{key}"
            )
        exclusion_keys.add(key)
        mask = (
            block_ids.eq(block_id)
            & row_orders.eq(row_order)
            & column_ordinals.eq(column_ordinal)
        )
        matches = list(raw.index[mask])
        if len(matches) != 1:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_DRIFT:{capture_dir.name}:{key}:matches={len(matches)}"
            )
        index = matches[0]
        if index in excluded_indices:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_OVERLAP:{capture_dir.name}:{key}"
            )
        excluded_indices.add(index)
        source_row = raw.loc[index]
        expected_item_value = exclusion.get("normalized_item")
        actual_item_value = source_row.get("normalized_item")
        expected_raw_value = exclusion.get("value_raw")
        actual_raw_value = source_row.get("value_raw")
        expected_item = "" if expected_item_value is None else str(expected_item_value).strip()
        actual_item = (
            "" if actual_item_value is None or pd.isna(actual_item_value)
            else str(actual_item_value).strip()
        )
        expected_raw = "" if expected_raw_value is None else str(expected_raw_value).strip()
        actual_raw = (
            "" if actual_raw_value is None or pd.isna(actual_raw_value)
            else str(actual_raw_value).strip()
        )
        expected_number = exclusion.get("parsed_number")
        actual_number = pd.to_numeric(
            pd.Series([source_row.get("value")]), errors="coerce"
        ).iloc[0]
        evidence_mismatch = (
            not expected_item
            or actual_item != expected_item
            or actual_raw != expected_raw
            or expected_number is None
            or pd.isna(actual_number)
            or abs(float(actual_number) - float(expected_number))
            > max(1e-9, abs(float(expected_number)) * 1e-12)
        )
        if evidence_mismatch:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_EVIDENCE_MISMATCH:{capture_dir.name}:{key}:"
                f"item={actual_item!r}/{expected_item!r}:"
                f"raw={actual_raw!r}/{expected_raw!r}:"
                f"value={actual_number!r}/{expected_number!r}"
            )
        applied.append({
            **dict(exclusion),
            "capture_run_id": capture_dir.name,
            "table_block_id": block_id,
            "row_order": row_order,
            "column_ordinal": column_ordinal,
            "source_row_id": str(source_row.get("row_id") or ""),
            "source_value_raw": source_row.get("value_raw"),
            "source_value": source_row.get("value"),
            "exclusion_policy": "NON_SOURCE_DERIVED_OBSERVATION",
        })

    metadata["merge_row_exclusions_applied"] = applied
    metadata["merge_row_excluded_cell_count"] = len(applied)
    return raw.drop(index=list(excluded_indices)).reset_index(drop=True)


def load_capture_long(capture_dir: Path, metadata: dict[str, Any], table_id: str) -> pd.DataFrame:
    capture_meta_path = Path(capture_dir) / "capture_metadata.json"
    if capture_meta_path.exists():
        try:
            capture_meta = json.loads(capture_meta_path.read_text(encoding="utf-8"))
        except Exception:
            capture_meta = {}
        lifecycle = str(capture_meta.get("lifecycle_status") or "ACTIVE")
        if lifecycle != "ACTIVE":
            raise ValueError(
                f"CAPTURE_LIFECYCLE_BLOCKED：{Path(capture_dir).name} 当前状态={lifecycle}，"
                "不能进入正式 canonical Merge。请在数据资产管理中心恢复为 ACTIVE 后再合表。"
            )

    path = capture_dir / "table_raw_long.csv"
    if not path.exists():
        raise FileNotFoundError(f"缺少 {path}")
    df = pd.read_csv(
        path,
        dtype={
            "source_period_label": "string",
            "period_label": "string",
            "period_identity": "string",
            "period_kind": "string",
            "year": "string",
            "data_year": "string",
        },
    )
    if df.empty:
        return df

    for col in ("period_year", "period_month", "period_day", "report_year"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    for col in ("year", "data_year", "period_label", "source_period_label"):
        if col in df.columns:
            cleaned = df[col].astype("string").str.strip().replace({"": pd.NA, "<NA>": pd.NA, "nan": pd.NA, "None": pd.NA})
            cleaned = cleaned.str.replace(r"^((?:19|20)\d{2})\.0$", r"\1", regex=True)
            df[col] = cleaned

    df = _apply_merge_row_exclusions(df, metadata, capture_dir)
    if df.empty:
        return df

    # Preserve capture-layer mapping fields without colliding with merge-layer
    # canonicalization columns.
    rename_map = {}
    if "canonical_item" in df.columns:
        rename_map["canonical_item"] = "capture_canonical_item"
    if "mapping_status" in df.columns:
        rename_map["mapping_status"] = "capture_mapping_status"
    if "mapping_note" in df.columns:
        rename_map["mapping_note"] = "capture_mapping_note"
    if rename_map:
        df = df.rename(columns=rename_map)

    effective_document_year = (
        _absolute_year(metadata.get("document_year"))
        or _infer_document_year_from_capture_source(
            pdf_name=metadata.get("pdf_name"),
            capture_dir=capture_dir,
        )
        or ""
    )
    metadata["document_year"] = effective_document_year

    # Canonical Observation Contract v6.6: keep explicit source/report and
    # observation-period fields. Legacy aliases survive only for compatibility
    # with existing templates and resolvers.
    if "report_year" not in df.columns:
        df["report_year"] = effective_document_year
    else:
        df["report_year"] = effective_document_year
    if "data_year" not in df.columns:
        df["data_year"] = df.get("year", "")
    if "statement_scope" not in df.columns:
        df["statement_scope"] = df.get("scope", "")
    if "restated_flag" not in df.columns:
        df["restated_flag"] = df.get("restated", False)
    if "currency_unit" not in df.columns:
        unit_map = {
            "元": "CNY", "千元": "CNY_THOUSAND", "万元": "CNY_TEN_THOUSAND",
            "百万元": "CNY_MILLION", "亿元": "CNY_HUNDRED_MILLION", "%": "PERCENT",
        }
        df["currency_unit"] = df.get("unit", "").map(unit_map).fillna("")
    df = _ensure_period_identity_columns(df)
    # Canonical scope values avoid mixing display labels with observation keys.
    df["statement_scope"] = df["statement_scope"].replace({"本集团": "CONSOLIDATED", "集团": "CONSOLIDATED", "本公司": "COMPANY", "公司": "COMPANY"})

    df.insert(0, "capture_run_id", metadata["capture_run_id"])
    df.insert(1, "company", metadata.get("company", ""))
    df.insert(2, "document_year", effective_document_year)
    df.insert(3, "table_id", table_id)
    # Do not infer this identity from a detail item later in canonicalization.
    # It is the semantic source of the entire Capture.
    df["table_family"] = _identity_text(metadata.get("table_family")) or str(table_id)
    df["member_table"] = _identity_text(metadata.get("member_table")) or _identity_text(metadata.get("table_query"))
    df["member_table_role"] = _identity_text(metadata.get("member_table_role")) or "COMPONENT"
    df["source_table_title"] = _identity_text(metadata.get("source_table_title")) or df["member_table"]
    df["note_reference"] = _identity_text(metadata.get("note_reference")) or _identity_text(metadata.get("note_number"))
    df["source_pdf"] = _identity_text(metadata.get("source_pdf")) or _identity_text(metadata.get("pdf_name"))
    df["member_table_order"] = metadata.get("member_table_order")
    df = annotate_financial_investment_identity(df)
    block_defaults = {
        "container_id": _identity_text(metadata.get("container_id") or metadata.get("note_container_id")),
        "table_block_id": _identity_text(metadata.get("table_block_id") or metadata.get("member_subtable_id")),
        "block_order": (
            metadata.get("block_order")
            if metadata.get("block_order") not in (None, "")
            else -1
        ),
        "classification_axis": _identity_text(metadata.get("classification_axis")) or "UNRESOLVED",
        "block_role": _identity_text(metadata.get("block_role")) or "UNRESOLVED",
        "block_terminal_type": _identity_text(metadata.get("block_terminal_type")) or "UNRESOLVED",
    }
    for column, default in block_defaults.items():
        if column not in df.columns:
            df[column] = default
        else:
            if column == "block_order":
                df[column] = pd.to_numeric(df[column], errors="coerce").fillna(
                    -1 if default in (None, "") else default
                )
            else:
                empty = df[column].isna() | df[column].astype(str).str.strip().isin(
                    {"", "nan", "None", "<NA>"}
                )
                df.loc[empty, column] = default
    if "period_type" not in df.columns:
        df["period_type"] = _identity_text(metadata.get("period_type")) or "ANNUAL"
    if "currency" not in df.columns:
        df["currency"] = _identity_text(metadata.get("currency"))

    # Keep legacy aliases synchronized after source metadata is authoritative.
    df["year"] = df["data_year"]
    df["scope"] = df["statement_scope"]
    df["restated"] = df["restated_flag"]

    df = _resolve_relative_period_years(df, effective_document_year)
    # The resolver canonicalises legacy `year`; make the explicit v6.6
    # observation field authoritative after that resolution as well.
    df["data_year"] = df["year"].map(lambda value: _absolute_year(value) or value)
    df["report_year"] = df["report_year"].map(lambda value: _absolute_year(value) or value)
    _validate_absolute_year_resolution(
        df,
        document_year=effective_document_year,
        capture_run_id=str(metadata["capture_run_id"]),
    )
    # UI and Merge consume the same certified source-row graph projection.
    # Legacy immutable Captures enter an explicitly labelled compatibility
    # adapter; new Captures never infer parents from old hierarchy fields.
    df = assign_semantic_row_keys(df)
    return df


def build_mapping_queue(
    raw_long: pd.DataFrame,
    table_id: str,
    taxonomy: Optional[dict[str, Any]] = None,
) -> pd.DataFrame:
    taxonomy = taxonomy or {"tables": {}}
    lookup = _taxonomy_lookup(taxonomy, table_id)

    numeric = raw_long[raw_long["value"].notna()].copy()
    groups = []
    for source_key, g in numeric.groupby("source_key", sort=False):
        parent_section = normalize_section(
            g["hierarchy_parent_label"].iloc[0]
            if "hierarchy_parent_label" in g else ""
        )
        normalized_item = str(g["normalized_item"].iloc[0] or "")
        companies = sorted({str(x) for x in g["company"].dropna().tolist() if str(x).strip()})
        raw_examples = []
        for x in g["raw_item"].dropna().tolist():
            sx = str(x)
            if sx not in raw_examples:
                raw_examples.append(sx)

        existing = lookup.get(source_key)
        if existing:
            canonical_section = existing.get("canonical_section", parent_section)
            canonical_item = existing.get("canonical_item", normalized_item)
            category = existing.get("category", "")
            status = "AUTO_TAXONOMY"
            suggested = canonical_item
            suggestion_score = 1.0
        else:
            canonical_section = parent_section
            canonical_item = normalized_item
            category = ""
            # Exact same normalized item/context across >=2 captures is safe identity alignment.
            capture_count = g["capture_run_id"].nunique()
            if capture_count >= 2:
                status = "AUTO_EXACT_IDENTITY"
            else:
                status = "UNMAPPED_PRESERVED"
            suggested = ""
            suggestion_score = None

        groups.append({
            "source_key": source_key,
            "parent_section": parent_section,
            "normalized_item": normalized_item,
            "occurrences": int(len(g)),
            "capture_count": int(g["capture_run_id"].nunique()),
            "companies": " | ".join(companies),
            "example_raw_items": " | ".join(raw_examples[:6]),
            "suggested_canonical_section": canonical_section,
            "suggested_canonical_item": suggested,
            "suggestion_score": suggestion_score,
            "canonical_section": canonical_section,
            "canonical_item": canonical_item,
            "category": category,
            "mapping_status": status,
            "mapping_note": "",
        })

    queue = pd.DataFrame(groups)
    if queue.empty:
        return queue

    # Suggest-only fuzzy matches among source labels and known taxonomy canonicals.
    candidates = set(queue["normalized_item"].astype(str).tolist())
    table_tax = taxonomy.get("tables", {}).get(table_id, {})
    for m in table_tax.get("mappings", []):
        if m.get("canonical_item"):
            candidates.add(str(m["canonical_item"]))

    for idx, row in queue.iterrows():
        if row["mapping_status"] != "UNMAPPED_PRESERVED":
            continue
        item = str(row["normalized_item"])
        best_name, best_score = "", 0.0
        for candidate in candidates:
            if candidate == item:
                continue
            score = difflib.SequenceMatcher(None, normalize_text(item), normalize_text(candidate)).ratio()
            if score > best_score:
                best_name, best_score = candidate, score
        if best_score >= 0.72:
            queue.at[idx, "suggested_canonical_item"] = best_name
            queue.at[idx, "suggestion_score"] = round(best_score, 4)

    return queue


def apply_mapping(raw_long: pd.DataFrame, mapping_queue: pd.DataFrame) -> pd.DataFrame:
    if raw_long.empty:
        return raw_long.copy()

    # Legacy callers may provide only the old source_key. Materialise the
    # shared hierarchy projection once, while preserving that key solely for
    # taxonomy lookup compatibility.
    projected_source_key = raw_long.get("source_key")
    if "semantic_row_key" not in raw_long.columns:
        projected = assign_semantic_row_keys(raw_long)
        if projected_source_key is not None:
            existing = projected_source_key.astype(str).str.strip()
            projected.loc[existing.ne(""), "source_key"] = existing[existing.ne("")]
        raw_long = projected

    map_cols = [
        "source_key", "canonical_section", "canonical_item", "category",
        "mapping_status", "mapping_note",
    ]
    mapping = mapping_queue[map_cols].drop_duplicates("source_key")
    # Capture Long already carries machine-level canonical/mapping placeholders.
    # The reviewed mapping queue is authoritative at merge time; remove those
    # placeholders before joining so pandas does not create unusable _x/_y
    # columns and silently drop the canonical contract expected below.
    mapping_payload_columns = [
        column for column in map_cols[1:] if column in raw_long.columns
    ]
    out = raw_long.drop(columns=mapping_payload_columns).merge(
        mapping,
        on="source_key",
        how="left",
    )

    certified_parent = (
        out["hierarchy_parent_label"]
        if "hierarchy_parent_label" in out.columns
        else pd.Series("", index=out.index)
    )
    out["canonical_section"] = out["canonical_section"].fillna(
        certified_parent.fillna("").map(normalize_section)
    )
    out["canonical_item"] = out["canonical_item"].fillna(out["normalized_item"])
    out["mapping_status"] = out["mapping_status"].fillna("UNMAPPED_PRESERVED")
    out["category"] = out["category"].fillna("")

    accepted = {
        "AUTO_TAXONOMY",
        "AUTO_EXACT_IDENTITY",
        "CONFIRMED",
        "CONFIRMED_OVERRIDE",
    }

    def key_for(row) -> str:
        section = normalize_section(row.get("canonical_section"))
        item = str(row.get("canonical_item") or row.get("normalized_item") or "").strip()
        if row.get("mapping_status") in accepted:
            prefix = "CANON"
        else:
            prefix = "RAW"
        parent_path = _normalize_semantic_path(row.get("semantic_parent_path"))
        family = _identity_text(row.get("table_family"))
        member = _identity_text(row.get("member_table"))
        role = _identity_text(row.get("member_table_role"))
        # Missing member identity is deliberately isolated by Capture rather
        # than collapsed into a false value comparison. materialize_canonical
        # will emit REVIEW_REQUIRED_SOURCE_IDENTITY for it.
        if not family or not member or not role:
            source_scope = f"MISSING_SOURCE::{_identity_text(row.get('capture_run_id')) or 'UNKNOWN'}"
        else:
            source_scope = f"FAMILY::{family}::MEMBER::{member}::ROLE::{role}"
        axis = _identity_text(row.get("classification_axis"))
        if axis and axis != "UNRESOLVED":
            source_scope += f"::AXIS::{axis}"
        else:
            block_id = _identity_text(row.get("table_block_id"))
            if block_id:
                source_scope += f"::BLOCK::{block_id}::AXIS::UNRESOLVED"
        occurrence = int(row.get("semantic_occurrence") or 1)
        occurrence_suffix = ""
        source_key = _identity_text(row.get("semantic_row_key"))
        if "||OCC::" in source_key:
            occurrence_suffix = f"::OCC::{occurrence}"
        unresolved_suffix = ""
        if _identity_text(row.get("semantic_identity_status")) not in {
            "CERTIFIED", "LEGACY_IDENTITY_COMPATIBILITY",
        }:
            unresolved_suffix = (
                f"::UNRESOLVED_SOURCE::{_identity_text(row.get('capture_run_id'))}"
                f"::{_identity_text(row.get('source_row_id'))}"
            )
        return (
            f"{prefix}::{row.get('table_id')}::{source_scope}::"
            f"PARENT::{parent_path or section or 'ROOT'}::ITEM::{item}"
            f"{occurrence_suffix}{unresolved_suffix}"
        )

    out["canonical_key"] = out.apply(key_for, axis=1)
    return out


def _dimension_label(row: pd.Series) -> str:
    values = {
        "company": row.get("company"),
        "report_year": row.get("report_year", row.get("document_year")),
        "period_label": row.get("period_label"),
        "period_identity": row.get("period_identity"),
        "period_year": row.get("period_year"),
        "period_month": row.get("period_month"),
        "period_day": row.get("period_day"),
        "period_precision": row.get("period_precision"),
        "period_date": row.get("period_date"),
        "data_year": row.get("data_year", row.get("year")),
        "period_type": row.get("period_type"),
        "currency_unit": row.get("currency_unit"),
        "statement_scope": row.get("statement_scope", row.get("scope")),
        "restated_flag": row.get("restated_flag", row.get("restated")),
        "measure": row.get("measure"),
    }
    parts = []
    for key, value in values.items():
        if key == "restated_flag":
            parts.append(f"{key}={'TRUE' if _as_bool(value) else 'FALSE'}")
        elif value is not None and str(value).strip() not in {"", "nan", "None"}:
            parts.append(f"{key}={str(value).strip()}")
    return " | ".join(parts)


def _as_bool(value: Any) -> bool:
    """Interpret persisted CSV flags without treating the string 'False' as true."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "已重述", "经重述", "重述"}
    return bool(value)


def _dimension_missing(value: Any) -> bool:
    return value is None or (isinstance(value, float) and pd.isna(value)) or str(value).strip() in {"", "nan", "None", "<NA>"}


def _unique_source_sequence(mapped_long: pd.DataFrame, capture_run_id: str) -> list[dict[str, Any]]:
    """
    Collapse long-form observations to one certified source row.
    Preserve exact source order.
    """
    g = mapped_long[mapped_long["capture_run_id"].astype(str) == str(capture_run_id)].copy()
    if g.empty:
        return []

    g["_row_order_num"] = pd.to_numeric(g.get("row_order"), errors="coerce")
    g = g.sort_values(["_row_order_num"], kind="stable")

    rows = []
    seen_source_rows = set()
    for _, r in g.iterrows():
        ro = r.get("_row_order_num")
        if pd.isna(ro):
            continue
        ro_int = int(ro)
        source_row_id = _identity_text(r.get("source_row_id")) or f"LEGACY_ROW_ORDER::{ro_int}"
        if source_row_id in seen_source_rows:
            continue
        seen_source_rows.add(source_row_id)
        key = str(r.get("canonical_key") or "").strip()
        if not key:
            continue
        rows.append({
            "canonical_key": key,
            "row_order": ro_int,
            "row_type": str(r.get("row_role") or r.get("row_type") or ""),
            "row_level": r.get("hierarchy_level") if r.get("hierarchy_level") is not None else r.get("row_level"),
            "canonical_section": str(r.get("canonical_section") or ""),
            "canonical_item": str(r.get("canonical_item") or r.get("normalized_item") or ""),
            "parent_section": str(r.get("hierarchy_parent_label") or r.get("parent_section") or ""),
            "raw_item": str(r.get("raw_item") or ""),
            "source_row_id": source_row_id,
            "parent_row_id": _identity_text(r.get("parent_row_id")),
            "semantic_row_key": _identity_text(r.get("semantic_row_key")),
            "table_family": _identity_text(r.get("table_family")),
            "member_table": _identity_text(r.get("member_table")),
            "member_table_role": _identity_text(r.get("member_table_role")),
            "source_table_title": _identity_text(r.get("source_table_title")),
            "note_reference": _identity_text(r.get("note_reference")),
            "member_table_order": r.get("member_table_order"),
            "row_path": _identity_text(r.get("hierarchy_path") or r.get("row_path")),
            "table_block_id": _identity_text(r.get("table_block_id")),
            "classification_axis": _identity_text(r.get("classification_axis")),
            "block_role": _identity_text(r.get("block_role")),
            "total_confidence": _identity_text(r.get("total_confidence")),
        })
    return rows


def _dedupe_key_sequence(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    seq = []
    first_meta = {}
    duplicates = []
    for row in rows:
        key = row["canonical_key"]
        if key in first_meta:
            duplicates.append({
                "conflict_type": "DUPLICATE_CANONICAL_KEY_IN_SOURCE",
                "canonical_key": key,
                "first_row_order": first_meta[key]["row_order"],
                "duplicate_row_order": row["row_order"],
                "detail": (
                    "同一来源中多个原始行被映射到同一 canonical_key；"
                    "顺序层仅保留首次出现位置，数值层仍由 VALUE_CONFLICT/UNIT_CONFLICT 独立检查。"
                ),
            })
            continue
        first_meta[key] = row
        seq.append(key)
    return seq, duplicates


def _shared_order_conflicts(
    reference_seq: list[str],
    other_seq: list[str],
    capture_run_id: str,
) -> list[dict[str, Any]]:
    """
    Detect inversions among shared keys. Reference order is authoritative and
    will never be reordered to satisfy a conflicting source.
    """
    ref_pos = {k: i for i, k in enumerate(reference_seq)}
    shared = [k for k in other_seq if k in ref_pos]
    conflicts = []
    max_seen = -1
    max_key = None
    for key in shared:
        pos = ref_pos[key]
        if pos < max_seen:
            conflicts.append({
                "conflict_type": "ORDER_CONFLICT",
                "capture_run_id": capture_run_id,
                "canonical_key": key,
                "reference_predecessor_key": max_key,
                "detail": (
                    "该来源对共同项目的相对顺序与排序基准表冲突；"
                    "最终合表继续严格采用排序基准表顺序，不自动重排。"
                ),
            })
        if pos > max_seen:
            max_seen = pos
            max_key = key
    return conflicts


def _row_partition_key(meta: dict[str, Any]) -> tuple[str, str, str]:
    table_family = _identity_text(meta.get("table_family"))
    member_table = _identity_text(meta.get("member_table"))
    classification_axis = _identity_text(meta.get("classification_axis"))
    block_dim = classification_axis or _identity_text(meta.get("table_block_id")) or ""
    return (table_family, member_table, block_dim)


def _is_terminal_total(meta: dict[str, Any]) -> bool:
    row_type = str(meta.get("row_type") or "").upper()
    if row_type in ("TOTAL", "GRAND_TOTAL"):
        return True
    if row_type == "IMPLICIT_TOTAL":
        confidence = str(meta.get("total_confidence") or "").upper()
        if confidence == "HIGH":
            return True
    return False


def _is_subtotal(meta: dict[str, Any]) -> bool:
    return str(meta.get("row_type") or "").upper() == "SUBTOTAL"


def _is_footnote_or_memo(meta: dict[str, Any]) -> bool:
    row_type = str(meta.get("row_type") or "").upper()
    return row_type in ("FOOTNOTE", "NARRATIVE", "MEMO", "SPACER")


def _merge_missing_keys_preserving_context(
    base: list[str],
    incoming: list[str],
    base_meta: Optional[dict[str, dict[str, Any]]] = None,
    incoming_meta: Optional[dict[str, dict[str, Any]]] = None,
) -> list[str]:
    """
    Insert source-unique keys while respecting:
    1. Partition boundary (table_family, member_table, table_block_id)
    2. Hierarchy subtree boundary (parent_section / canonical_section)
    3. Terminal Total boundary (TOTAL / GRAND_TOTAL terminal row)
    4. Subtotal boundary (SUBTOTAL of subtree)
    Existing keys are never reordered.
    """
    out = list(base)
    base_meta_dict = dict(base_meta or {})
    incoming_meta_dict = dict(incoming_meta or {})

    incoming = [k for i, k in enumerate(incoming) if k and k not in incoming[:i]]

    i = 0
    while i < len(incoming):
        if incoming[i] in out:
            i += 1
            continue

        # Collect a consecutive missing block belonging to same partition/parent.
        j = i
        block = []
        first_k = incoming[i]
        first_m = incoming_meta_dict.get(first_k, {})
        first_part = _row_partition_key(first_m)
        first_parent = _identity_text(first_m.get("parent_section") or first_m.get("canonical_section"))

        while j < len(incoming) and incoming[j] not in out:
            curr_k = incoming[j]
            curr_m = incoming_meta_dict.get(curr_k, {})
            curr_part = _row_partition_key(curr_m)
            curr_parent = _identity_text(curr_m.get("parent_section") or curr_m.get("canonical_section"))
            if j > i and (curr_part != first_part or curr_parent != first_parent):
                break
            block.append(curr_k)
            j += 1

        # 1. Locate partition indices in out
        part_indices = [
            idx for idx, k in enumerate(out)
            if _row_partition_key(base_meta_dict.get(k, incoming_meta_dict.get(k, {}))) == first_part
        ]

        # 2. Locate Terminal Total index in partition
        total_idx = None
        if part_indices:
            for idx in part_indices:
                k = out[idx]
                m = base_meta_dict.get(k, incoming_meta_dict.get(k, {}))
                if _is_terminal_total(m):
                    total_idx = idx
                    break

        # 3. Find nearest known anchors in incoming
        prev_known = None
        for k in reversed(incoming[:i]):
            if k in out:
                prev_known = k
                break

        next_known = None
        for k in incoming[j:]:
            if k in out:
                next_known = k
                break

        # 4. Check if parent exists in out
        parent_idx = None
        last_descendant_idx = None
        subtotal_idx = None
        if first_parent:
            for idx in part_indices:
                k = out[idx]
                m = base_meta_dict.get(k, incoming_meta_dict.get(k, {}))
                item_name = _identity_text(m.get("canonical_item") or m.get("normalized_item") or m.get("canonical_key"))
                if item_name == first_parent or k == first_parent:
                    parent_idx = idx
                    break

            if parent_idx is not None:
                for idx in range(parent_idx + 1, len(out)):
                    k = out[idx]
                    m = base_meta_dict.get(k, incoming_meta_dict.get(k, {}))
                    if _row_partition_key(m) != first_part or _is_terminal_total(m):
                        break
                    p_sec = _identity_text(m.get("parent_section") or m.get("canonical_section"))
                    if p_sec == first_parent:
                        if _is_subtotal(m):
                            subtotal_idx = idx
                            break
                        last_descendant_idx = idx
                    elif last_descendant_idx is not None:
                        break

        # Decide insertion point
        if parent_idx is not None:
            if last_descendant_idx is not None:
                insert_at = last_descendant_idx + 1
            else:
                insert_at = parent_idx + 1
            if subtotal_idx is not None and insert_at > subtotal_idx:
                insert_at = subtotal_idx
            if total_idx is not None and not _is_footnote_or_memo(first_m) and insert_at > total_idx:
                insert_at = total_idx
        elif prev_known is not None and next_known is not None:
            prev_idx = out.index(prev_known)
            next_idx = out.index(next_known)
            if prev_idx < next_idx:
                insert_at = next_idx
            else:
                insert_at = prev_idx + 1
        elif prev_known is not None:
            insert_at = out.index(prev_known) + 1
            while insert_at < len(out) and out[insert_at] in incoming[:i]:
                insert_at += 1
        elif next_known is not None:
            insert_at = out.index(next_known)
        elif part_indices:
            if total_idx is not None and not _is_footnote_or_memo(first_m):
                insert_at = total_idx
            else:
                insert_at = part_indices[-1] + 1
        else:
            insert_at = len(out)

        # Enforce partition & total boundaries
        if total_idx is not None and not _is_footnote_or_memo(first_m):
            if insert_at > total_idx:
                insert_at = total_idx

        if part_indices:
            p_start = part_indices[0]
            p_end = part_indices[-1]
            if insert_at < p_start:
                insert_at = p_start
            elif insert_at > p_end + 1:
                insert_at = p_end + 1

        out[insert_at:insert_at] = block
        for bk in block:
            base_meta_dict[bk] = incoming_meta_dict.get(bk, {})
        i = j

    return out


def _normalize_logical_table_row_order(
    ordered_keys: list[str],
    meta_by_key: dict[str, dict[str, Any]],
) -> list[str]:
    """
    Perform deterministic, minimal repairs per logical partition:
    1. Ensure TOTAL is positioned as the terminal row of Data Body (after all DETAIL/PARENT/SUBTOTAL, before FOOTNOTE).
    Preserves all other relative orderings.
    """
    if not ordered_keys:
        return []

    partitions: list[tuple[tuple[str, str, str], list[str]]] = []
    curr_part = None
    curr_keys: list[str] = []

    for k in ordered_keys:
        m = meta_by_key.get(k, {})
        part = _row_partition_key(m)
        if part != curr_part:
            if curr_keys:
                partitions.append((curr_part, curr_keys))
            curr_part = part
            curr_keys = [k]
        else:
            curr_keys.append(k)
    if curr_keys:
        partitions.append((curr_part, curr_keys))

    normalized_keys: list[str] = []
    for part, p_keys in partitions:
        data_keys = []
        footnote_keys = []
        total_keys = []

        for k in p_keys:
            m = meta_by_key.get(k, {})
            if _is_footnote_or_memo(m):
                footnote_keys.append(k)
            elif _is_terminal_total(m):
                total_keys.append(k)
            else:
                data_keys.append(k)

        normalized_partition = data_keys + total_keys + footnote_keys
        normalized_keys.extend(normalized_partition)

    return normalized_keys


def _validate_logical_table_order(
    ordered_keys: list[str],
    meta_by_key: dict[str, dict[str, Any]],
    reference_id: str,
    reference_seq: list[str],
) -> list[dict[str, Any]]:
    """
    Validate the 6 QA Invariants on the merged structural row order:
    - Invariant 1: No DETAIL/PARENT/SUBTOTAL after TOTAL in Data Body.
    - Invariant 2: TOTAL is Data Body terminal row.
    - Invariant 3: SUBTOTAL must be positioned after its descendants.
    - Invariant 4: DETAIL must stay within its ancestor subtree.
    - Invariant 5: Missing historical rows do not cross partition boundaries.
    - Invariant 6: Relative order of reference benchmark rows is preserved.
    """
    violations = []

    # Invariant 6: Benchmark relative order check
    ref_pos = {k: i for i, k in enumerate(ordered_keys) if k in reference_seq}
    ref_indices = [ref_pos[k] for k in reference_seq if k in ref_pos]
    if ref_indices != sorted(ref_indices):
        violations.append({
            "conflict_type": "BENCHMARK_ORDER_INVERSION",
            "capture_run_id": reference_id,
            "canonical_key": "",
            "detail": "基准表原始相对行序被意外改变，违反 Invariant 6。",
        })

    # Group by partition to check Invariants 1-5
    curr_part = None
    part_keys: list[str] = []

    def _check_partition(p_keys: list[str]) -> None:
        seen_total = False
        total_key = None
        for k in p_keys:
            m = meta_by_key.get(k, {})
            if _is_terminal_total(m):
                seen_total = True
                total_key = k
            elif seen_total and not _is_footnote_or_memo(m):
                violations.append({
                    "conflict_type": "DATA_ROW_AFTER_TOTAL",
                    "capture_run_id": str(m.get("capture_run_id") or ""),
                    "canonical_key": k,
                    "reference_predecessor_key": total_key,
                    "detail": f"明细/父级/小计项 [{k}] 出现在合计行 [{total_key}] 之后，违反 Invariant 1/2。",
                })

    for k in ordered_keys:
        m = meta_by_key.get(k, {})
        part = _row_partition_key(m)
        if part != curr_part:
            if part_keys:
                _check_partition(part_keys)
            curr_part = part
            part_keys = [k]
        else:
            part_keys.append(k)
    if part_keys:
        _check_partition(part_keys)

    return violations


def build_structural_order(
    mapped_long: pd.DataFrame,
    manifest: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build one authoritative canonical row order.

    Policy:
    1. Reference capture order is immutable.
    2. Other captures may contribute missing keys around nearest known anchors.
    3. Shared-key inversions are reported as ORDER_CONFLICT, never used to
       reorder the reference.
    """
    if mapped_long.empty:
        return (
            pd.DataFrame(columns=[
                "canonical_order", "canonical_key", "row_type", "row_level",
                "canonical_section", "canonical_item", "parent_section", "order_source",
                "reference_row_order",
            ]),
            pd.DataFrame(),
        )

    manifest_sources = [x for x in (manifest.get("sources") or []) if x.get("capture_run_id")]
    # A Family is ordered by its explicit member order, then by the stable
    # caller/plan order. Never use one member's rows as the reference rows for
    # a different member table.
    def source_order(item: tuple[int, dict[str, Any]]) -> tuple[int, int]:
        index, source = item
        try:
            return (int(source.get("member_table_order")), index)
        except (TypeError, ValueError):
            return (10**9, index)
    manifest_sources = [source for _, source in sorted(enumerate(manifest_sources), key=source_order)]
    source_ids = [str(x.get("capture_run_id")) for x in manifest_sources]
    if not source_ids:
        source_ids = list(dict.fromkeys(mapped_long["capture_run_id"].astype(str).tolist()))

    requested_ref = str(manifest.get("reference_capture_run_id") or "")
    reference_id = requested_ref if requested_ref in source_ids else source_ids[0]

    sequences = {}
    row_meta_by_source = {}
    conflicts = []

    for source_id in source_ids:
        rows = _unique_source_sequence(mapped_long, source_id)
        seq, dup_conflicts = _dedupe_key_sequence(rows)
        for c in dup_conflicts:
            c["capture_run_id"] = source_id
        conflicts.extend(dup_conflicts)
        sequences[source_id] = seq
        row_meta_by_source[source_id] = {r["canonical_key"]: r for r in rows if r["canonical_key"]}

    base: list[str] = []
    order_source: dict[str, str] = {}
    note_ordinal_map: dict[str, Any] = {}
    order_policy = str(
        manifest.get("order_policy")
        or "REFERENCE_CAPTURE_PRESERVE_WITH_CONTEXTUAL_INSERTION"
    )
    reference_year = str(manifest.get("reference_report_year") or "").strip()
    if order_policy == NOTE_ORDINAL_ORDER_POLICY:
        built = _note_ordinal_base_order(
            mapped_long, manifest, reference_year,
            row_meta_by_source=row_meta_by_source,
            reference_id=reference_id,
        )
        if built is not None:
            base, order_source, note_ordinal_map = built
    if not base:
        base = list(sequences.get(reference_id, []))
        order_source = {k: f"REFERENCE:{reference_id}" for k in base}

    ref_meta = row_meta_by_source.get(reference_id, {})
    current_meta = dict(ref_meta)

    for source_id in source_ids:
        if source_id == reference_id:
            continue
        seq = sequences.get(source_id, [])
        conflicts.extend(_shared_order_conflicts(base, seq, source_id))
        before = set(base)
        src_meta = row_meta_by_source.get(source_id, {})
        merged = _merge_missing_keys_preserving_context(
            base,
            seq,
            base_meta=current_meta,
            incoming_meta=src_meta,
        )
        for k in merged:
            if k not in before and k not in order_source:
                order_source[k] = f"INSERTED_FROM:{source_id}"
            if k in src_meta and k not in current_meta:
                current_meta[k] = src_meta[k]
        base = merged

    # Include any unexpected keys not represented in manifest sources.
    all_keys = list(dict.fromkeys(mapped_long["canonical_key"].dropna().astype(str).tolist()))
    for key in all_keys:
        if key not in base:
            base.append(key)
            order_source[key] = "APPENDED_UNREFERENCED"

    # Merge metadata across all sources
    all_meta_by_key = {}
    for key in base:
        meta = ref_meta.get(key)
        if meta is None:
            for source_id in source_ids:
                if key in row_meta_by_source.get(source_id, {}):
                    meta = row_meta_by_source[source_id][key]
                    break
        all_meta_by_key[key] = meta or {"canonical_key": key}

    # Phase 2: Normalization
    base = _normalize_logical_table_row_order(base, all_meta_by_key)

    # Phase 2: QA Invariant Validation
    ref_seq = sequences.get(reference_id, [])
    order_violations = _validate_logical_table_order(
        base, all_meta_by_key, reference_id=reference_id, reference_seq=ref_seq,
    )
    conflicts.extend(order_violations)

    # Structural metadata: reference source first, then first source containing key.
    order_rows = []
    for idx, key in enumerate(base, start=1):
        meta = ref_meta.get(key)
        meta_source = reference_id if meta else None
        if meta is None:
            for source_id in source_ids:
                if key in row_meta_by_source.get(source_id, {}):
                    meta = row_meta_by_source[source_id][key]
                    meta_source = source_id
                    break
        meta = meta or {}
        order_rows.append({
            "canonical_order": idx,
            "canonical_key": key,
            "row_type": meta.get("row_type", ""),
            "row_level": meta.get("row_level"),
            "canonical_section": meta.get("canonical_section", ""),
            "parent_section": meta.get("parent_section", ""),
            "canonical_item": meta.get("canonical_item", ""),
            "table_family": meta.get("table_family", ""),
            "member_table": meta.get("member_table", ""),
            "member_table_role": meta.get("member_table_role", ""),
            "source_table_title": meta.get("source_table_title", ""),
            "note_reference": meta.get("note_reference", ""),
            "member_table_order": meta.get("member_table_order"),
            "row_path": meta.get("row_path", ""),
            "table_block_id": meta.get("table_block_id", ""),
            "classification_axis": meta.get("classification_axis", ""),
            "note_ordinal": note_ordinal_map.get(key),
            "order_source": order_source.get(key, f"FIRST_SEEN:{meta_source or 'UNKNOWN'}"),
            "reference_capture_run_id": reference_id,
            "reference_row_order": (
                ref_meta.get(key, {}).get("row_order")
                if key in ref_meta else None
            ),
            "metadata_source_capture_run_id": meta_source,
        })

    order_df = pd.DataFrame(order_rows)
    conflict_cols = [
        "conflict_type", "capture_run_id", "canonical_key",
        "reference_predecessor_key", "first_row_order",
        "duplicate_row_order", "detail",
    ]
    conflicts_df = pd.DataFrame(conflicts)
    for col in conflict_cols:
        if col not in conflicts_df.columns:
            conflicts_df[col] = None
    conflicts_df = conflicts_df[conflict_cols]

    return order_df, conflicts_df


NOTE_ORDINAL_ORDER_POLICY = "NOTE_ORDINAL_REFERENCE_YEAR"


def _note_ordinal_value(note_reference: Any) -> int | None:
    from table_boundary_resolver import parse_note_ordinal

    try:
        return parse_note_ordinal(str(note_reference or "").strip())
    except Exception:
        return None


def _note_ordinal_base_order(
    mapped_long: pd.DataFrame,
    manifest: dict[str, Any],
    reference_year: str,
    *,
    row_meta_by_source: dict[str, dict[str, dict[str, Any]]],
    reference_id: str,
) -> tuple[list[str], dict[str, str], dict[str, int | None]] | None:
    """Order members by the note ordinal of a user-selected reference year.

    Top-level order follows the selected year's note sequence in the annual
    report (via ``note_reference`` ordinal); within a member, rows keep the
    selected-year capture order, then the reference capture, then first seen.
    Members with no capture in the selected year are appended by their
    member_table_order hint.  Returns ``None`` when the policy cannot be
    applied so the caller falls back to the legacy reference-capture policy.
    """
    if not reference_year:
        return None
    year_sources = [
        source for source in (manifest.get("sources") or [])
        if str(
            source.get("document_year") or source.get("report_year") or ""
        ) == reference_year
    ]
    if not year_sources:
        return None
    year_run_ids = {
        str(source.get("capture_run_id")) for source in year_sources
    }
    member_order_hint: dict[str, int] = {}
    for source in year_sources:
        member = str(source.get("member_table") or "")
        try:
            member_order_hint.setdefault(
                member, int(source.get("member_table_order") or 0)
            )
        except (TypeError, ValueError):
            member_order_hint.setdefault(member, 10**9)

    key_member: dict[str, str] = {}
    key_first_order: dict[str, int] = {}
    key_year_order: dict[str, int] = {}
    key_ref_order: dict[str, int] = {}
    for source_id, meta_by_key in row_meta_by_source.items():
        for key, meta in meta_by_key.items():
            member = str(meta.get("member_table") or "")
            if member and key not in key_member:
                key_member[key] = member
            try:
                row_order = int(meta.get("row_order") or 0)
            except (TypeError, ValueError):
                row_order = 10**9
            key_first_order.setdefault(key, row_order)
            if source_id in year_run_ids:
                key_year_order.setdefault(key, row_order)
            if source_id == reference_id:
                key_ref_order.setdefault(key, row_order)

    member_note: dict[str, tuple[int, str]] = {}
    for source in year_sources:
        member = str(source.get("member_table") or "")
        note_ref = str(source.get("note_reference") or "")
        ordinal = _note_ordinal_value(note_ref)
        if ordinal is None:
            continue
        current = member_note.get(member)
        if current is None or ordinal < current[0]:
            member_note[member] = (ordinal, note_ref)
    all_members = sorted({member for member in key_member.values() if member})
    members_with_note = sorted(
        member_note.items(),
        key=lambda item: (
            item[1][0], member_order_hint.get(item[0], 0),
        ),
    )
    members_without_note = sorted(
        [member for member in all_members if member not in member_note],
        key=lambda member: (member_order_hint.get(member, 0), member),
    )
    ordered_members = [
        member for member, _ in members_with_note
    ] + members_without_note

    base: list[str] = []
    order_source: dict[str, str] = {}
    note_ordinal_map: dict[str, int | None] = {}
    for member in ordered_members:
        note_ord, note_ref = member_note.get(member, (None, ""))
        year_keys = [
            key for key, owner in key_member.items()
            if owner == member and key in key_year_order
        ]
        if year_keys:
            member_keys = sorted(year_keys, key=lambda k: (key_year_order[k], k))
        else:
            ref_keys = [
                key for key, owner in key_member.items()
                if owner == member and key in key_ref_order
            ]
            if ref_keys:
                member_keys = sorted(ref_keys, key=lambda k: (key_ref_order[k], k))
            else:
                first_keys = [
                    key for key, owner in key_member.items()
                    if owner == member and key in key_first_order
                ]
                member_keys = sorted(first_keys, key=lambda k: (key_first_order[k], k))

        for key in member_keys:
            base.append(key)
            order_source[key] = (
                f"NOTE_ORDINAL:{note_ref or 'UNRESOLVED'}:{reference_year}"
            )
            note_ordinal_map[key] = note_ord
    return base, order_source, note_ordinal_map


def materialize_canonical(
    mapped_long: pd.DataFrame,
    structural_order: Optional[pd.DataFrame] = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    numeric = _ensure_period_identity_columns(
        mapped_long[mapped_long["value"].notna()].copy()
    )
    if numeric.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    dims = [
        "table_id", "table_family", "member_table", "member_table_role",
        "presentation_member_id", "presentation_regime",
        "canonical_analysis_bucket", "comparability_status",
        "analysis_bridge_groups", "member_contract_version",
        "bridge_certification_status", "certified_bridge_rule_id",
        *CANONICAL_BLOCK_IDENTITY_COLUMNS,
        "semantic_parent_path", "semantic_occurrence",
        "source_table_title", "canonical_key", "canonical_section", "canonical_item",
        "company", "report_year", "period_identity", "period_label",
        "period_year", "period_month", "period_day", "period_precision",
        "period_date", "period_kind", "statement_scope", "restated_flag",
        "period_type", "currency_unit", "unit",
        "measure",
    ]
    # Accept old capture exports while always materialising the explicit v6.6
    # canonical observation fields.
    aliases = {
        "report_year": "document_year", "data_year": "year",
        "statement_scope": "scope", "restated_flag": "restated",
    }
    for canonical, legacy in aliases.items():
        if canonical not in numeric.columns and legacy in numeric.columns:
            numeric[canonical] = numeric[legacy]
    if "currency_unit" not in numeric.columns:
        unit_map = {"元": "CNY", "千元": "CNY_THOUSAND", "万元": "CNY_TEN_THOUSAND", "百万元": "CNY_MILLION", "亿元": "CNY_HUNDRED_MILLION", "%": "PERCENT"}
        numeric["currency_unit"] = (numeric["unit"].map(unit_map).fillna("") if "unit" in numeric.columns else "")
    for column, default in {
        "table_family": "", "member_table": "", "member_table_role": "",
        "presentation_member_id": "", "presentation_regime": "",
        "canonical_analysis_bucket": "", "comparability_status": "",
        "analysis_bridge_groups": "[]", "member_contract_version": "",
        "bridge_certification_status": "", "certified_bridge_rule_id": "",
        "container_id": "", "table_block_id": "", "block_order": -1,
        "classification_axis": "UNRESOLVED", "block_role": "UNRESOLVED",
        "block_terminal_type": "UNRESOLVED",
        "semantic_parent_path": "", "semantic_occurrence": 1,
        "row_path": "", "period_type": "", "currency_unit": "", "unit": "", "measure": "",
        "source_table_title": "", "note_reference": "", "source_pdf": "",
        "semantic_row_key": "", "semantic_identity_status": "LEGACY_IDENTITY_COMPATIBILITY",
    }.items():
        if column not in numeric.columns:
            numeric[column] = default
        elif column in BLOCK_IDENTITY_COLUMNS:
            # Non-compound callers legitimately have no block identity.
            # Normalise nulls to explicit legacy-safe sentinels before the
            # wide groupby/pivot; pandas otherwise drops the complete numeric
            # observation merely because one grouping dimension is null.
            numeric[column] = numeric[column].where(
                numeric[column].notna(),
                default,
            )
    # Old merge projects may be refreshed from CSV artifacts whose years were
    # parsed as floats and whose flags were persisted as display strings.
    # Normalise these fields before they become a wide-column identity.
    numeric["report_year"] = numeric["report_year"].map(lambda value: _absolute_year(value) or value)
    numeric["data_year"] = numeric["period_year"].astype("string")
    numeric["statement_scope"] = numeric["statement_scope"].replace({
        "本集团": "CONSOLIDATED", "集团": "CONSOLIDATED",
        "本公司": "COMPANY", "公司": "COMPANY",
    })
    numeric["restated_flag"] = numeric["restated_flag"].map(_as_bool)

    unresolved_amount_unit = (
        numeric["measure"].astype(str).str.strip().eq("金额")
        & numeric["unit"].astype(str).str.strip().isin({"", "nan", "None"})
    )
    if unresolved_amount_unit.any():
        bad = numeric.loc[
            unresolved_amount_unit,
            ["capture_run_id", "canonical_item", "report_year", "data_year"],
        ].drop_duplicates()
        raise ValueError(
            "UNIT_UNRESOLVED_AMOUNT_OBSERVATION:"
            + bad.to_json(orient="records", force_ascii=False)
        )

    resolved_rows = []
    conflicts = []

    unit_identity_columns = [
        column for column in dims
        if column not in {"currency_unit", "unit"}
    ]
    def _unit_identity_key(values: Any) -> tuple[str, ...]:
        if not isinstance(values, tuple):
            values = (values,)
        return tuple(
            "" if value is None or pd.isna(value) else str(value).strip()
            for value in values
        )

    unit_conflict_identities: set[tuple[str, ...]] = set()
    for identity, unit_group in numeric.groupby(
        unit_identity_columns, dropna=False, sort=False,
    ):
        units = {
            str(value).strip()
            for value in unit_group["unit"].dropna().tolist()
            if str(value).strip() and str(value).lower() != "nan"
        }
        if len(units) > 1:
            unit_conflict_identities.add(_unit_identity_key(identity))

    # v6.10: cross-family merge guard.  The same item label (e.g. "定期存款")
    # may appear in both financial_investment and investment_portfolio
    # families.  Detect when the same canonical_key prefix (without family
    # qualifier) maps to different table_family values and emit a warning.
    if "table_family" in numeric.columns and "canonical_item" in numeric.columns:
        item_families = (
            numeric[["canonical_item", "table_family"]]
            .drop_duplicates()
            .groupby("canonical_item")["table_family"]
            .apply(set)
        )
        cross_family_items = item_families[item_families.map(len) > 1]
        for item, families in cross_family_items.items():
            conflicts.append({
                "conflict_type": "CROSS_FAMILY_MERGE_CONFLICT",
                "canonical_item": item,
                "families": sorted(families),
                "detail": (
                    f"同一项目 '{item}' 出现在多个 Table Family 中："
                    f"{'、'.join(sorted(families))}。"
                    "不得仅按字符串相等直接合并；必须通过 bridge contract 建立跨表族映射。"
                ),
            })

    for key, g in numeric.groupby(dims, dropna=False, sort=False):
        values = [float(x) for x in g["value"].dropna().tolist()]
        unique_values = []
        for v in values:
            if not any(abs(v - u) <= max(1e-8, abs(u) * 1e-10) for u in unique_values):
                unique_values.append(v)

        conflict_reasons = []
        unit_identity = _unit_identity_key(tuple(
            key[dims.index(column)] for column in unit_identity_columns
        ))
        if unit_identity in unit_conflict_identities:
            conflict_reasons.append("REVIEW_REQUIRED_UNIT_CONFLICT")
        identity_missing = any(_source_identity_missing(row) for _, row in g.iterrows())
        if identity_missing:
            # Identity is incomplete, so this group has no right to reach value
            # conflict comparison. Rows remain auditable but require recovery.
            conflict_reasons.append("REVIEW_REQUIRED_SOURCE_IDENTITY")
        semantic_identity_unresolved = any(
            _identity_text(row.get("semantic_identity_status")) not in {
                "CERTIFIED", "LEGACY_IDENTITY_COMPATIBILITY",
            }
            for _, row in g.iterrows()
        )
        if semantic_identity_unresolved:
            conflict_reasons.append("REVIEW_REQUIRED_SEMANTIC_ROW_IDENTITY")
        # Same key with different physical columns but incomplete dimensions is
        # not proof of inconsistent values.  Preserve evidence and request a
        # human dimension review instead of presenting it as a hard block.
        physical_columns = g.get("column_ordinal", pd.Series(dtype=float)).dropna().nunique()
        missing_dimensions = (_dimension_missing(key[dims.index("period_identity")]) if "period_identity" in dims else False) or (_dimension_missing(key[dims.index("statement_scope")]) if "statement_scope" in dims else False)
        dimension_ambiguous = physical_columns > 1 and missing_dimensions
        if len(unique_values) > 1 and not identity_missing:
            conflict_reasons.append("REVIEW_REQUIRED_DIMENSION_AMBIGUITY" if dimension_ambiguous else "VALUE_CONFLICT")

        base = {col: val for col, val in zip(dims, key)}
        base["row_path"] = _identity_text(
            g["hierarchy_path"].iloc[0]
            if "hierarchy_path" in g.columns else
            g["row_path"].iloc[0]
            if "row_path" in g.columns else ""
        )
        base["data_year"] = (
            str(base.get("period_year"))
            if not _dimension_missing(base.get("period_year")) else ""
        )
        for column in PHYSICAL_BLOCK_LINEAGE_COLUMNS:
            values = list(dict.fromkeys(
                value
                for value in g[column].tolist()
                if _identity_text(value)
            ))
            if not values:
                base[column] = ""
            elif len(values) == 1:
                base[column] = values[0]
            else:
                base[column] = "MULTIPLE[" + "|".join(map(str, values)) + "]"
        base["source_identity_status"] = (
            "MISSING_MEMBER_TABLE_IDENTITY"
            if identity_missing else
            "SEMANTIC_ROW_IDENTITY_UNRESOLVED"
            if semantic_identity_unresolved else
            "SOURCE_IDENTITY_COMPLETE"
        )
        base["source_count"] = int(len(g))
        base["source_row_ids"] = "|".join(sorted({
            _identity_text(value)
            for value in g.get("source_row_id", pd.Series(dtype="object")).tolist()
            if _identity_text(value)
        }))
        base["mapping_status"] = " | ".join(sorted(set(g["mapping_status"].astype(str))))
        base["conflict_status"] = "|".join(conflict_reasons) if conflict_reasons else "OK"
        base["conflict_severity"] = (
            "WARNING" if conflict_reasons and not any(
                reason in {"VALUE_CONFLICT", "REVIEW_REQUIRED_SEMANTIC_ROW_IDENTITY"}
                for reason in conflict_reasons
            )
            else ("BLOCKING" if conflict_reasons else "OK")
        )
        base["dimension_review_required"] = bool(dimension_ambiguous)
        base["final_value"] = unique_values[0] if len(unique_values) == 1 and not conflict_reasons else None
        base["source_rows"] = " | ".join(
            f"{r.capture_run_id}:p{r.page}:{r.raw_item}"
            for r in g.itertuples()
        )
        base["source_provenance"] = json.dumps([
            {
                "capture_run_id": str(r.get("capture_run_id") or ""),
                "pdf": str(r.get("source_pdf") or r.get("pdf_name") or ""),
                "page": r.get("page"),
                "bbox": r.get("bbox"),
                "raw_item": str(r.get("raw_item") or ""),
                "normalized_item": str(r.get("normalized_item") or ""),
                "row_item_raw": str(r.get("row_item_raw") or ""),
                "row_item_normalized": str(r.get("row_item_normalized") or ""),
                "footnote_markers": r.get("footnote_markers"),
                "footnote_evidence": r.get("footnote_evidence"),
                "normalization_status": r.get("normalization_status"),
                "source_row_id": r.get("source_row_id"),
                "parent_row_id": r.get("parent_row_id"),
                "presentation_member_id": r.get("presentation_member_id"),
                "presentation_regime": r.get("presentation_regime"),
                "canonical_analysis_bucket": r.get("canonical_analysis_bucket"),
                "comparability_status": r.get("comparability_status"),
                "semantic_row_key": r.get("semantic_row_key"),
                "hierarchy_status": r.get("hierarchy_status"),
                "context_source_page": r.get("context_source_page"),
                "container_id": r.get("container_id"),
                "table_block_id": r.get("table_block_id"),
                "block_order": r.get("block_order"),
                "classification_axis": r.get("classification_axis"),
                "block_role": r.get("block_role"),
                "block_terminal_type": r.get("block_terminal_type"),
            }
            for _, r in g.iterrows()
        ], ensure_ascii=False)
        resolved_rows.append(base)

        if conflict_reasons:
            conflicts.append({
                **base,
                "values_found": " | ".join(map(str, unique_values)),
                "units_found": str(base.get("unit") or ""),
            })

    resolved = pd.DataFrame(resolved_rows)

    order_cols = [
        "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
        "order_source", "reference_capture_run_id", "reference_row_order",
    ]
    if structural_order is not None and not structural_order.empty:
        available_order_cols = [c for c in order_cols if c in structural_order.columns]
        resolved = resolved.merge(
            structural_order[available_order_cols].drop_duplicates("canonical_key"),
            on="canonical_key",
            how="left",
        )
        resolved = resolved.sort_values(
            ["canonical_order", "company", "report_year", "period_identity"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)
    else:
        resolved["canonical_order"] = range(1, len(resolved) + 1)
        resolved["row_type"] = ""
        resolved["order_source"] = "UNSPECIFIED"

    conflict_columns = list(resolved.columns) + ["values_found", "units_found"]
    conflicts_df = pd.DataFrame(conflicts, columns=conflict_columns)

    safe = resolved[
        (resolved["conflict_status"] == "OK")
        & resolved["final_value"].notna()
    ].copy()
    if safe.empty:
        empty_wide = pd.DataFrame(
            columns=[
                *BLOCK_IDENTITY_COLUMNS,
                "canonical_key", "canonical_section", "canonical_item", "unit",
            ]
        )
        return resolved, empty_wide, conflicts_df

    safe["document_column"] = safe.apply(_dimension_label, axis=1)

    # Stable one-row-per-canonical-key metadata. Do NOT use multiple metadata
    # fields directly as pivot_table index with dropna=False, which can create a
    # Cartesian product.
    key_meta_cols = [
        "table_family", "member_table", "member_table_role", "row_path",
        *BLOCK_IDENTITY_COLUMNS,
        "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
        "canonical_section", "canonical_item", "order_source", "source_identity_status",
    ]
    semantic_meta_cols = [
        column for column in key_meta_cols
        if column not in PHYSICAL_BLOCK_LINEAGE_COLUMNS
    ]
    if structural_order is not None and not structural_order.empty:
        key_meta = (
            structural_order[[c for c in semantic_meta_cols if c in structural_order.columns]]
            .drop_duplicates(subset=["canonical_key"], keep="first")
        )
    else:
        key_meta = (
            safe[[c for c in semantic_meta_cols if c in safe.columns]]
            .drop_duplicates(subset=["canonical_key"], keep="first")
        )

    def combine_lineage(series: pd.Series) -> Any:
        values = list(dict.fromkeys(
            value for value in series.tolist() if _identity_text(value)
        ))
        if not values:
            return ""
        if len(values) == 1:
            return values[0]
        return "MULTIPLE[" + "|".join(map(str, values)) + "]"

    lineage_meta = (
        safe.groupby("canonical_key", sort=False, dropna=False)[
            PHYSICAL_BLOCK_LINEAGE_COLUMNS
        ]
        .agg(combine_lineage)
        .reset_index()
    )
    key_meta = key_meta.merge(lineage_meta, on="canonical_key", how="left")

    wide_key_columns = ["canonical_key"]

    wide_values = safe.pivot_table(
        index=wide_key_columns,
        columns="document_column",
        values="final_value",
        aggfunc="first",
        dropna=True,
    ).reset_index()
    wide_values.columns.name = None

    wide = (
        key_meta
        .merge(wide_values, on=wide_key_columns, how="left")
    )
    if "canonical_order" in wide.columns:
        wide = wide.sort_values("canonical_order", kind="stable", na_position="last").reset_index(drop=True)

    fixed = [
        "table_family", "member_table", "presentation_member_id",
        "presentation_regime", "canonical_analysis_bucket",
        "comparability_status", "member_table_role", "row_path",
        *BLOCK_IDENTITY_COLUMNS,
        "canonical_order", "row_type", "row_level", "parent_section", "canonical_section",
        "canonical_item", "canonical_key", "order_source", "source_identity_status",
    ]
    fixed = [c for c in fixed if c in wide.columns]
    wide = wide[fixed + [c for c in wide.columns if c not in fixed]]

    return resolved, wide, conflicts_df


def coverage_report(raw_long: pd.DataFrame, mapped_long: pd.DataFrame, conflicts: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for run_id, g in mapped_long[mapped_long["value"].notna()].groupby("capture_run_id", sort=False):
        total = len(g)
        confirmed = g["mapping_status"].isin(
            ["AUTO_TAXONOMY", "AUTO_EXACT_IDENTITY", "CONFIRMED", "CONFIRMED_OVERRIDE"]
        ).sum()
        preserved = (g["mapping_status"] == "UNMAPPED_PRESERVED").sum()
        rows.append({
            "capture_run_id": run_id,
            "company": g["company"].iloc[0] if len(g) else "",
            "document_year": g["document_year"].iloc[0] if len(g) else "",
            "numeric_source_rows": int(total),
            "mapped_or_exact_rows": int(confirmed),
            "unmapped_preserved_rows": int(preserved),
            "mapping_coverage": float(confirmed / total) if total else 0.0,
        })

    report = pd.DataFrame(rows)
    if not report.empty:
        report["project_conflict_count"] = int(len(conflicts))
    return report


def persist_confirmed_taxonomy(
    taxonomy_path: Path,
    table_id: str,
    mapping_queue: pd.DataFrame,
) -> dict[str, Any]:
    taxonomy = load_taxonomy(taxonomy_path)
    table = taxonomy.setdefault("tables", {}).setdefault(
        table_id,
        {"mappings": []},
    )
    existing = {
        (m.get("canonical_section", ""), m.get("canonical_item", "")): m
        for m in table.get("mappings", [])
    }

    accepted = {"CONFIRMED", "CONFIRMED_OVERRIDE", "AUTO_TAXONOMY"}
    written = 0
    for row in mapping_queue.to_dict("records"):
        if row.get("mapping_status") not in accepted:
            continue
        canonical_item = str(row.get("canonical_item") or "").strip()
        if not canonical_item:
            continue
        canonical_section = normalize_section(row.get("canonical_section"))
        key = (canonical_section, canonical_item)
        entry = existing.get(key)
        if entry is None:
            entry = {
                "canonical_section": canonical_section,
                "canonical_item": canonical_item,
                "category": str(row.get("category") or ""),
                "source_keys": [],
            }
            table["mappings"].append(entry)
            existing[key] = entry
        source_key = str(row.get("source_key"))
        if source_key not in entry["source_keys"]:
            entry["source_keys"].append(source_key)
            written += 1
        if row.get("category"):
            entry["category"] = str(row.get("category"))

    _atomic_json(Path(taxonomy_path), taxonomy)
    return {"written_source_keys": written, "taxonomy_path": str(taxonomy_path)}


def collect_merge_reconciliation(manifest: dict[str, Any]) -> pd.DataFrame:
    frames=[]
    for source in manifest.get("sources") or []:
        capture_dir=Path(str(source.get("capture_dir") or ""))
        path=capture_dir/"table_reconciliation_audit.csv"
        if not path.exists():
            try:
                from reconciliation import write_reconciliation_audit
                path=write_reconciliation_audit(capture_dir)
            except Exception:
                continue
        try:
            df=pd.read_csv(path)
        except Exception:
            continue
        if df.empty:
            continue
        df.insert(0,"capture_run_id",source.get("capture_run_id"))
        df.insert(1,"company",source.get("company"))
        df.insert(2,"document_year",source.get("document_year"))
        frames.append(df)
    return pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()


def source_identity_qa(raw_long: pd.DataFrame) -> pd.DataFrame:
    """Expose source provenance before any row/value comparison is attempted."""
    columns = SOURCE_IDENTITY_COLUMNS + [
        "member_table_order", "row_path_count", "source_identity_status", "qa_detail",
    ]
    if raw_long.empty:
        return pd.DataFrame(columns=columns)
    work = raw_long.copy()
    for field in SOURCE_IDENTITY_COLUMNS:
        if field not in work.columns:
            work[field] = ""
    if "member_table_order" not in work.columns:
        work["member_table_order"] = None
    rows = []
    for capture_id, group in work.groupby("capture_run_id", dropna=False, sort=False):
        first = group.iloc[0]
        missing = [field for field in ("table_family", "member_table", "member_table_role") if not _identity_text(first.get(field))]
        rows.append({
            **{field: first.get(field) for field in SOURCE_IDENTITY_COLUMNS},
            "capture_run_id": capture_id,
            "member_table_order": first.get("member_table_order"),
            "row_path_count": int(group.get("row_path", pd.Series(dtype=str)).dropna().astype(str).nunique()),
            "source_identity_status": "MISSING_MEMBER_TABLE_IDENTITY" if missing else "SOURCE_IDENTITY_COMPLETE",
            "qa_detail": ("缺少：" + "、".join(missing)) if missing else "可进入 source-aware row alignment",
        })
    return pd.DataFrame(rows, columns=columns)


RESEARCH_WIDE_FIXED_COLUMNS = ("member_table", "canonical_item")
# Canonical row_type values treated as summary/total rows in the user-facing
# wide workbook. Identification stays at the semantic layer; label text is
# never used to guess a total.
TOTAL_ROW_TYPES = frozenset({"TOTAL", "IMPLICIT_TOTAL", "SUBTOTAL"})
GRAND_TOTAL_ROW_TYPES = frozenset({"TOTAL", "IMPLICIT_TOTAL", "SUBTOTAL", "GRAND_TOTAL"})
PARENT_ROW_TYPES = frozenset({"PARENT", "PARENT_SECTION", "SECTION", "SECTION_HEADER"})
RESEARCH_WIDE_FIXED_COLUMN_LABELS = {
    "member_table": "附注表名",
    "canonical_item": "项目",
}
RESEARCH_WIDE_METADATA_DIMENSION_LABELS = {
    "company": "公司",
    "report_year": "报告年",
    "period_label": "期间",
    "statement_scope": "口径",
    "period_type": "期间类型",
    "currency": "币种",
    "currency_unit": "金额单位",
    "measure": "计量口径",
}


def build_research_wide_frame(research_wide: pd.DataFrame) -> pd.DataFrame:
    """Trim the research wide to row identity plus observation columns.

    The full presentation wide carries many fixed context columns (row_path,
    canonical_key, block identity, source title, ...).  For real research the
    user only needs the member, canonical item and values; each observation's
    multi-level header carries its period/company/scope/unit/measure contract.
    """
    value_ids = [
        column for column in research_wide.columns
        if str(column).startswith("COL_")
    ]
    fixed = [
        column for column in RESEARCH_WIDE_FIXED_COLUMNS
        if column in research_wide.columns
    ]
    return research_wide[fixed + value_ids].copy()


def write_presentation_wide_sheet(
    writer: Any,
    research_wide: pd.DataFrame,
    fixed_columns: list[Any],
    column_dimensions: pd.DataFrame,
    header_policy: VisibleHeaderDimensionPolicy,
    sheet_name: str = "canonical_wide",
    row_types: Optional[Sequence[str]] = None,
    row_levels: Optional[Sequence[Any]] = None,
) -> None:
    """Write the human-facing multi-level wide sheet without COL ids.

    ``row_types`` and ``row_levels`` carry semantic hierarchy hints for Excel
    visual rendering (native indentation, bolding, background fill, and outline
    grouping). The CSV data contract is untouched.
    """
    header_rows = max(1, len(header_policy.visible_header_dimensions))
    research_wide.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        header=False,
        startrow=header_rows + 2,
    )
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="研究宽表元数据")
    meta_text = "；".join(
        f"{RESEARCH_WIDE_METADATA_DIMENSION_LABELS.get(dimension, dimension)}={value}"
        for dimension, value in header_policy.metadata_values.items()
        if value
    )
    ws.cell(row=1, column=2, value=meta_text)
    start_col = len(fixed_columns) + 1
    last_col = len(fixed_columns) + len(column_dimensions)
    if last_col > 2:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    for column_index, column_name in enumerate(fixed_columns, start=1):
        ws.cell(
            row=2,
            column=column_index,
            value=RESEARCH_WIDE_FIXED_COLUMN_LABELS.get(
                str(column_name), str(column_name),
            ),
        )
        if header_rows > 1:
            ws.merge_cells(
                start_row=2,
                start_column=column_index,
                end_row=header_rows + 1,
                end_column=column_index,
            )
    dimension_runs: dict[int, list[tuple[int, int]]] = {}
    display_labels = {
        dimension: column_dimensions.get(
            f"display_{dimension}",
            pd.Series([""] * len(column_dimensions)),
        ).tolist()
        for dimension in header_policy.visible_header_dimensions
    }
    for level_index, dimension in enumerate(header_policy.visible_header_dimensions):
        row_index = level_index + 2
        labels = display_labels[dimension]
        for offset, value in enumerate(labels):
            ws.cell(row=row_index, column=start_col + offset, value=value)
        run_start = start_col
        dimension_runs[row_index] = []
        for offset in range(1, len(labels) + 1):
            parent_changed = any(
                display_labels[parent][offset] != display_labels[parent][offset - 1]
                for parent in header_policy.visible_header_dimensions[:level_index]
            ) if offset < len(labels) else False
            changed = (
                offset == len(labels)
                or labels[offset] != labels[offset - 1]
                or parent_changed
            )
            if changed:
                run_end = start_col + offset - 1
                dimension_runs[row_index].append((run_start, run_end))
                if run_end > run_start and labels[offset - 1]:
                    ws.merge_cells(
                        start_row=row_index,
                        start_column=run_start,
                        end_row=row_index,
                        end_column=run_end,
                    )
                run_start = start_col + offset

    header_end_row = header_rows + 1
    data_start_row = header_rows + 3
    title_fill = PatternFill("solid", fgColor="1F4E78")
    metadata_fill = PatternFill("solid", fgColor="EAF0F6")
    report_fill = PatternFill("solid", fgColor="D9E5F2")
    period_fill = PatternFill("solid", fgColor="EDF2F7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    parent_level0_fill = PatternFill("solid", fgColor="F2F4F8")
    thin_side = Side(style="thin", color="AAB7C4")
    medium_side = Side(style="medium", color="5B6573")
    double_side = Side(style="double", color="5B6573")

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_view.showOutlineSymbols = True

    ws.row_dimensions[1].height = 22
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).font = Font(
        name="等线", size=11, bold=True, color="FFFFFF",
    )
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center",
    )
    for column_index in range(2, last_col + 1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = metadata_fill
        cell.font = Font(name="等线", size=10, color="374151")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_index in range(2, header_end_row + 1):
        ws.row_dimensions[row_index].height = 22
        for column_index in range(1, last_col + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.fill = report_fill if row_index == 2 else period_fill
            cell.font = Font(name="等线", size=10.5, bold=True, color="1F2937")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            cell.border = Border(
                top=medium_side if row_index == 2 else thin_side,
                bottom=medium_side if row_index == header_end_row else thin_side,
                left=cell.border.left,
                right=cell.border.right,
            )

    for column_index in range(1, len(fixed_columns) + 1):
        for row_index in range(2, header_end_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = Border(
                left=medium_side,
                right=medium_side,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    report_row = next(
        (
            index + 2
            for index, dimension in enumerate(header_policy.visible_header_dimensions)
            if dimension == "report_year"
        ),
        2,
    )
    for group_start, group_end in dimension_runs.get(report_row, []):
        for row_index in range(report_row, header_end_row + 1):
            left_cell = ws.cell(row=row_index, column=group_start)
            left_cell.border = Border(
                left=medium_side,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=row_index, column=group_end)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=medium_side,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )

    data_year_row = next(
        (
            index + 2
            for index, dimension in enumerate(header_policy.visible_header_dimensions)
            if dimension == "data_year"
        ),
        None,
    )
    if data_year_row is not None:
        for group_start, group_end in dimension_runs.get(data_year_row, []):
            left_cell = ws.cell(row=data_year_row, column=group_start)
            left_cell.border = Border(
                left=medium_side,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=data_year_row, column=group_end)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=medium_side,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )

    # Direct portfolio measurement labels can exceed 30 CJK characters.
    # Keep the fixed identity columns readable without wrapping data rows.
    widths = (36, 48)
    for column_index, width in enumerate(widths[:len(fixed_columns)], start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width
    for column_index in range(start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 16

    row_types_list = list(row_types) if row_types is not None else None

    def _parse_level(val: Any) -> int:
        try:
            if val is None or pd.isna(val):
                return 0
            return max(0, int(val))
        except (ValueError, TypeError):
            return 0

    data_row_count = max(0, ws.max_row - data_start_row + 1)
    row_levels_list = (
        [_parse_level(lvl) for lvl in row_levels]
        if row_levels is not None
        else [0] * data_row_count
    )

    has_children_list = [False] * len(row_levels_list)
    for i in range(len(row_levels_list)):
        if i + 1 < len(row_levels_list) and row_levels_list[i + 1] > row_levels_list[i]:
            has_children_list[i] = True

    fixed_count = len(fixed_columns)
    for row_index in range(data_start_row, ws.max_row + 1):
        row_offset = row_index - data_start_row
        row_type = (
            str(row_types_list[row_offset]).upper()
            if row_types_list is not None and row_offset < len(row_types_list)
            else ""
        )
        row_level = (
            row_levels_list[row_offset]
            if row_offset < len(row_levels_list)
            else 0
        )
        has_children = (
            has_children_list[row_offset]
            if row_offset < len(has_children_list)
            else False
        )

        ws.row_dimensions[row_index].outlineLevel = row_level

        is_grand_total = row_type in GRAND_TOTAL_ROW_TYPES
        is_parent = not is_grand_total and (has_children or row_type in PARENT_ROW_TYPES)

        visual_indent = min(row_level, 4)

        if is_grand_total:
            row_font = Font(name="等线", size=10.5, bold=True, color="1F2937")
            row_fill = total_fill
            row_border = Border(top=thin_side, bottom=double_side)
        elif is_parent:
            row_font = Font(name="等线", size=10.5, bold=True, color="1F2937")
            row_fill = parent_level0_fill if row_level == 0 else None
            row_border = Border(top=thin_side, bottom=thin_side) if row_level == 0 else None
        else:
            row_font = Font(name="等线", size=10.5, bold=False, color="1F2937")
            row_fill = None
            row_border = None

        for column_index in range(1, last_col + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if column_index <= fixed_count:
                col_name = str(fixed_columns[column_index - 1]) if column_index - 1 < len(fixed_columns) else ""
                indent_val = (
                    visual_indent
                    if col_name in ("canonical_item", "项目") or (column_index == fixed_count and "canonical_item" not in fixed_columns)
                    else 0
                )
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent_val)
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'

            cell.font = row_font
            if row_fill is not None:
                cell.fill = row_fill
            if row_border is not None:
                cell.border = row_border

    ws.freeze_panes = f"{get_column_letter(start_col)}{data_start_row}"


def write_merge_outputs(
    output_dir: Path,
    manifest: dict[str, Any],
    raw_long: pd.DataFrame,
    mapping_queue: pd.DataFrame,
    taxonomy_path: Optional[Path] = None,
    member_display_map: Optional[dict[str, str]] = None,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)

    # v5.8 invariant: canonicalization never sees relative labels in `year`.
    # This also repairs already-created v5.7 merge_raw_long.csv on refresh.
    raw_long, manifest = _repair_manifest_and_raw_periods(raw_long, manifest)

    mapped = apply_mapping(raw_long, mapping_queue)
    structural_order, order_conflicts = build_structural_order(mapped, manifest)

    if not structural_order.empty:
        mapped = mapped.merge(
            structural_order[
                [
                    "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
                    "order_source", "reference_capture_run_id", "reference_row_order",
                ]
            ].drop_duplicates("canonical_key"),
            on="canonical_key",
            how="left",
            suffixes=("", "_order"),
        )
        # Preserve source evidence ordering inside each capture, while exposing
        # canonical_order for cross-company research views.
        mapped = mapped.sort_values(
            ["capture_run_id", "row_order", "canonical_order"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    resolved, wide, conflicts = materialize_canonical(
        mapped,
        structural_order=structural_order,
    )
    (
        financial_original,
        financial_bridge,
        financial_bridge_wide,
        financial_bridge_audit,
    ) = project_financial_investment_views(resolved)
    manifest["financial_investment_standards_bridge"] = {
        "schema_version": BRIDGE_SCHEMA_VERSION,
        "delivery_policy": "DUAL_VIEW_SOURCE_PRESENTATION_AND_EXPLICIT_BRIDGE",
        "original_row_count": int(len(financial_original)),
        "bridge_row_count": int(len(financial_bridge)),
        "bridge_value_count": int(
            financial_bridge["final_value"].notna().sum()
            if "final_value" in financial_bridge.columns else 0
        ),
        "audit_row_count": int(len(financial_bridge_audit)),
        "no_same_period_sum": True,
    }
    coverage = coverage_report(raw_long, mapped, conflicts)
    reconciliation = collect_merge_reconciliation(manifest)
    identity_qa = source_identity_qa(raw_long)
    precision_audit = period_precision_audit(mapped)

    # Canonical Long is the source of truth. CSV cannot represent a real
    # hierarchical header, so its wide view uses stable encoded ids and an
    # explicit dimension mapping rather than a lossy concatenated label.
    fixed_columns = [c for c in wide.columns if not str(c).startswith("company=")]
    value_columns = [c for c in wide.columns if str(c).startswith("company=")]
    column_rows = []
    rename_columns = {}
    for ordinal, column in enumerate(value_columns, start=1):
        column_id = f"COL_{ordinal:05d}"
        rename_columns[column] = column_id
        dims = {}
        for token in str(column).split(" | "):
            if "=" in token:
                key, value = token.split("=", 1); dims[key] = value
        # Currency is retained independently even when the older v6.6
        # presentation label only carried currency_unit.
        dims.setdefault("currency", "CNY" if str(dims.get("currency_unit") or "").startswith("CNY") else "")
        column_rows.append({"column_id": column_id, "source_column_label": column, **{key: dims.get(key, "") for key in OBSERVATION_DIMENSIONS}})
    research_wide = wide.rename(columns=rename_columns)
    column_dimensions = pd.DataFrame(column_rows)
    header_policy = VisibleHeaderDimensionPolicy.from_column_dimensions(column_dimensions)
    if not column_dimensions.empty:
        display_rows = [header_policy.label_for_column(row) for row in column_dimensions.to_dict("records")]
        for dimension in header_policy.visible_header_dimensions:
            column_dimensions[f"display_{dimension}"] = [row.get(dimension, "") for row in display_rows]
    # Research-wide export: only member/canonical item plus values, with
    # the same multi-level header architecture as the presentation export.
    research_wide_trim = build_research_wide_frame(research_wide)
    if member_display_map:
        research_wide_trim["member_table"] = research_wide_trim[
            "member_table"
        ].map(
            lambda value: member_display_map.get(str(value), value)
        )
    research_wide_fixed = [
        column for column in RESEARCH_WIDE_FIXED_COLUMNS
        if column in research_wide_trim.columns
    ]
    # row_type and row_level are styling-only hints for the Excel writer;
    # they never enter the trimmed research CSV.
    research_row_types = (
        research_wide["row_type"].tolist()
        if "row_type" in research_wide.columns
        else [""] * len(research_wide)
    )
    research_row_levels = (
        research_wide["row_level"].tolist()
        if "row_level" in research_wide.columns
        else [0] * len(research_wide)
    )

    paths = {
        "manifest": output_dir / "merge_manifest.json",
        "raw_long": output_dir / "merge_raw_long.csv",
        "mapping_queue": output_dir / "merge_mapping_queue.csv",
        "canonical_long": output_dir / "merge_canonical_long.csv",
        "canonical_research_long": output_dir / "canonical_research_long.csv",
        "resolved_long": output_dir / "merge_resolved_long.csv",
        "canonical_wide": output_dir / "merge_canonical_wide.csv",
        "column_dimensions": output_dir / "column_dimensions.csv",
        "wide_metadata": output_dir / "research_wide_metadata.json",
        "research_wide_csv": output_dir / "research_wide.csv",
        "research_wide_xlsx": output_dir / "research_wide.xlsx",
        "conflicts": output_dir / "merge_conflicts.csv",
        "coverage": output_dir / "merge_coverage.csv",
        "structural_order": output_dir / "merge_structural_order.csv",
        "order_conflicts": output_dir / "merge_order_conflicts.csv",
        "reconciliation": output_dir / "merge_reconciliation_audit.csv",
        "source_identity_qa": output_dir / "merge_source_identity_qa.csv",
        "period_precision_audit": output_dir / "merge_period_precision_audit.csv",
        "financial_original": output_dir / "financial_investment_original_long.csv",
        "financial_bridge": output_dir / "financial_investment_standards_bridge_long.csv",
        "financial_bridge_wide": output_dir / "financial_investment_standards_bridge_wide.csv",
        "financial_bridge_audit": output_dir / "financial_investment_standards_bridge_audit.csv",
        "xlsx": output_dir / "merge_project.xlsx",
    }

    paths["manifest"].write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    raw_long.to_csv(paths["raw_long"], index=False, encoding="utf-8-sig")
    mapping_queue.to_csv(paths["mapping_queue"], index=False, encoding="utf-8-sig")
    mapped.to_csv(paths["canonical_long"], index=False, encoding="utf-8-sig")
    # Canonical Research Long is the source-of-truth observation layer. It
    # retains source rows and provenance; wide/resolved outputs are views.
    mapped.to_csv(paths["canonical_research_long"], index=False, encoding="utf-8-sig")
    resolved.to_csv(paths["resolved_long"], index=False, encoding="utf-8-sig")
    research_wide.to_csv(paths["canonical_wide"], index=False, encoding="utf-8-sig")
    research_wide_trim.to_csv(
        paths["research_wide_csv"], index=False, encoding="utf-8-sig",
    )
    column_dimensions.to_csv(paths["column_dimensions"], index=False, encoding="utf-8-sig")
    paths["wide_metadata"].write_text(json.dumps({"metadata_dimensions": list(header_policy.metadata_dimensions), "visible_header_dimensions": list(header_policy.visible_header_dimensions), "display_order": list(header_policy.display_order), "metadata_values": header_policy.metadata_values, "presentation_export_version": 4, "fixed_columns": ["member_table", "canonical_item"], "unit_contract": "OBSERVATION_COLUMN_CURRENCY_UNIT_AND_MEASURE", "period_contract": "POINT_PERIOD_IDENTITY_V4"}, ensure_ascii=False, indent=2), encoding="utf-8")
    conflicts.to_csv(paths["conflicts"], index=False, encoding="utf-8-sig")
    coverage.to_csv(paths["coverage"], index=False, encoding="utf-8-sig")
    structural_order.to_csv(paths["structural_order"], index=False, encoding="utf-8-sig")
    order_conflicts.to_csv(paths["order_conflicts"], index=False, encoding="utf-8-sig")
    reconciliation.to_csv(paths["reconciliation"], index=False, encoding="utf-8-sig")
    identity_qa.to_csv(paths["source_identity_qa"], index=False, encoding="utf-8-sig")
    precision_audit.to_csv(
        paths["period_precision_audit"], index=False, encoding="utf-8-sig",
    )
    financial_original.to_csv(paths["financial_original"], index=False, encoding="utf-8-sig")
    financial_bridge.to_csv(paths["financial_bridge"], index=False, encoding="utf-8-sig")
    financial_bridge_wide.to_csv(paths["financial_bridge_wide"], index=False, encoding="utf-8-sig")
    financial_bridge_audit.to_csv(paths["financial_bridge_audit"], index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(paths["xlsx"], engine="openpyxl") as writer:
        raw_long.to_excel(writer, sheet_name="raw_long", index=False)
        mapping_queue.to_excel(writer, sheet_name="mapping_queue", index=False)
        mapped.to_excel(writer, sheet_name="canonical_long", index=False)
        mapped.to_excel(writer, sheet_name="canonical_research_long", index=False)
        resolved.to_excel(writer, sheet_name="resolved_long", index=False)
        precision_audit.to_excel(writer, sheet_name="period_precision_audit", index=False)
        write_presentation_wide_sheet(
            writer,
            research_wide,
            fixed_columns,
            column_dimensions,
            header_policy,
            row_types=research_row_types,
            row_levels=research_row_levels,
        )
        column_dimensions.to_excel(writer, sheet_name="column_dimensions", index=False)
        conflicts.to_excel(writer, sheet_name="conflicts", index=False)
        coverage.to_excel(writer, sheet_name="coverage", index=False)
        structural_order.to_excel(writer, sheet_name="structural_order", index=False)
        order_conflicts.to_excel(writer, sheet_name="order_conflicts", index=False)
        reconciliation.to_excel(writer, sheet_name="reconciliation", index=False)
        identity_qa.to_excel(writer, sheet_name="source_identity_qa", index=False)
        financial_original.to_excel(writer, sheet_name="金融投资_原始口径", index=False)
        financial_bridge_wide.to_excel(writer, sheet_name="金融投资_跨准则桥接", index=False)
        financial_bridge_audit.to_excel(writer, sheet_name="金融投资_桥接审计", index=False)

    if research_wide_fixed:
        with pd.ExcelWriter(
            paths["research_wide_xlsx"], engine="openpyxl",
        ) as writer:
            write_presentation_wide_sheet(
                writer,
                research_wide_trim,
                research_wide_fixed,
                column_dimensions,
                header_policy,
                sheet_name="research_wide",
                row_types=research_row_types,
                row_levels=research_row_levels,
            )
            financial_original.to_excel(writer, sheet_name="金融投资_原始口径", index=False)
            financial_bridge_wide.to_excel(writer, sheet_name="金融投资_跨准则桥接", index=False)
            financial_bridge_audit.to_excel(writer, sheet_name="金融投资_桥接审计", index=False)

    if taxonomy_path and Path(taxonomy_path).exists():
        snapshot = output_dir / "taxonomy_snapshot.json"
        snapshot.write_text(Path(taxonomy_path).read_text(encoding="utf-8"), encoding="utf-8")
        paths["taxonomy_snapshot"] = snapshot

    try:
        from registry_bridge import sync_merge_run
        sync_merge_run(output_dir)
    except Exception:
        pass

    return {k: str(v) for k, v in paths.items()}


def create_merge_project(
    capture_dirs: list[Path],
    metadata_rows: list[dict[str, Any]],
    output_dir: Path,
    table_id: str,
    taxonomy_path: Path,
    reference_capture_run_id: Optional[str] = None,
    merge_lineage: Optional[dict[str, Any]] = None,
    member_display_map: Optional[dict[str, str]] = None,
    order_policy: Optional[str] = None,
    reference_report_year: Optional[str] = None,
) -> dict[str, str]:
    table_id = normalize_table_id(table_id)
    metadata_map = {str(r["capture_run_id"]): r for r in metadata_rows}

    frames = []
    manifest_sources = []
    for capture_dir in capture_dirs:
        inferred = infer_capture_metadata(capture_dir)
        user_meta = metadata_map.get(capture_dir.name, {})
        meta = {**inferred, **user_meta}
        frame = load_capture_long(capture_dir, meta, table_id)
        frames.append(frame)
        manifest_sources.append(meta)

    raw_long = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    taxonomy = load_taxonomy(taxonomy_path)
    queue = build_mapping_queue(raw_long, table_id, taxonomy)

    source_ids = [str(x.get("capture_run_id")) for x in manifest_sources]
    if reference_capture_run_id not in source_ids:
        reference_capture_run_id = source_ids[0] if source_ids else None

    bundle_expansion = dict(merge_lineage or {})
    if bundle_expansion:
        applied_exclusions = [
            dict(exclusion)
            for source in manifest_sources
            for exclusion in (source.get("merge_row_exclusions_applied") or [])
        ]
        expected_count = int(bundle_expansion.get("row_cell_exclusion_count") or 0)
        if len(applied_exclusions) != expected_count:
            raise ValueError(
                "MERGE_ROW_EXCLUSION_COUNT_MISMATCH:"
                f"expected={expected_count}:applied={len(applied_exclusions)}"
            )
        bundle_expansion["row_cell_exclusions_applied"] = applied_exclusions
        bundle_expansion["row_cell_exclusions_applied_count"] = len(applied_exclusions)

    manifest = {
        "version": "v6.8",
        "merge_schema_version": "6.9_FINANCIAL_PRESENTATION_REGIME_DUAL_VIEW",
        "canonical_observation_schema_version": "6.9_PRESENTATION_MEMBER_REGIME_LINEAGE",
        "table_id": table_id,
        "sources": manifest_sources,
        "taxonomy_path": str(taxonomy_path),
        "order_policy": "REFERENCE_CAPTURE_PRESERVE_WITH_CONTEXTUAL_INSERTION",
        "identity_policy": "TABLE_FAMILY__PRESENTATION_MEMBER__PRESENTATION_REGIME__MEMBER_ROLE__CLASSIFICATION_AXIS__ROW_PATH__DIMENSIONS__PHYSICAL_BLOCK_LINEAGE_ONLY",
        "reference_capture_run_id": reference_capture_run_id,
    }
    if member_display_map:
        manifest["member_display_map"] = dict(member_display_map)
    if order_policy:
        manifest["order_policy"] = str(order_policy)
    if reference_report_year:
        manifest["reference_report_year"] = str(reference_report_year)
    if bundle_expansion:
        manifest["bundle_expansion"] = bundle_expansion
    return write_merge_outputs(
        output_dir=output_dir,
        manifest=manifest,
        raw_long=raw_long,
        mapping_queue=queue,
        taxonomy_path=taxonomy_path,
        member_display_map=member_display_map,
    )


def refresh_merge_project(
    output_dir: Path,
    mapping_queue: Optional[pd.DataFrame] = None,
    persist_taxonomy: bool = False,
    order_policy: Optional[str] = None,
    reference_report_year: Optional[str] = None,
    member_display_map: Optional[dict[str, str]] = None,
) -> dict[str, str]:
    output_dir = Path(output_dir)
    manifest = json.loads((output_dir / "merge_manifest.json").read_text(encoding="utf-8"))
    # A refresh is also the non-destructive migration path for old derived
    # merge artifacts. Source capture evidence and human mapping decisions are
    # retained; only generated outputs are recomputed under the current
    # observation contract.
    manifest["merge_schema_version"] = "6.9_FINANCIAL_PRESENTATION_REGIME_DUAL_VIEW"
    manifest["canonical_observation_schema_version"] = "6.9_PRESENTATION_MEMBER_REGIME_LINEAGE"
    manifest["identity_policy"] = "TABLE_FAMILY__PRESENTATION_MEMBER__PRESENTATION_REGIME__MEMBER_ROLE__CLASSIFICATION_AXIS__ROW_PATH__DIMENSIONS__PHYSICAL_BLOCK_LINEAGE_ONLY"
    if order_policy:
        manifest["order_policy"] = str(order_policy)
    if reference_report_year is not None:
        manifest["reference_report_year"] = str(reference_report_year)
    if member_display_map is not None:
        manifest["member_display_map"] = dict(member_display_map)
    raw = pd.read_csv(output_dir / "merge_raw_long.csv")
    queue = (
        mapping_queue.copy()
        if mapping_queue is not None
        else pd.read_csv(output_dir / "merge_mapping_queue.csv")
    )
    taxonomy_path = Path(manifest.get("taxonomy_path", output_dir / "table_taxonomy.json"))

    if persist_taxonomy:
        persist_confirmed_taxonomy(taxonomy_path, manifest["table_id"], queue)

    return write_merge_outputs(
        output_dir=output_dir,
        manifest=manifest,
        raw_long=raw,
        mapping_queue=queue,
        taxonomy_path=taxonomy_path,
        member_display_map=(
            member_display_map
            if member_display_map is not None
            else manifest.get("member_display_map") or None
        ),
    )

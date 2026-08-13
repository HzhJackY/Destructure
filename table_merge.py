#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
table_merge.py — v5.1 multi-company / multi-year table merge engine.

Safety model:
- raw capture rows are immutable evidence.
- exact normalized identities may auto-align.
- non-identical labels are NEVER silently merged.
- fuzzy similarity is suggestion-only.
- confirmed mappings can be persisted to a table taxonomy.
- duplicate/conflicting canonical keys are surfaced, not silently overwritten.
"""

from __future__ import annotations

import difflib
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Optional, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from batch_pipeline import infer_company_year
from financial_metric_pdf_resolver import normalize_text
from financial_structure_resolver import ensure_row_paths
from visible_header_policy import VisibleHeaderDimensionPolicy, OBSERVATION_DIMENSIONS


TAXONOMY_VERSION = 1

# A Family Merge observation is source-aware before it is row-aware.  These
# fields intentionally survive all long/canonical/wide materializations.
SOURCE_IDENTITY_COLUMNS = [
    "table_family", "member_table", "member_table_role",
    "source_table_title", "note_reference", "capture_run_id", "source_pdf",
]
BLOCK_IDENTITY_COLUMNS = [
    "container_id", "table_block_id", "block_order",
    "classification_axis", "block_role", "block_terminal_type",
]
PHYSICAL_BLOCK_LINEAGE_COLUMNS = [
    "container_id", "table_block_id", "block_order",
    "block_role", "block_terminal_type",
]
CANONICAL_BLOCK_IDENTITY_COLUMNS = ["classification_axis"]
SOURCE_IDENTITY_COLUMNS += BLOCK_IDENTITY_COLUMNS


def _identity_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "<na>"} else text


def _source_identity_missing(row: Any) -> bool:
    return not all(_identity_text(row.get(field)) for field in (
        "table_family", "member_table", "member_table_role",
    ))


def _semantic_block_scope(row: Any) -> str:
    """Use a stable semantic axis across Captures; isolate unresolved blocks."""
    axis = _identity_text(row.get("classification_axis"))
    if axis and axis != "UNRESOLVED":
        return f"AXIS::{axis}||"
    block_id = _identity_text(row.get("table_block_id"))
    return f"BLOCK::{block_id}||AXIS::UNRESOLVED||" if block_id else ""


def normalize_table_id(text: str) -> str:
    s = normalize_text(str(text or ""))
    return s or "UNNAMED_TABLE"


def normalize_section(text: Any) -> str:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    return re.sub(r"\s+", "", str(text)).strip("：:")


def source_mapping_key(parent_section: Any, normalized_item: Any) -> str:
    """Legacy key retained for taxonomy compatibility."""
    return f"{normalize_section(parent_section)}||{str(normalized_item or '').strip()}"


def assign_conditional_source_keys(df: pd.DataFrame) -> pd.DataFrame:
    """
    v5.6 identity policy:
    - if an item name appears only once in a source table, identity is the item
      name itself; parent_section is context, not a primary key.
    - only when the same normalized_item appears multiple times in the SAME
      source table do we activate contextual disambiguation using parent_section,
      row_type, and local occurrence order.

    This prevents harmless parent-section parsing differences across companies
    from blocking exact-name merges.
    """
    out = ensure_row_paths(df)
    out["source_key"] = ""
    if out.empty:
        return out

    # One structural record per original row, ignoring repeated period/value rows.
    structural_columns = [
        "row_order", "normalized_item", "parent_section", "row_type",
        *[column for column in BLOCK_IDENTITY_COLUMNS if column in out.columns],
    ]
    structural = (
        out[structural_columns]
        .drop_duplicates(subset=["row_order"], keep="first")
        .copy()
    )
    structural["normalized_item"] = structural["normalized_item"].fillna("").astype(str).str.strip()
    structural["_row_order_num"] = pd.to_numeric(structural["row_order"], errors="coerce")
    structural = structural.sort_values("_row_order_num", kind="stable")

    for item, g in structural.groupby("normalized_item", sort=False):
        if not item:
            continue
        rows = g["row_order"].tolist()
        if len(rows) == 1:
            record = g.iloc[0]
            block_scope = _semantic_block_scope(record)
            key = f"{block_scope}UNIQUE||{item}"
            out.loc[out["row_order"].isin(rows), "source_key"] = key
            continue

        # Same name occurs more than once in the same source: context is now
        # required. Occurrence index prevents accidental collapse even if two
        # repeated rows also share the same parent text.
        for occurrence, (_, r) in enumerate(g.iterrows(), start=1):
            parent = normalize_section(r.get("parent_section"))
            row_path = str(r.get("row_path") or "").strip()
            row_type = str(r.get("row_type") or "").strip()
            block_scope = _semantic_block_scope(r)
            # v6.2: repeated names are identities only with their full derived
            # row path.  This prevents e.g. two "交易性金融资产" rows under
            # 股息收入 and 利息收入 from collapsing.
            key = f"{block_scope}CONTEXT_PATH||{item}||{row_path or parent}||{row_type}||OCC{occurrence}"
            out.loc[out["row_order"] == r["row_order"], "source_key"] = key

    # Fallback for unusual rows without normalized item.
    missing = out["source_key"].astype(str).str.len() == 0
    out.loc[missing, "source_key"] = out.loc[missing].apply(
        lambda r: (
            _semantic_block_scope(r)
            + f"LEGACY||{source_mapping_key(r.get('parent_section'), r.get('normalized_item'))}"
        ),
        axis=1,
    )
    return out


def _atomic_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)


def load_taxonomy(path: Path) -> dict[str, Any]:
    path = Path(path)
    if not path.exists():
        return {"version": TAXONOMY_VERSION, "tables": {}}
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("version", TAXONOMY_VERSION)
    data.setdefault("tables", {})
    return data


def _taxonomy_lookup(taxonomy: dict[str, Any], table_id: str) -> dict[str, dict[str, Any]]:
    table = taxonomy.get("tables", {}).get(table_id, {})
    mappings = table.get("mappings", [])
    lookup: dict[str, dict[str, Any]] = {}

    # Exact stored keys.
    for m in mappings:
        for alias_key in m.get("source_keys", []):
            lookup[str(alias_key)] = m

    # Backward compatibility: old versions stored parent_section||item. For a
    # v5.6 UNIQUE||item key, reuse an old mapping only if all legacy mappings for
    # that item point to the same canonical target.
    by_item: dict[str, list[dict[str, Any]]] = {}
    for m in mappings:
        for alias_key in m.get("source_keys", []):
            key = str(alias_key)
            if key.startswith("UNIQUE||"):
                item = key.split("||",1)[1]
            elif "||" in key and not key.startswith("CONTEXT||"):
                item = key.split("||")[-1]
            else:
                continue
            by_item.setdefault(item, []).append(m)

    for item, candidates in by_item.items():
        targets = {
            (
                str(m.get("canonical_section") or ""),
                str(m.get("canonical_item") or ""),
            )
            for m in candidates
        }
        if len(targets) == 1:
            lookup.setdefault(f"UNIQUE||{item}", candidates[0])

    return lookup

def infer_capture_metadata(capture_dir: Path) -> dict[str, Any]:
    result_path = capture_dir / "table_capture_result.json"
    data = json.loads(result_path.read_text(encoding="utf-8"))
    pdf_name = data.get("pdf_name", "")
    company, filename_year = infer_company_year(Path(pdf_name), "")

    # v5.8: report/document_year must be an absolute four-digit year.
    # Relative column labels such as 本年累计数 / 上年累计数 / 去年累计数
    # are never valid document identity.
    absolute_candidates = []
    if re.fullmatch(r"20\d{2}", str(filename_year or "").strip()):
        absolute_candidates.append(str(filename_year).strip())

    for c in data.get("columns", []) or []:
        for field in ["year", "period_label", "header_raw"]:
            value = str(c.get(field) or "").strip()
            m = re.search(r"(20\d{2})", value)
            if m:
                absolute_candidates.append(m.group(1))

    document_year = max(absolute_candidates) if absolute_candidates else ""
    capture_meta_path = capture_dir / "capture_metadata.json"
    try:
        capture_meta = json.loads(capture_meta_path.read_text(encoding="utf-8")) if capture_meta_path.exists() else {}
    except (OSError, json.JSONDecodeError):
        capture_meta = {}
    stats = data.get("stats") or {}
    return {
        "capture_run_id": capture_dir.name,
        "capture_dir": str(capture_dir),
        "pdf_name": pdf_name,
        "pdf_sha256": data.get("pdf_sha256"),
        "company": company,
        "document_year": document_year,
        "table_query": data.get("table_query", ""),
        "note_number": data.get("note_number"),
        "located_title": data.get("located_title", ""),
        "table_family": capture_meta.get("table_family"),
        "member_table": capture_meta.get("member_table"),
        "member_table_role": capture_meta.get("member_table_role"),
        "source_table_title": capture_meta.get("source_table_title"),
        "note_reference": capture_meta.get("note_reference"),
        "source_pdf": capture_meta.get("source_pdf_path") or stats.get("source_pdf_path") or pdf_name,
        "member_table_order": capture_meta.get("member_table_order"),
    }


CURRENT_RELATIVE_PERIOD_TOKENS = {
    "本年累计数", "本年度累计数", "本期累计数", "本期数", "本年数",
    "本期", "本年", "本年度", "当期累计数", "当期",
    "期末", "年末", "本期期末", "本年末", "本年度末",
}
PRIOR_RELATIVE_PERIOD_TOKENS = {
    "上年累计数", "上年度累计数", "上期累计数", "上期数", "上年数",
    "上期", "上年", "上年度", "去年", "去年累计数", "去年数",
    "去年同期", "上年同期", "上年度同期", "上期同期",
}


def _is_missing_scalar(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _clean_scalar_text(value: Any) -> str:
    if _is_missing_scalar(value):
        return ""
    return str(value).strip()


def _clean_period_token(value: Any) -> str:
    token = _clean_scalar_text(value)
    if re.fullmatch(r"20\d{2}\.0", token):
        return token[:4]
    token = re.sub(r"\s+", "", token)
    token = token.replace("（", "(").replace("）", ")")
    token = re.sub(r"\(?(?:已重述|经重述|重述后|重述)\)?", "", token)
    token = re.sub(r"(?:人民币)?(?:亿元|百万元|万元|千元|元)$", "", token)
    return token.strip()


def _absolute_year(value: Any) -> Optional[str]:
    token = _clean_period_token(value)
    # CSV round trips commonly turn a year into `2022.0`.  It remains an
    # absolute reporting period, not a distinct observation dimension.
    m = re.fullmatch(r"(20\d{2})(?:\.0)?(?:年度|年)?", token)
    return m.group(1) if m else None


def _relative_period_kind(value: Any) -> Optional[str]:
    token = _clean_period_token(value)
    if token in CURRENT_RELATIVE_PERIOD_TOKENS:
        return "CURRENT"
    if token in PRIOR_RELATIVE_PERIOD_TOKENS:
        return "PRIOR"
    return None


def _infer_document_year_from_capture_source(
    *,
    pdf_name: Any = "",
    capture_dir: Any = "",
    columns: Optional[list[dict[str, Any]]] = None,
) -> str:
    """Infer report year from absolute evidence only."""
    _, filename_year = infer_company_year(Path(str(pdf_name or "")), "")
    y = _absolute_year(filename_year)
    if y:
        return y

    absolute_candidates: list[str] = []
    for c in columns or []:
        for field in ["year", "period_label", "header_raw"]:
            y = _absolute_year(c.get(field))
            if y:
                absolute_candidates.append(y)
    if absolute_candidates:
        return max(absolute_candidates)

    cap = Path(str(capture_dir or ""))
    result_path = cap / "table_capture_result.json"
    if result_path.exists():
        try:
            data = json.loads(result_path.read_text(encoding="utf-8"))
            _, filename_year = infer_company_year(Path(str(data.get("pdf_name") or "")), "")
            y = _absolute_year(filename_year)
            if y:
                return y
            for c in data.get("columns") or []:
                for field in ["year", "period_label", "header_raw"]:
                    y = _absolute_year(c.get(field))
                    if y:
                        absolute_candidates.append(y)
            if absolute_candidates:
                return max(absolute_candidates)
        except Exception:
            pass
    return ""


def _resolve_source_document_year(source: dict[str, Any]) -> str:
    existing = _absolute_year(source.get("document_year"))
    if existing:
        return existing
    return _infer_document_year_from_capture_source(
        pdf_name=source.get("pdf_name"),
        capture_dir=source.get("capture_dir"),
    )


def _resolve_relative_period_years(
    df: pd.DataFrame,
    document_year: Any,
) -> pd.DataFrame:
    """
    Convert relative period labels to actual years before canonical merge.

    document_year=2023:
      本年 / 本年累计数 / 本期累计数 -> 2023
      去年 / 去年累计数 / 上年累计数 / 上期累计数 -> 2022

    Original wording is retained in `source_period_label`.
    """
    out = df.copy()
    if "year" not in out.columns:
        return out

    doc = _absolute_year(document_year)

    if "source_period_label" not in out.columns:
        if "period_label" in out.columns:
            out["source_period_label"] = out["period_label"]
        else:
            out["source_period_label"] = out["year"]

    def resolve(value: Any) -> Any:
        absolute = _absolute_year(value)
        if absolute:
            return absolute
        kind = _relative_period_kind(value)
        if kind is None or doc is None:
            return value
        return doc if kind == "CURRENT" else str(int(doc) - 1)

    out["year"] = out["year"].map(resolve)

    if "column_dimension_key" in out.columns:
        def dimension_key(r: pd.Series) -> Any:
            if _is_missing_scalar(r.get("column_ordinal")):
                return r.get("column_dimension_key")
            restated_raw = r.get("restated")
            restated = False if _is_missing_scalar(restated_raw) else bool(restated_raw)
            return (
                f"{_clean_scalar_text(r.get('year'))}|"
                f"{_clean_scalar_text(r.get('scope'))}|"
                f"{'RESTATED' if restated else 'ORIGINAL'}"
            )

        out["column_dimension_key"] = out.apply(dimension_key, axis=1)
    return out


def _validate_absolute_year_resolution(
    df: pd.DataFrame,
    *,
    document_year: Any,
    capture_run_id: str,
) -> None:
    if "year" not in df.columns:
        return

    relative_values = sorted({
        str(v)
        for v in df["year"].dropna().tolist()
        if _relative_period_kind(v)
    })
    if not relative_values:
        return

    doc = _absolute_year(document_year)
    if not doc:
        raise ValueError(
            f"PERIOD_RESOLUTION_REQUIRED：来源 {capture_run_id} 包含相对期间 "
            f"{relative_values}，但报告年份 document_year 未识别为四位年份。"
            " 请在合表来源信息中填写该年报实际年份，例如 2023。"
            " 系统随后会自动将本年→2023、去年/上年→2022。"
        )

    raise ValueError(
        f"PERIOD_RESOLUTION_INVARIANT_VIOLATION：来源 {capture_run_id} "
        f"在 document_year={doc} 已知时仍残留相对 year={relative_values}。"
        " 已阻止其进入 canonical 合表。"
    )


def _repair_manifest_and_raw_periods(
    raw_long: pd.DataFrame,
    manifest: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """
    Repair existing v5.7 Merge Projects on rematerialization as well as new ones.
    """
    raw = raw_long.copy()
    # Legacy CSV projects commonly deserialize period identity columns as
    # int64.  v6.7 deliberately normalizes report/document years to canonical
    # strings so they can coexist with relative labels before resolution.  On
    # pandas 3, assigning "2023" into an int64 block is a hard TypeError.
    # Coerce only identity/presentation fields; financial value columns remain
    # untouched and no source amount is rewritten.
    for period_column in (
        "document_year", "report_year", "data_year", "year",
        "source_period_label",
    ):
        if period_column in raw.columns:
            raw[period_column] = raw[period_column].astype("string")
    manifest = dict(manifest)
    sources = [dict(s) for s in (manifest.get("sources") or [])]

    for source in sources:
        run_id = str(source.get("capture_run_id") or "")
        doc_year = _resolve_source_document_year(source)
        if doc_year:
            source["document_year"] = doc_year

        if "capture_run_id" not in raw.columns:
            continue
        mask = raw["capture_run_id"].astype(str) == run_id
        if not mask.any():
            continue

        if doc_year and "document_year" in raw.columns:
            raw.loc[mask, "document_year"] = doc_year

        repaired = _resolve_relative_period_years(
            raw.loc[mask].copy(),
            source.get("document_year"),
        )
        for col in repaired.columns:
            if col not in raw.columns:
                raw[col] = None
        raw.loc[mask, repaired.columns] = repaired.values

        _validate_absolute_year_resolution(
            raw.loc[mask],
            document_year=source.get("document_year"),
            capture_run_id=run_id,
        )

    manifest["sources"] = sources
    manifest["period_resolution_policy"] = (
        "RELATIVE_PERIOD_TO_ABSOLUTE_YEAR_BEFORE_CANONICAL_MERGE"
    )
    return raw, manifest


def _apply_merge_row_exclusions(
    raw: pd.DataFrame,
    metadata: dict[str, Any],
    capture_dir: Path,
) -> pd.DataFrame:
    exclusions = list(metadata.get("merge_row_exclusions") or [])
    metadata["merge_row_exclusions_applied"] = []
    metadata["merge_row_excluded_cell_count"] = 0
    if not exclusions:
        return raw

    required = {
        "table_block_id", "row_order", "column_ordinal",
        "normalized_item", "value_raw", "value",
    }
    missing = sorted(required - set(raw.columns))
    if missing:
        raise ValueError(
            f"MERGE_ROW_EXCLUSION_COLUMNS_MISSING:{capture_dir.name}:{missing}"
        )

    row_orders = pd.to_numeric(raw["row_order"], errors="coerce")
    column_ordinals = pd.to_numeric(raw["column_ordinal"], errors="coerce")
    block_ids = raw["table_block_id"].fillna("").astype(str)
    excluded_indices: set[Any] = set()
    exclusion_keys: set[tuple[str, int, int]] = set()
    applied: list[dict[str, Any]] = []

    for exclusion in exclusions:
        capture_run_id = str(exclusion.get("capture_run_id") or "")
        if capture_run_id and capture_run_id != capture_dir.name:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_CAPTURE_MISMATCH:{capture_dir.name}:{capture_run_id}"
            )
        try:
            block_id = str(exclusion["table_block_id"])
            row_order = int(exclusion["row_order"])
            column_ordinal = int(exclusion["column_ordinal"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_KEY_INVALID:{capture_dir.name}:{exclusion}"
            ) from exc
        key = (block_id, row_order, column_ordinal)
        if key in exclusion_keys:
            raise ValueError(
                f"DUPLICATE_MERGE_ROW_EXCLUSION:{capture_dir.name}:{key}"
            )
        exclusion_keys.add(key)
        mask = (
            block_ids.eq(block_id)
            & row_orders.eq(row_order)
            & column_ordinals.eq(column_ordinal)
        )
        matches = list(raw.index[mask])
        if len(matches) != 1:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_DRIFT:{capture_dir.name}:{key}:matches={len(matches)}"
            )
        index = matches[0]
        if index in excluded_indices:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_OVERLAP:{capture_dir.name}:{key}"
            )
        excluded_indices.add(index)
        source_row = raw.loc[index]
        expected_item_value = exclusion.get("normalized_item")
        actual_item_value = source_row.get("normalized_item")
        expected_raw_value = exclusion.get("value_raw")
        actual_raw_value = source_row.get("value_raw")
        expected_item = "" if expected_item_value is None else str(expected_item_value).strip()
        actual_item = (
            "" if actual_item_value is None or pd.isna(actual_item_value)
            else str(actual_item_value).strip()
        )
        expected_raw = "" if expected_raw_value is None else str(expected_raw_value).strip()
        actual_raw = (
            "" if actual_raw_value is None or pd.isna(actual_raw_value)
            else str(actual_raw_value).strip()
        )
        expected_number = exclusion.get("parsed_number")
        actual_number = pd.to_numeric(
            pd.Series([source_row.get("value")]), errors="coerce"
        ).iloc[0]
        evidence_mismatch = (
            not expected_item
            or actual_item != expected_item
            or actual_raw != expected_raw
            or expected_number is None
            or pd.isna(actual_number)
            or abs(float(actual_number) - float(expected_number))
            > max(1e-9, abs(float(expected_number)) * 1e-12)
        )
        if evidence_mismatch:
            raise ValueError(
                f"MERGE_ROW_EXCLUSION_EVIDENCE_MISMATCH:{capture_dir.name}:{key}:"
                f"item={actual_item!r}/{expected_item!r}:"
                f"raw={actual_raw!r}/{expected_raw!r}:"
                f"value={actual_number!r}/{expected_number!r}"
            )
        applied.append({
            **dict(exclusion),
            "capture_run_id": capture_dir.name,
            "table_block_id": block_id,
            "row_order": row_order,
            "column_ordinal": column_ordinal,
            "source_row_id": str(source_row.get("row_id") or ""),
            "source_value_raw": source_row.get("value_raw"),
            "source_value": source_row.get("value"),
            "exclusion_policy": "NON_SOURCE_DERIVED_OBSERVATION",
        })

    metadata["merge_row_exclusions_applied"] = applied
    metadata["merge_row_excluded_cell_count"] = len(applied)
    return raw.drop(index=list(excluded_indices)).reset_index(drop=True)


def load_capture_long(capture_dir: Path, metadata: dict[str, Any], table_id: str) -> pd.DataFrame:
    capture_meta_path = Path(capture_dir) / "capture_metadata.json"
    if capture_meta_path.exists():
        try:
            capture_meta = json.loads(capture_meta_path.read_text(encoding="utf-8"))
        except Exception:
            capture_meta = {}
        lifecycle = str(capture_meta.get("lifecycle_status") or "ACTIVE")
        if lifecycle != "ACTIVE":
            raise ValueError(
                f"CAPTURE_LIFECYCLE_BLOCKED：{Path(capture_dir).name} 当前状态={lifecycle}，"
                "不能进入正式 canonical Merge。请在数据资产管理中心恢复为 ACTIVE 后再合表。"
            )

    path = capture_dir / "table_raw_long.csv"
    if not path.exists():
        raise FileNotFoundError(f"缺少 {path}")
    df = pd.read_csv(path)
    if df.empty:
        return df
    df = _apply_merge_row_exclusions(df, metadata, capture_dir)
    if df.empty:
        return df

    # Preserve capture-layer mapping fields without colliding with merge-layer
    # canonicalization columns.
    rename_map = {}
    if "canonical_item" in df.columns:
        rename_map["canonical_item"] = "capture_canonical_item"
    if "mapping_status" in df.columns:
        rename_map["mapping_status"] = "capture_mapping_status"
    if "mapping_note" in df.columns:
        rename_map["mapping_note"] = "capture_mapping_note"
    if rename_map:
        df = df.rename(columns=rename_map)

    effective_document_year = (
        _absolute_year(metadata.get("document_year"))
        or _infer_document_year_from_capture_source(
            pdf_name=metadata.get("pdf_name"),
            capture_dir=capture_dir,
        )
        or ""
    )
    metadata["document_year"] = effective_document_year

    # Canonical Observation Contract v6.6: keep explicit source/report and
    # observation-period fields. Legacy aliases survive only for compatibility
    # with existing templates and resolvers.
    if "report_year" not in df.columns:
        df["report_year"] = effective_document_year
    else:
        df["report_year"] = effective_document_year
    if "data_year" not in df.columns:
        df["data_year"] = df.get("year", "")
    if "statement_scope" not in df.columns:
        df["statement_scope"] = df.get("scope", "")
    if "restated_flag" not in df.columns:
        df["restated_flag"] = df.get("restated", False)
    if "currency_unit" not in df.columns:
        unit_map = {
            "元": "CNY", "千元": "CNY_THOUSAND", "万元": "CNY_TEN_THOUSAND",
            "百万元": "CNY_MILLION", "亿元": "CNY_HUNDRED_MILLION", "%": "PERCENT",
        }
        df["currency_unit"] = df.get("unit", "").map(unit_map).fillna("")
    # Canonical scope values avoid mixing display labels with observation keys.
    df["statement_scope"] = df["statement_scope"].replace({"本集团": "CONSOLIDATED", "集团": "CONSOLIDATED", "本公司": "COMPANY", "公司": "COMPANY"})

    df.insert(0, "capture_run_id", metadata["capture_run_id"])
    df.insert(1, "company", metadata.get("company", ""))
    df.insert(2, "document_year", effective_document_year)
    df.insert(3, "table_id", table_id)
    # Do not infer this identity from a detail item later in canonicalization.
    # It is the semantic source of the entire Capture.
    df["table_family"] = _identity_text(metadata.get("table_family")) or str(table_id)
    df["member_table"] = _identity_text(metadata.get("member_table")) or _identity_text(metadata.get("table_query"))
    df["member_table_role"] = _identity_text(metadata.get("member_table_role")) or "COMPONENT"
    df["source_table_title"] = _identity_text(metadata.get("source_table_title")) or df["member_table"]
    df["note_reference"] = _identity_text(metadata.get("note_reference")) or _identity_text(metadata.get("note_number"))
    df["source_pdf"] = _identity_text(metadata.get("source_pdf")) or _identity_text(metadata.get("pdf_name"))
    df["member_table_order"] = metadata.get("member_table_order")
    block_defaults = {
        "container_id": _identity_text(metadata.get("container_id") or metadata.get("note_container_id")),
        "table_block_id": _identity_text(metadata.get("table_block_id") or metadata.get("member_subtable_id")),
        "block_order": (
            metadata.get("block_order")
            if metadata.get("block_order") not in (None, "")
            else -1
        ),
        "classification_axis": _identity_text(metadata.get("classification_axis")) or "UNRESOLVED",
        "block_role": _identity_text(metadata.get("block_role")) or "UNRESOLVED",
        "block_terminal_type": _identity_text(metadata.get("block_terminal_type")) or "UNRESOLVED",
    }
    for column, default in block_defaults.items():
        if column not in df.columns:
            df[column] = default
        else:
            if column == "block_order":
                df[column] = pd.to_numeric(df[column], errors="coerce").fillna(
                    -1 if default in (None, "") else default
                )
            else:
                empty = df[column].isna() | df[column].astype(str).str.strip().isin(
                    {"", "nan", "None", "<NA>"}
                )
                df.loc[empty, column] = default
    if "period_type" not in df.columns:
        df["period_type"] = _identity_text(metadata.get("period_type")) or "ANNUAL"
    if "currency" not in df.columns:
        df["currency"] = _identity_text(metadata.get("currency"))

    # Keep legacy aliases synchronized after source metadata is authoritative.
    df["year"] = df["data_year"]
    df["scope"] = df["statement_scope"]
    df["restated"] = df["restated_flag"]

    df = _resolve_relative_period_years(df, effective_document_year)
    # The resolver canonicalises legacy `year`; make the explicit v6.6
    # observation field authoritative after that resolution as well.
    df["data_year"] = df["year"].map(lambda value: _absolute_year(value) or value)
    df["report_year"] = df["report_year"].map(lambda value: _absolute_year(value) or value)
    _validate_absolute_year_resolution(
        df,
        document_year=effective_document_year,
        capture_run_id=str(metadata["capture_run_id"]),
    )
    # Derived row-path metadata is computed from raw capture exports only.  It
    # never rewrites table_capture_result.json or machine evidence files.
    df = ensure_row_paths(df)
    df = assign_conditional_source_keys(df)
    return df


def build_mapping_queue(
    raw_long: pd.DataFrame,
    table_id: str,
    taxonomy: Optional[dict[str, Any]] = None,
) -> pd.DataFrame:
    taxonomy = taxonomy or {"tables": {}}
    lookup = _taxonomy_lookup(taxonomy, table_id)

    numeric = raw_long[raw_long["value"].notna()].copy()
    groups = []
    for source_key, g in numeric.groupby("source_key", sort=False):
        parent_section = normalize_section(g["parent_section"].iloc[0] if "parent_section" in g else "")
        normalized_item = str(g["normalized_item"].iloc[0] or "")
        companies = sorted({str(x) for x in g["company"].dropna().tolist() if str(x).strip()})
        raw_examples = []
        for x in g["raw_item"].dropna().tolist():
            sx = str(x)
            if sx not in raw_examples:
                raw_examples.append(sx)

        existing = lookup.get(source_key)
        if existing:
            canonical_section = existing.get("canonical_section", parent_section)
            canonical_item = existing.get("canonical_item", normalized_item)
            category = existing.get("category", "")
            status = "AUTO_TAXONOMY"
            suggested = canonical_item
            suggestion_score = 1.0
        else:
            canonical_section = parent_section
            canonical_item = normalized_item
            category = ""
            # Exact same normalized item/context across >=2 captures is safe identity alignment.
            capture_count = g["capture_run_id"].nunique()
            if capture_count >= 2:
                status = "AUTO_EXACT_IDENTITY"
            else:
                status = "UNMAPPED_PRESERVED"
            suggested = ""
            suggestion_score = None

        groups.append({
            "source_key": source_key,
            "parent_section": parent_section,
            "normalized_item": normalized_item,
            "occurrences": int(len(g)),
            "capture_count": int(g["capture_run_id"].nunique()),
            "companies": " | ".join(companies),
            "example_raw_items": " | ".join(raw_examples[:6]),
            "suggested_canonical_section": canonical_section,
            "suggested_canonical_item": suggested,
            "suggestion_score": suggestion_score,
            "canonical_section": canonical_section,
            "canonical_item": canonical_item,
            "category": category,
            "mapping_status": status,
            "mapping_note": "",
        })

    queue = pd.DataFrame(groups)
    if queue.empty:
        return queue

    # Suggest-only fuzzy matches among source labels and known taxonomy canonicals.
    candidates = set(queue["normalized_item"].astype(str).tolist())
    table_tax = taxonomy.get("tables", {}).get(table_id, {})
    for m in table_tax.get("mappings", []):
        if m.get("canonical_item"):
            candidates.add(str(m["canonical_item"]))

    for idx, row in queue.iterrows():
        if row["mapping_status"] != "UNMAPPED_PRESERVED":
            continue
        item = str(row["normalized_item"])
        best_name, best_score = "", 0.0
        for candidate in candidates:
            if candidate == item:
                continue
            score = difflib.SequenceMatcher(None, normalize_text(item), normalize_text(candidate)).ratio()
            if score > best_score:
                best_name, best_score = candidate, score
        if best_score >= 0.72:
            queue.at[idx, "suggested_canonical_item"] = best_name
            queue.at[idx, "suggestion_score"] = round(best_score, 4)

    return queue


def apply_mapping(raw_long: pd.DataFrame, mapping_queue: pd.DataFrame) -> pd.DataFrame:
    if raw_long.empty:
        return raw_long.copy()

    map_cols = [
        "source_key", "canonical_section", "canonical_item", "category",
        "mapping_status", "mapping_note",
    ]
    mapping = mapping_queue[map_cols].drop_duplicates("source_key")
    # Capture Long already carries machine-level canonical/mapping placeholders.
    # The reviewed mapping queue is authoritative at merge time; remove those
    # placeholders before joining so pandas does not create unusable _x/_y
    # columns and silently drop the canonical contract expected below.
    mapping_payload_columns = [
        column for column in map_cols[1:] if column in raw_long.columns
    ]
    out = raw_long.drop(columns=mapping_payload_columns).merge(
        mapping,
        on="source_key",
        how="left",
    )

    out["canonical_section"] = out["canonical_section"].fillna(
        out["parent_section"].fillna("").map(normalize_section)
    )
    out["canonical_item"] = out["canonical_item"].fillna(out["normalized_item"])
    out["mapping_status"] = out["mapping_status"].fillna("UNMAPPED_PRESERVED")
    out["category"] = out["category"].fillna("")

    accepted = {
        "AUTO_TAXONOMY",
        "AUTO_EXACT_IDENTITY",
        "CONFIRMED",
        "CONFIRMED_OVERRIDE",
    }

    def key_for(row) -> str:
        section = normalize_section(row.get("canonical_section"))
        item = str(row.get("canonical_item") or row.get("normalized_item") or "").strip()
        if row.get("mapping_status") in accepted:
            prefix = "CANON"
        else:
            prefix = "RAW"
        row_path = str(row.get("row_path") or row.get("normalized_item") or "").strip()
        family = _identity_text(row.get("table_family"))
        member = _identity_text(row.get("member_table"))
        role = _identity_text(row.get("member_table_role"))
        # Missing member identity is deliberately isolated by Capture rather
        # than collapsed into a false value comparison. materialize_canonical
        # will emit REVIEW_REQUIRED_SOURCE_IDENTITY for it.
        if not family or not member or not role:
            source_scope = f"MISSING_SOURCE::{_identity_text(row.get('capture_run_id')) or 'UNKNOWN'}"
        else:
            source_scope = f"FAMILY::{family}::MEMBER::{member}::ROLE::{role}"
        axis = _identity_text(row.get("classification_axis"))
        if axis and axis != "UNRESOLVED":
            source_scope += f"::AXIS::{axis}"
        else:
            block_id = _identity_text(row.get("table_block_id"))
            if block_id:
                source_scope += f"::BLOCK::{block_id}::AXIS::UNRESOLVED"
        return f"{prefix}::{row.get('table_id')}::{source_scope}::{section}::{item}::{row_path}"

    out["canonical_key"] = out.apply(key_for, axis=1)
    return out


def _dimension_label(row: pd.Series) -> str:
    values = {
        "company": row.get("company"),
        "report_year": row.get("report_year", row.get("document_year")),
        "data_year": row.get("data_year", row.get("year")),
        "period_type": row.get("period_type"),
        "currency_unit": row.get("currency_unit"),
        "statement_scope": row.get("statement_scope", row.get("scope")),
        "restated_flag": row.get("restated_flag", row.get("restated")),
        "measure": row.get("measure"),
    }
    parts = []
    for key, value in values.items():
        if key == "restated_flag":
            parts.append(f"{key}={'TRUE' if _as_bool(value) else 'FALSE'}")
        elif value is not None and str(value).strip() not in {"", "nan", "None"}:
            parts.append(f"{key}={str(value).strip()}")
    return " | ".join(parts)


def _as_bool(value: Any) -> bool:
    """Interpret persisted CSV flags without treating the string 'False' as true."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "已重述", "经重述", "重述"}
    return bool(value)


def _dimension_missing(value: Any) -> bool:
    return value is None or (isinstance(value, float) and pd.isna(value)) or str(value).strip() in {"", "nan", "None", "<NA>"}


def _unique_source_sequence(mapped_long: pd.DataFrame, capture_run_id: str) -> list[dict[str, Any]]:
    """
    Collapse long-form period rows to one structural row per original row_order.
    Preserve exact source order.
    """
    g = mapped_long[mapped_long["capture_run_id"].astype(str) == str(capture_run_id)].copy()
    if g.empty:
        return []

    g["_row_order_num"] = pd.to_numeric(g.get("row_order"), errors="coerce")
    g = g.sort_values(["_row_order_num"], kind="stable")

    rows = []
    seen_row_orders = set()
    for _, r in g.iterrows():
        ro = r.get("_row_order_num")
        if pd.isna(ro):
            continue
        ro_int = int(ro)
        if ro_int in seen_row_orders:
            continue
        seen_row_orders.add(ro_int)
        key = str(r.get("canonical_key") or "").strip()
        if not key:
            continue
        rows.append({
            "canonical_key": key,
            "row_order": ro_int,
            "row_type": str(r.get("row_type") or ""),
            "row_level": r.get("row_level"),
            "canonical_section": str(r.get("canonical_section") or ""),
            "canonical_item": str(r.get("canonical_item") or r.get("normalized_item") or ""),
            "parent_section": str(r.get("parent_section") or ""),
            "raw_item": str(r.get("raw_item") or ""),
            "table_family": _identity_text(r.get("table_family")),
            "member_table": _identity_text(r.get("member_table")),
            "member_table_role": _identity_text(r.get("member_table_role")),
            "source_table_title": _identity_text(r.get("source_table_title")),
            "note_reference": _identity_text(r.get("note_reference")),
            "member_table_order": r.get("member_table_order"),
            "row_path": _identity_text(r.get("row_path")),
        })
    return rows


def _dedupe_key_sequence(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    seq = []
    first_meta = {}
    duplicates = []
    for row in rows:
        key = row["canonical_key"]
        if key in first_meta:
            duplicates.append({
                "conflict_type": "DUPLICATE_CANONICAL_KEY_IN_SOURCE",
                "canonical_key": key,
                "first_row_order": first_meta[key]["row_order"],
                "duplicate_row_order": row["row_order"],
                "detail": (
                    "同一来源中多个原始行被映射到同一 canonical_key；"
                    "顺序层仅保留首次出现位置，数值层仍由 VALUE_CONFLICT/UNIT_CONFLICT 独立检查。"
                ),
            })
            continue
        first_meta[key] = row
        seq.append(key)
    return seq, duplicates


def _shared_order_conflicts(
    reference_seq: list[str],
    other_seq: list[str],
    capture_run_id: str,
) -> list[dict[str, Any]]:
    """
    Detect inversions among shared keys. Reference order is authoritative and
    will never be reordered to satisfy a conflicting source.
    """
    ref_pos = {k: i for i, k in enumerate(reference_seq)}
    shared = [k for k in other_seq if k in ref_pos]
    conflicts = []
    max_seen = -1
    max_key = None
    for key in shared:
        pos = ref_pos[key]
        if pos < max_seen:
            conflicts.append({
                "conflict_type": "ORDER_CONFLICT",
                "capture_run_id": capture_run_id,
                "canonical_key": key,
                "reference_predecessor_key": max_key,
                "detail": (
                    "该来源对共同项目的相对顺序与排序基准表冲突；"
                    "最终合表继续严格采用排序基准表顺序，不自动重排。"
                ),
            })
        if pos > max_seen:
            max_seen = pos
            max_key = key
    return conflicts


def _merge_missing_keys_preserving_context(
    base: list[str],
    incoming: list[str],
) -> list[str]:
    """
    Insert source-unique keys around the nearest already-known anchors while
    preserving their local source order. Existing keys are never reordered.

    This prevents unique child/detail rows from being dumped at the end of the
    merged table.
    """
    out = list(base)
    incoming = [k for i, k in enumerate(incoming) if k and k not in incoming[:i]]

    i = 0
    while i < len(incoming):
        if incoming[i] in out:
            i += 1
            continue

        # Collect a consecutive missing block.
        j = i
        block = []
        while j < len(incoming) and incoming[j] not in out:
            block.append(incoming[j])
            j += 1

        prev_known = None
        for k in reversed(incoming[:i]):
            if k in out:
                prev_known = k
                break

        next_known = None
        for k in incoming[j:]:
            if k in out:
                next_known = k
                break

        if prev_known is not None and next_known is not None:
            prev_idx = out.index(prev_known)
            next_idx = out.index(next_known)
            if prev_idx < next_idx:
                # Insert immediately before next anchor, keeping block order.
                insert_at = next_idx
            else:
                # Incoming source itself conflicts with established order.
                # Never reorder established keys; place after previous anchor.
                insert_at = prev_idx + 1
        elif prev_known is not None:
            insert_at = out.index(prev_known) + 1
            # Keep prior insertions after the same anchor in stable order.
            while insert_at < len(out) and out[insert_at] in incoming[:i]:
                insert_at += 1
        elif next_known is not None:
            insert_at = out.index(next_known)
        else:
            insert_at = len(out)

        out[insert_at:insert_at] = block
        i = j

    return out


def build_structural_order(
    mapped_long: pd.DataFrame,
    manifest: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build one authoritative canonical row order.

    Policy:
    1. Reference capture order is immutable.
    2. Other captures may contribute missing keys around nearest known anchors.
    3. Shared-key inversions are reported as ORDER_CONFLICT, never used to
       reorder the reference.
    """
    if mapped_long.empty:
        return (
            pd.DataFrame(columns=[
                "canonical_order", "canonical_key", "row_type", "row_level",
                "canonical_section", "canonical_item", "parent_section", "order_source",
                "reference_row_order",
            ]),
            pd.DataFrame(),
        )

    manifest_sources = [x for x in (manifest.get("sources") or []) if x.get("capture_run_id")]
    # A Family is ordered by its explicit member order, then by the stable
    # caller/plan order. Never use one member's rows as the reference rows for
    # a different member table.
    def source_order(item: tuple[int, dict[str, Any]]) -> tuple[int, int]:
        index, source = item
        try:
            return (int(source.get("member_table_order")), index)
        except (TypeError, ValueError):
            return (10**9, index)
    manifest_sources = [source for _, source in sorted(enumerate(manifest_sources), key=source_order)]
    source_ids = [str(x.get("capture_run_id")) for x in manifest_sources]
    if not source_ids:
        source_ids = list(dict.fromkeys(mapped_long["capture_run_id"].astype(str).tolist()))

    requested_ref = str(manifest.get("reference_capture_run_id") or "")
    reference_id = requested_ref if requested_ref in source_ids else source_ids[0]

    sequences = {}
    row_meta_by_source = {}
    conflicts = []

    for source_id in source_ids:
        rows = _unique_source_sequence(mapped_long, source_id)
        seq, dup_conflicts = _dedupe_key_sequence(rows)
        for c in dup_conflicts:
            c["capture_run_id"] = source_id
        conflicts.extend(dup_conflicts)
        sequences[source_id] = seq
        row_meta_by_source[source_id] = {r["canonical_key"]: r for r in rows if r["canonical_key"]}

    base: list[str] = []
    order_source: dict[str, str] = {}
    note_ordinal_map: dict[str, Any] = {}
    order_policy = str(
        manifest.get("order_policy")
        or "REFERENCE_CAPTURE_PRESERVE_WITH_CONTEXTUAL_INSERTION"
    )
    reference_year = str(manifest.get("reference_report_year") or "").strip()
    if order_policy == NOTE_ORDINAL_ORDER_POLICY:
        built = _note_ordinal_base_order(
            mapped_long, manifest, reference_year,
            row_meta_by_source=row_meta_by_source,
            reference_id=reference_id,
        )
        if built is not None:
            base, order_source, note_ordinal_map = built
    if not base:
        base = list(sequences.get(reference_id, []))
        order_source = {k: f"REFERENCE:{reference_id}" for k in base}

    for source_id in source_ids:
        if source_id == reference_id:
            continue
        seq = sequences.get(source_id, [])
        conflicts.extend(_shared_order_conflicts(base, seq, source_id))
        before = set(base)
        merged = _merge_missing_keys_preserving_context(base, seq)
        for k in merged:
            if k not in before and k not in order_source:
                order_source[k] = f"INSERTED_FROM:{source_id}"
        base = merged

    # Include any unexpected keys not represented in manifest sources.
    all_keys = list(dict.fromkeys(mapped_long["canonical_key"].dropna().astype(str).tolist()))
    for key in all_keys:
        if key not in base:
            base.append(key)
            order_source[key] = "APPENDED_UNREFERENCED"

    # Structural metadata: reference source first, then first source containing key.
    order_rows = []
    ref_meta = row_meta_by_source.get(reference_id, {})
    for idx, key in enumerate(base, start=1):
        meta = ref_meta.get(key)
        meta_source = reference_id if meta else None
        if meta is None:
            for source_id in source_ids:
                if key in row_meta_by_source.get(source_id, {}):
                    meta = row_meta_by_source[source_id][key]
                    meta_source = source_id
                    break
        meta = meta or {}
        order_rows.append({
            "canonical_order": idx,
            "canonical_key": key,
            "row_type": meta.get("row_type", ""),
            "row_level": meta.get("row_level"),
            "canonical_section": meta.get("canonical_section", ""),
            "parent_section": meta.get("parent_section", ""),
            "canonical_item": meta.get("canonical_item", ""),
            "table_family": meta.get("table_family", ""),
            "member_table": meta.get("member_table", ""),
            "member_table_role": meta.get("member_table_role", ""),
            "source_table_title": meta.get("source_table_title", ""),
            "note_reference": meta.get("note_reference", ""),
            "member_table_order": meta.get("member_table_order"),
            "row_path": meta.get("row_path", ""),
            "note_ordinal": note_ordinal_map.get(key),
            "order_source": order_source.get(key, f"FIRST_SEEN:{meta_source or 'UNKNOWN'}"),
            "reference_capture_run_id": reference_id,
            "reference_row_order": (
                ref_meta.get(key, {}).get("row_order")
                if key in ref_meta else None
            ),
            "metadata_source_capture_run_id": meta_source,
        })

    order_df = pd.DataFrame(order_rows)
    conflict_cols = [
        "conflict_type", "capture_run_id", "canonical_key",
        "reference_predecessor_key", "first_row_order",
        "duplicate_row_order", "detail",
    ]
    conflicts_df = pd.DataFrame(conflicts)
    for col in conflict_cols:
        if col not in conflicts_df.columns:
            conflicts_df[col] = None
    conflicts_df = conflicts_df[conflict_cols]

    return order_df, conflicts_df


NOTE_ORDINAL_ORDER_POLICY = "NOTE_ORDINAL_REFERENCE_YEAR"


def _note_ordinal_value(note_reference: Any) -> int | None:
    from table_boundary_resolver import parse_note_ordinal

    try:
        return parse_note_ordinal(str(note_reference or "").strip())
    except Exception:
        return None


def _note_ordinal_base_order(
    mapped_long: pd.DataFrame,
    manifest: dict[str, Any],
    reference_year: str,
    *,
    row_meta_by_source: dict[str, dict[str, dict[str, Any]]],
    reference_id: str,
) -> tuple[list[str], dict[str, str], dict[str, int | None]] | None:
    """Order members by the note ordinal of a user-selected reference year.

    Top-level order follows the selected year's note sequence in the annual
    report (via ``note_reference`` ordinal); within a member, rows keep the
    selected-year capture order, then the reference capture, then first seen.
    Members with no capture in the selected year are appended by their
    member_table_order hint.  Returns ``None`` when the policy cannot be
    applied so the caller falls back to the legacy reference-capture policy.
    """
    if not reference_year:
        return None
    year_sources = [
        source for source in (manifest.get("sources") or [])
        if str(
            source.get("document_year") or source.get("report_year") or ""
        ) == reference_year
    ]
    if not year_sources:
        return None
    year_run_ids = {
        str(source.get("capture_run_id")) for source in year_sources
    }
    member_order_hint: dict[str, int] = {}
    for source in year_sources:
        member = str(source.get("member_table") or "")
        try:
            member_order_hint.setdefault(
                member, int(source.get("member_table_order") or 0)
            )
        except (TypeError, ValueError):
            member_order_hint.setdefault(member, 10**9)

    key_member: dict[str, str] = {}
    key_first_order: dict[str, int] = {}
    key_year_order: dict[str, int] = {}
    key_ref_order: dict[str, int] = {}
    for source_id, meta_by_key in row_meta_by_source.items():
        for key, meta in meta_by_key.items():
            member = str(meta.get("member_table") or "")
            if member and key not in key_member:
                key_member[key] = member
            try:
                row_order = int(meta.get("row_order") or 0)
            except (TypeError, ValueError):
                row_order = 10**9
            key_first_order.setdefault(key, row_order)
            if source_id in year_run_ids:
                key_year_order.setdefault(key, row_order)
            if source_id == reference_id:
                key_ref_order.setdefault(key, row_order)

    member_note: dict[str, tuple[int, str]] = {}
    for source in year_sources:
        member = str(source.get("member_table") or "")
        note_ref = str(source.get("note_reference") or "")
        ordinal = _note_ordinal_value(note_ref)
        if ordinal is None:
            continue
        current = member_note.get(member)
        if current is None or ordinal < current[0]:
            member_note[member] = (ordinal, note_ref)
    all_members = sorted({member for member in key_member.values() if member})
    members_with_note = sorted(
        member_note.items(),
        key=lambda item: (
            item[1][0], member_order_hint.get(item[0], 0),
        ),
    )
    members_without_note = sorted(
        [member for member in all_members if member not in member_note],
        key=lambda member: (member_order_hint.get(member, 0), member),
    )
    ordered_members = [
        member for member, _ in members_with_note
    ] + members_without_note

    base: list[str] = []
    order_source: dict[str, str] = {}
    note_ordinal_map: dict[str, int | None] = {}
    for member in ordered_members:
        note_ord, note_ref = member_note.get(member, (None, ""))
        member_keys = sorted(
            [
                key for key, owner in key_member.items()
                if owner == member
            ],
            key=lambda key: (
                key_year_order.get(key, 10**9),
                key_ref_order.get(key, 10**9),
                key_first_order.get(key, 10**9),
                key,
            ),
        )
        for key in member_keys:
            base.append(key)
            order_source[key] = (
                f"NOTE_ORDINAL:{note_ref or 'UNRESOLVED'}:{reference_year}"
            )
            note_ordinal_map[key] = note_ord
    return base, order_source, note_ordinal_map


def materialize_canonical(
    mapped_long: pd.DataFrame,
    structural_order: Optional[pd.DataFrame] = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    numeric = mapped_long[mapped_long["value"].notna()].copy()
    if numeric.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    dims = [
        "table_id", "table_family", "member_table", "member_table_role",
        *CANONICAL_BLOCK_IDENTITY_COLUMNS,
        "source_table_title", "row_path", "canonical_key", "canonical_section", "canonical_item",
        "company", "report_year", "data_year", "statement_scope", "restated_flag",
        "period_type", "currency_unit", "unit",
        "measure",
    ]
    # Accept old capture exports while always materialising the explicit v6.6
    # canonical observation fields.
    aliases = {
        "report_year": "document_year", "data_year": "year",
        "statement_scope": "scope", "restated_flag": "restated",
    }
    for canonical, legacy in aliases.items():
        if canonical not in numeric.columns and legacy in numeric.columns:
            numeric[canonical] = numeric[legacy]
    if "currency_unit" not in numeric.columns:
        unit_map = {"元": "CNY", "千元": "CNY_THOUSAND", "万元": "CNY_TEN_THOUSAND", "百万元": "CNY_MILLION", "亿元": "CNY_HUNDRED_MILLION", "%": "PERCENT"}
        numeric["currency_unit"] = (numeric["unit"].map(unit_map).fillna("") if "unit" in numeric.columns else "")
    for column, default in {
        "table_family": "", "member_table": "", "member_table_role": "",
        "container_id": "", "table_block_id": "", "block_order": -1,
        "classification_axis": "UNRESOLVED", "block_role": "UNRESOLVED",
        "block_terminal_type": "UNRESOLVED",
        "row_path": "", "period_type": "", "currency_unit": "", "unit": "", "measure": "",
        "source_table_title": "", "note_reference": "", "source_pdf": "",
    }.items():
        if column not in numeric.columns:
            numeric[column] = default
        elif column in BLOCK_IDENTITY_COLUMNS:
            # Non-compound callers legitimately have no block identity.
            # Normalise nulls to explicit legacy-safe sentinels before the
            # wide groupby/pivot; pandas otherwise drops the complete numeric
            # observation merely because one grouping dimension is null.
            numeric[column] = numeric[column].where(
                numeric[column].notna(),
                default,
            )
    # Old merge projects may be refreshed from CSV artifacts whose years were
    # parsed as floats and whose flags were persisted as display strings.
    # Normalise these fields before they become a wide-column identity.
    numeric["report_year"] = numeric["report_year"].map(lambda value: _absolute_year(value) or value)
    numeric["data_year"] = numeric["data_year"].map(lambda value: _absolute_year(value) or value)
    numeric["statement_scope"] = numeric["statement_scope"].replace({
        "本集团": "CONSOLIDATED", "集团": "CONSOLIDATED",
        "本公司": "COMPANY", "公司": "COMPANY",
    })
    numeric["restated_flag"] = numeric["restated_flag"].map(_as_bool)

    resolved_rows = []
    conflicts = []

    # v6.10: cross-family merge guard.  The same item label (e.g. "定期存款")
    # may appear in both financial_investment and investment_portfolio
    # families.  Detect when the same canonical_key prefix (without family
    # qualifier) maps to different table_family values and emit a warning.
    if "table_family" in numeric.columns and "canonical_item" in numeric.columns:
        item_families = (
            numeric[["canonical_item", "table_family"]]
            .drop_duplicates()
            .groupby("canonical_item")["table_family"]
            .apply(set)
        )
        cross_family_items = item_families[item_families.map(len) > 1]
        for item, families in cross_family_items.items():
            conflicts.append({
                "conflict_type": "CROSS_FAMILY_MERGE_CONFLICT",
                "canonical_item": item,
                "families": sorted(families),
                "detail": (
                    f"同一项目 '{item}' 出现在多个 Table Family 中："
                    f"{'、'.join(sorted(families))}。"
                    "不得仅按字符串相等直接合并；必须通过 bridge contract 建立跨表族映射。"
                ),
            })

    for key, g in numeric.groupby(dims, dropna=False, sort=False):
        values = [float(x) for x in g["value"].dropna().tolist()]
        unique_values = []
        for v in values:
            if not any(abs(v - u) <= max(1e-8, abs(u) * 1e-10) for u in unique_values):
                unique_values.append(v)

        conflict_reasons = []
        identity_missing = any(_source_identity_missing(row) for _, row in g.iterrows())
        if identity_missing:
            # Identity is incomplete, so this group has no right to reach value
            # conflict comparison. Rows remain auditable but require recovery.
            conflict_reasons.append("REVIEW_REQUIRED_SOURCE_IDENTITY")
        # Same key with different physical columns but incomplete dimensions is
        # not proof of inconsistent values.  Preserve evidence and request a
        # human dimension review instead of presenting it as a hard block.
        physical_columns = g.get("column_ordinal", pd.Series(dtype=float)).dropna().nunique()
        missing_dimensions = (_dimension_missing(key[dims.index("data_year")]) if "data_year" in dims else False) or (_dimension_missing(key[dims.index("statement_scope")]) if "statement_scope" in dims else False)
        dimension_ambiguous = physical_columns > 1 and missing_dimensions
        if len(unique_values) > 1 and not identity_missing:
            conflict_reasons.append("REVIEW_REQUIRED_DIMENSION_AMBIGUITY" if dimension_ambiguous else "VALUE_CONFLICT")

        base = {col: val for col, val in zip(dims, key)}
        for column in PHYSICAL_BLOCK_LINEAGE_COLUMNS:
            values = list(dict.fromkeys(
                value
                for value in g[column].tolist()
                if _identity_text(value)
            ))
            if not values:
                base[column] = ""
            elif len(values) == 1:
                base[column] = values[0]
            else:
                base[column] = "MULTIPLE[" + "|".join(map(str, values)) + "]"
        base["source_identity_status"] = (
            "MISSING_MEMBER_TABLE_IDENTITY" if identity_missing else "SOURCE_IDENTITY_COMPLETE"
        )
        base["source_count"] = int(len(g))
        base["mapping_status"] = " | ".join(sorted(set(g["mapping_status"].astype(str))))
        base["conflict_status"] = "|".join(conflict_reasons) if conflict_reasons else "OK"
        base["conflict_severity"] = (
            "WARNING" if conflict_reasons and not any(reason == "VALUE_CONFLICT" for reason in conflict_reasons)
            else ("BLOCKING" if conflict_reasons else "OK")
        )
        base["dimension_review_required"] = bool(dimension_ambiguous)
        base["final_value"] = unique_values[0] if len(unique_values) == 1 and not conflict_reasons else None
        base["source_rows"] = " | ".join(
            f"{r.capture_run_id}:p{r.page}:{r.raw_item}"
            for r in g.itertuples()
        )
        base["source_provenance"] = json.dumps([
            {
                "capture_run_id": str(r.get("capture_run_id") or ""),
                "pdf": str(r.get("source_pdf") or r.get("pdf_name") or ""),
                "page": r.get("page"),
                "bbox": r.get("bbox"),
                "context_source_page": r.get("context_source_page"),
                "container_id": r.get("container_id"),
                "table_block_id": r.get("table_block_id"),
                "block_order": r.get("block_order"),
                "classification_axis": r.get("classification_axis"),
                "block_role": r.get("block_role"),
                "block_terminal_type": r.get("block_terminal_type"),
            }
            for _, r in g.iterrows()
        ], ensure_ascii=False)
        resolved_rows.append(base)

        if conflict_reasons:
            conflicts.append({
                **base,
                "values_found": " | ".join(map(str, unique_values)),
                "units_found": str(base.get("unit") or ""),
            })

    resolved = pd.DataFrame(resolved_rows)

    order_cols = [
        "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
        "order_source", "reference_capture_run_id", "reference_row_order",
    ]
    if structural_order is not None and not structural_order.empty:
        available_order_cols = [c for c in order_cols if c in structural_order.columns]
        resolved = resolved.merge(
            structural_order[available_order_cols].drop_duplicates("canonical_key"),
            on="canonical_key",
            how="left",
        )
        resolved = resolved.sort_values(
            ["canonical_order", "company", "report_year", "data_year"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)
    else:
        resolved["canonical_order"] = range(1, len(resolved) + 1)
        resolved["row_type"] = ""
        resolved["order_source"] = "UNSPECIFIED"

    conflict_columns = list(resolved.columns) + ["values_found", "units_found"]
    conflicts_df = pd.DataFrame(conflicts, columns=conflict_columns)

    safe = resolved[
        (resolved["conflict_status"] == "OK")
        & resolved["final_value"].notna()
    ].copy()
    if safe.empty:
        empty_wide = pd.DataFrame(
            columns=[
                *BLOCK_IDENTITY_COLUMNS,
                "canonical_key", "canonical_section", "canonical_item", "unit",
            ]
        )
        return resolved, empty_wide, conflicts_df

    safe["document_column"] = safe.apply(_dimension_label, axis=1)

    # Stable one-row-per-canonical-key metadata. Do NOT use multiple metadata
    # fields directly as pivot_table index with dropna=False, which can create a
    # Cartesian product.
    key_meta_cols = [
        "table_family", "member_table", "member_table_role", "row_path",
        *BLOCK_IDENTITY_COLUMNS,
        "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
        "canonical_section", "canonical_item", "order_source", "source_identity_status",
    ]
    semantic_meta_cols = [
        column for column in key_meta_cols
        if column not in PHYSICAL_BLOCK_LINEAGE_COLUMNS
    ]
    key_meta = (
        safe[[c for c in semantic_meta_cols if c in safe.columns]]
        .drop_duplicates(subset=["canonical_key"], keep="first")
    )

    def combine_lineage(series: pd.Series) -> Any:
        values = list(dict.fromkeys(
            value for value in series.tolist() if _identity_text(value)
        ))
        if not values:
            return ""
        if len(values) == 1:
            return values[0]
        return "MULTIPLE[" + "|".join(map(str, values)) + "]"

    lineage_meta = (
        safe.groupby("canonical_key", sort=False, dropna=False)[
            PHYSICAL_BLOCK_LINEAGE_COLUMNS
        ]
        .agg(combine_lineage)
        .reset_index()
    )
    key_meta = key_meta.merge(lineage_meta, on="canonical_key", how="left")

    def combine_units(series: pd.Series) -> str:
        units = sorted({
            str(x).strip()
            for x in series.dropna().tolist()
            if str(x).strip() and str(x).lower() != "nan"
        })
        if not units:
            return ""
        if len(units) == 1:
            return units[0]
        return "REVIEW_REQUIRED[" + "|".join(units) + "]"

    wide_key_columns = ["canonical_key"]
    unit_by_key = (
        safe.groupby(wide_key_columns, sort=False)["unit"]
        .apply(combine_units)
        .rename("unit")
        .reset_index()
    )

    wide_values = safe.pivot_table(
        index=wide_key_columns,
        columns="document_column",
        values="final_value",
        aggfunc="first",
        dropna=True,
    ).reset_index()
    wide_values.columns.name = None

    wide = (
        key_meta
        .merge(unit_by_key, on=wide_key_columns, how="left")
        .merge(wide_values, on=wide_key_columns, how="left")
    )
    if "canonical_order" in wide.columns:
        wide = wide.sort_values("canonical_order", kind="stable", na_position="last").reset_index(drop=True)

    fixed = [
        "table_family", "member_table", "member_table_role", "row_path",
        *BLOCK_IDENTITY_COLUMNS,
        "canonical_order", "row_type", "row_level", "parent_section", "canonical_section",
        "canonical_item", "unit", "canonical_key", "order_source", "source_identity_status",
    ]
    fixed = [c for c in fixed if c in wide.columns]
    wide = wide[fixed + [c for c in wide.columns if c not in fixed]]

    return resolved, wide, conflicts_df


def coverage_report(raw_long: pd.DataFrame, mapped_long: pd.DataFrame, conflicts: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for run_id, g in mapped_long[mapped_long["value"].notna()].groupby("capture_run_id", sort=False):
        total = len(g)
        confirmed = g["mapping_status"].isin(
            ["AUTO_TAXONOMY", "AUTO_EXACT_IDENTITY", "CONFIRMED", "CONFIRMED_OVERRIDE"]
        ).sum()
        preserved = (g["mapping_status"] == "UNMAPPED_PRESERVED").sum()
        rows.append({
            "capture_run_id": run_id,
            "company": g["company"].iloc[0] if len(g) else "",
            "document_year": g["document_year"].iloc[0] if len(g) else "",
            "numeric_source_rows": int(total),
            "mapped_or_exact_rows": int(confirmed),
            "unmapped_preserved_rows": int(preserved),
            "mapping_coverage": float(confirmed / total) if total else 0.0,
        })

    report = pd.DataFrame(rows)
    if not report.empty:
        report["project_conflict_count"] = int(len(conflicts))
    return report


def persist_confirmed_taxonomy(
    taxonomy_path: Path,
    table_id: str,
    mapping_queue: pd.DataFrame,
) -> dict[str, Any]:
    taxonomy = load_taxonomy(taxonomy_path)
    table = taxonomy.setdefault("tables", {}).setdefault(
        table_id,
        {"mappings": []},
    )
    existing = {
        (m.get("canonical_section", ""), m.get("canonical_item", "")): m
        for m in table.get("mappings", [])
    }

    accepted = {"CONFIRMED", "CONFIRMED_OVERRIDE", "AUTO_TAXONOMY"}
    written = 0
    for row in mapping_queue.to_dict("records"):
        if row.get("mapping_status") not in accepted:
            continue
        canonical_item = str(row.get("canonical_item") or "").strip()
        if not canonical_item:
            continue
        canonical_section = normalize_section(row.get("canonical_section"))
        key = (canonical_section, canonical_item)
        entry = existing.get(key)
        if entry is None:
            entry = {
                "canonical_section": canonical_section,
                "canonical_item": canonical_item,
                "category": str(row.get("category") or ""),
                "source_keys": [],
            }
            table["mappings"].append(entry)
            existing[key] = entry
        source_key = str(row.get("source_key"))
        if source_key not in entry["source_keys"]:
            entry["source_keys"].append(source_key)
            written += 1
        if row.get("category"):
            entry["category"] = str(row.get("category"))

    _atomic_json(Path(taxonomy_path), taxonomy)
    return {"written_source_keys": written, "taxonomy_path": str(taxonomy_path)}


def collect_merge_reconciliation(manifest: dict[str, Any]) -> pd.DataFrame:
    frames=[]
    for source in manifest.get("sources") or []:
        capture_dir=Path(str(source.get("capture_dir") or ""))
        path=capture_dir/"table_reconciliation_audit.csv"
        if not path.exists():
            try:
                from reconciliation import write_reconciliation_audit
                path=write_reconciliation_audit(capture_dir)
            except Exception:
                continue
        try:
            df=pd.read_csv(path)
        except Exception:
            continue
        if df.empty:
            continue
        df.insert(0,"capture_run_id",source.get("capture_run_id"))
        df.insert(1,"company",source.get("company"))
        df.insert(2,"document_year",source.get("document_year"))
        frames.append(df)
    return pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()


def source_identity_qa(raw_long: pd.DataFrame) -> pd.DataFrame:
    """Expose source provenance before any row/value comparison is attempted."""
    columns = SOURCE_IDENTITY_COLUMNS + [
        "member_table_order", "row_path_count", "source_identity_status", "qa_detail",
    ]
    if raw_long.empty:
        return pd.DataFrame(columns=columns)
    work = raw_long.copy()
    for field in SOURCE_IDENTITY_COLUMNS:
        if field not in work.columns:
            work[field] = ""
    if "member_table_order" not in work.columns:
        work["member_table_order"] = None
    rows = []
    for capture_id, group in work.groupby("capture_run_id", dropna=False, sort=False):
        first = group.iloc[0]
        missing = [field for field in ("table_family", "member_table", "member_table_role") if not _identity_text(first.get(field))]
        rows.append({
            **{field: first.get(field) for field in SOURCE_IDENTITY_COLUMNS},
            "capture_run_id": capture_id,
            "member_table_order": first.get("member_table_order"),
            "row_path_count": int(group.get("row_path", pd.Series(dtype=str)).dropna().astype(str).nunique()),
            "source_identity_status": "MISSING_MEMBER_TABLE_IDENTITY" if missing else "SOURCE_IDENTITY_COMPLETE",
            "qa_detail": ("缺少：" + "、".join(missing)) if missing else "可进入 source-aware row alignment",
        })
    return pd.DataFrame(rows, columns=columns)


RESEARCH_WIDE_FIXED_COLUMNS = ("member_table", "canonical_item", "unit")
# Canonical row_type values treated as summary/total rows in the user-facing
# wide workbook. Identification stays at the semantic layer; label text is
# never used to guess a total.
TOTAL_ROW_TYPES = frozenset({"TOTAL", "IMPLICIT_TOTAL", "SUBTOTAL"})
RESEARCH_WIDE_FIXED_COLUMN_LABELS = {
    "member_table": "附注表名",
    "canonical_item": "项目",
    "unit": "单位",
}
RESEARCH_WIDE_METADATA_DIMENSION_LABELS = {
    "company": "公司",
    "statement_scope": "口径",
    "period_type": "期间类型",
    "currency": "币种",
    "currency_unit": "金额单位",
    "measure": "计量口径",
}


def build_research_wide_frame(research_wide: pd.DataFrame) -> pd.DataFrame:
    """Trim the research wide to identity/unit plus actual observation columns.

    The full presentation wide carries many fixed context columns (row_path,
    canonical_key, block identity, source title, ...).  For real research the
    user only needs the member, the canonical item, the unit and the values;
    the multi-level header keeps the period/company/scope/unit dimensions.
    """
    value_ids = [
        column for column in research_wide.columns
        if str(column).startswith("COL_")
    ]
    fixed = [
        column for column in RESEARCH_WIDE_FIXED_COLUMNS
        if column in research_wide.columns
    ]
    return research_wide[fixed + value_ids].copy()


def write_presentation_wide_sheet(
    writer: Any,
    research_wide: pd.DataFrame,
    fixed_columns: list[Any],
    column_dimensions: pd.DataFrame,
    header_policy: VisibleHeaderDimensionPolicy,
    sheet_name: str = "canonical_wide",
    row_types: Optional[Sequence[str]] = None,
) -> None:
    """Write the human-facing multi-level wide sheet without COL ids.

    ``row_types`` carries the Canonical semantic row type per data row as a
    styling-only hint.  Total-type rows are bolded and receive a light fill;
    the CSV data contract is untouched.
    """
    header_rows = max(1, len(header_policy.visible_header_dimensions))
    research_wide.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        header=False,
        startrow=header_rows + 2,
    )
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="研究宽表元数据")
    meta_text = "；".join(
        f"{RESEARCH_WIDE_METADATA_DIMENSION_LABELS.get(dimension, dimension)}={value}"
        for dimension, value in header_policy.metadata_values.items()
        if value
    )
    ws.cell(row=1, column=2, value=meta_text)
    start_col = len(fixed_columns) + 1
    last_col = len(fixed_columns) + len(column_dimensions)
    if last_col > 2:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    for column_index, column_name in enumerate(fixed_columns, start=1):
        ws.cell(
            row=2,
            column=column_index,
            value=RESEARCH_WIDE_FIXED_COLUMN_LABELS.get(
                str(column_name), str(column_name),
            ),
        )
        if header_rows > 1:
            ws.merge_cells(
                start_row=2,
                start_column=column_index,
                end_row=header_rows + 1,
                end_column=column_index,
            )
    dimension_runs: dict[int, list[tuple[int, int]]] = {}
    display_labels = {
        dimension: column_dimensions.get(
            f"display_{dimension}",
            pd.Series([""] * len(column_dimensions)),
        ).tolist()
        for dimension in header_policy.visible_header_dimensions
    }
    for level_index, dimension in enumerate(header_policy.visible_header_dimensions):
        row_index = level_index + 2
        labels = display_labels[dimension]
        for offset, value in enumerate(labels):
            ws.cell(row=row_index, column=start_col + offset, value=value)
        run_start = start_col
        dimension_runs[row_index] = []
        for offset in range(1, len(labels) + 1):
            parent_changed = any(
                display_labels[parent][offset] != display_labels[parent][offset - 1]
                for parent in header_policy.visible_header_dimensions[:level_index]
            ) if offset < len(labels) else False
            changed = (
                offset == len(labels)
                or labels[offset] != labels[offset - 1]
                or parent_changed
            )
            if changed:
                run_end = start_col + offset - 1
                dimension_runs[row_index].append((run_start, run_end))
                if run_end > run_start and labels[offset - 1]:
                    ws.merge_cells(
                        start_row=row_index,
                        start_column=run_start,
                        end_row=row_index,
                        end_column=run_end,
                    )
                run_start = start_col + offset

    header_end_row = header_rows + 1
    data_start_row = header_rows + 3
    title_fill = PatternFill("solid", fgColor="1F4E78")
    metadata_fill = PatternFill("solid", fgColor="EAF0F6")
    report_fill = PatternFill("solid", fgColor="D9E5F2")
    period_fill = PatternFill("solid", fgColor="EDF2F7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_side = Side(style="thin", color="AAB7C4")
    medium_side = Side(style="medium", color="5B6573")

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).font = Font(
        name="等线", size=11, bold=True, color="FFFFFF",
    )
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center",
    )
    for column_index in range(2, last_col + 1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = metadata_fill
        cell.font = Font(name="等线", size=10, color="374151")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_index in range(2, header_end_row + 1):
        ws.row_dimensions[row_index].height = 22
        for column_index in range(1, last_col + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.fill = report_fill if row_index == 2 else period_fill
            cell.font = Font(name="等线", size=10.5, bold=True, color="1F2937")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            cell.border = Border(
                top=medium_side if row_index == 2 else thin_side,
                bottom=medium_side if row_index == header_end_row else thin_side,
                left=cell.border.left,
                right=cell.border.right,
            )

    for column_index in range(1, len(fixed_columns) + 1):
        for row_index in range(2, header_end_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = Border(
                left=medium_side,
                right=medium_side,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    report_row = next(
        (
            index + 2
            for index, dimension in enumerate(header_policy.visible_header_dimensions)
            if dimension == "report_year"
        ),
        2,
    )
    for group_start, group_end in dimension_runs.get(report_row, []):
        for row_index in range(report_row, header_end_row + 1):
            left_cell = ws.cell(row=row_index, column=group_start)
            left_cell.border = Border(
                left=medium_side,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=row_index, column=group_end)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=medium_side,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )

    data_year_row = next(
        (
            index + 2
            for index, dimension in enumerate(header_policy.visible_header_dimensions)
            if dimension == "data_year"
        ),
        None,
    )
    if data_year_row is not None:
        for group_start, group_end in dimension_runs.get(data_year_row, []):
            left_cell = ws.cell(row=data_year_row, column=group_start)
            left_cell.border = Border(
                left=medium_side,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=data_year_row, column=group_end)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=medium_side,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )

    widths = (36, 30, 12)
    for column_index, width in enumerate(widths[:len(fixed_columns)], start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width
    for column_index in range(start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 16

    row_types_list = list(row_types) if row_types is not None else None
    for row_index in range(data_start_row, ws.max_row + 1):
        row_offset = row_index - data_start_row
        is_total = (
            row_types_list is not None
            and row_offset < len(row_types_list)
            and str(row_types_list[row_offset]).upper() in TOTAL_ROW_TYPES
        )
        for column_index in range(1, last_col + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if column_index <= 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif column_index == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'
            if is_total:
                cell.font = Font(name="等线", size=10.5, bold=True)
                cell.fill = total_fill

    ws.freeze_panes = f"{get_column_letter(start_col)}{data_start_row}"


def write_merge_outputs(
    output_dir: Path,
    manifest: dict[str, Any],
    raw_long: pd.DataFrame,
    mapping_queue: pd.DataFrame,
    taxonomy_path: Optional[Path] = None,
    member_display_map: Optional[dict[str, str]] = None,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)

    # v5.8 invariant: canonicalization never sees relative labels in `year`.
    # This also repairs already-created v5.7 merge_raw_long.csv on refresh.
    raw_long, manifest = _repair_manifest_and_raw_periods(raw_long, manifest)

    mapped = apply_mapping(raw_long, mapping_queue)
    structural_order, order_conflicts = build_structural_order(mapped, manifest)

    if not structural_order.empty:
        mapped = mapped.merge(
            structural_order[
                [
                    "canonical_key", "canonical_order", "row_type", "row_level", "parent_section",
                    "order_source", "reference_capture_run_id", "reference_row_order",
                ]
            ].drop_duplicates("canonical_key"),
            on="canonical_key",
            how="left",
            suffixes=("", "_order"),
        )
        # Preserve source evidence ordering inside each capture, while exposing
        # canonical_order for cross-company research views.
        mapped = mapped.sort_values(
            ["capture_run_id", "row_order", "canonical_order"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    resolved, wide, conflicts = materialize_canonical(
        mapped,
        structural_order=structural_order,
    )
    coverage = coverage_report(raw_long, mapped, conflicts)
    reconciliation = collect_merge_reconciliation(manifest)
    identity_qa = source_identity_qa(raw_long)

    # Canonical Long is the source of truth. CSV cannot represent a real
    # hierarchical header, so its wide view uses stable encoded ids and an
    # explicit dimension mapping rather than a lossy concatenated label.
    fixed_columns = [c for c in wide.columns if not str(c).startswith("company=")]
    value_columns = [c for c in wide.columns if str(c).startswith("company=")]
    column_rows = []
    rename_columns = {}
    for ordinal, column in enumerate(value_columns, start=1):
        column_id = f"COL_{ordinal:05d}"
        rename_columns[column] = column_id
        dims = {}
        for token in str(column).split(" | "):
            if "=" in token:
                key, value = token.split("=", 1); dims[key] = value
        # Currency is retained independently even when the older v6.6
        # presentation label only carried currency_unit.
        dims.setdefault("currency", "CNY" if str(dims.get("currency_unit") or "").startswith("CNY") else "")
        column_rows.append({"column_id": column_id, "source_column_label": column, **{key: dims.get(key, "") for key in OBSERVATION_DIMENSIONS}})
    research_wide = wide.rename(columns=rename_columns)
    column_dimensions = pd.DataFrame(column_rows)
    header_policy = VisibleHeaderDimensionPolicy.from_column_dimensions(column_dimensions)
    if not column_dimensions.empty:
        display_rows = [header_policy.label_for_column(row) for row in column_dimensions.to_dict("records")]
        for dimension in header_policy.visible_header_dimensions:
            column_dimensions[f"display_{dimension}"] = [row.get(dimension, "") for row in display_rows]
    # Research-wide export: only member/canonical item/unit plus values, with
    # the same multi-level header architecture as the presentation export.
    research_wide_trim = build_research_wide_frame(research_wide)
    if member_display_map:
        research_wide_trim["member_table"] = research_wide_trim[
            "member_table"
        ].map(
            lambda value: member_display_map.get(str(value), value)
        )
    research_wide_fixed = [
        column for column in RESEARCH_WIDE_FIXED_COLUMNS
        if column in research_wide_trim.columns
    ]
    # row_type is a styling-only hint for the Excel writer; it never enters
    # the trimmed research CSV.
    research_row_types = (
        research_wide["row_type"].tolist()
        if "row_type" in research_wide.columns
        else [""] * len(research_wide)
    )

    paths = {
        "manifest": output_dir / "merge_manifest.json",
        "raw_long": output_dir / "merge_raw_long.csv",
        "mapping_queue": output_dir / "merge_mapping_queue.csv",
        "canonical_long": output_dir / "merge_canonical_long.csv",
        "canonical_research_long": output_dir / "canonical_research_long.csv",
        "resolved_long": output_dir / "merge_resolved_long.csv",
        "canonical_wide": output_dir / "merge_canonical_wide.csv",
        "column_dimensions": output_dir / "column_dimensions.csv",
        "wide_metadata": output_dir / "research_wide_metadata.json",
        "research_wide_csv": output_dir / "research_wide.csv",
        "research_wide_xlsx": output_dir / "research_wide.xlsx",
        "conflicts": output_dir / "merge_conflicts.csv",
        "coverage": output_dir / "merge_coverage.csv",
        "structural_order": output_dir / "merge_structural_order.csv",
        "order_conflicts": output_dir / "merge_order_conflicts.csv",
        "reconciliation": output_dir / "merge_reconciliation_audit.csv",
        "source_identity_qa": output_dir / "merge_source_identity_qa.csv",
        "xlsx": output_dir / "merge_project.xlsx",
    }

    paths["manifest"].write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    raw_long.to_csv(paths["raw_long"], index=False, encoding="utf-8-sig")
    mapping_queue.to_csv(paths["mapping_queue"], index=False, encoding="utf-8-sig")
    mapped.to_csv(paths["canonical_long"], index=False, encoding="utf-8-sig")
    # Canonical Research Long is the source-of-truth observation layer. It
    # retains source rows and provenance; wide/resolved outputs are views.
    mapped.to_csv(paths["canonical_research_long"], index=False, encoding="utf-8-sig")
    resolved.to_csv(paths["resolved_long"], index=False, encoding="utf-8-sig")
    research_wide.to_csv(paths["canonical_wide"], index=False, encoding="utf-8-sig")
    research_wide_trim.to_csv(
        paths["research_wide_csv"], index=False, encoding="utf-8-sig",
    )
    column_dimensions.to_csv(paths["column_dimensions"], index=False, encoding="utf-8-sig")
    paths["wide_metadata"].write_text(json.dumps({"metadata_dimensions": list(header_policy.metadata_dimensions), "visible_header_dimensions": list(header_policy.visible_header_dimensions), "display_order": list(header_policy.display_order), "metadata_values": header_policy.metadata_values, "presentation_export_version": 2}, ensure_ascii=False, indent=2), encoding="utf-8")
    conflicts.to_csv(paths["conflicts"], index=False, encoding="utf-8-sig")
    coverage.to_csv(paths["coverage"], index=False, encoding="utf-8-sig")
    structural_order.to_csv(paths["structural_order"], index=False, encoding="utf-8-sig")
    order_conflicts.to_csv(paths["order_conflicts"], index=False, encoding="utf-8-sig")
    reconciliation.to_csv(paths["reconciliation"], index=False, encoding="utf-8-sig")
    identity_qa.to_csv(paths["source_identity_qa"], index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(paths["xlsx"], engine="openpyxl") as writer:
        raw_long.to_excel(writer, sheet_name="raw_long", index=False)
        mapping_queue.to_excel(writer, sheet_name="mapping_queue", index=False)
        mapped.to_excel(writer, sheet_name="canonical_long", index=False)
        mapped.to_excel(writer, sheet_name="canonical_research_long", index=False)
        resolved.to_excel(writer, sheet_name="resolved_long", index=False)
        write_presentation_wide_sheet(
            writer,
            research_wide,
            fixed_columns,
            column_dimensions,
            header_policy,
            row_types=research_row_types,
        )
        column_dimensions.to_excel(writer, sheet_name="column_dimensions", index=False)
        conflicts.to_excel(writer, sheet_name="conflicts", index=False)
        coverage.to_excel(writer, sheet_name="coverage", index=False)
        structural_order.to_excel(writer, sheet_name="structural_order", index=False)
        order_conflicts.to_excel(writer, sheet_name="order_conflicts", index=False)
        reconciliation.to_excel(writer, sheet_name="reconciliation", index=False)
        identity_qa.to_excel(writer, sheet_name="source_identity_qa", index=False)

    if research_wide_fixed:
        with pd.ExcelWriter(
            paths["research_wide_xlsx"], engine="openpyxl",
        ) as writer:
            write_presentation_wide_sheet(
                writer,
                research_wide_trim,
                research_wide_fixed,
                column_dimensions,
                header_policy,
                sheet_name="research_wide",
                row_types=research_row_types,
            )

    if taxonomy_path and Path(taxonomy_path).exists():
        snapshot = output_dir / "taxonomy_snapshot.json"
        snapshot.write_text(Path(taxonomy_path).read_text(encoding="utf-8"), encoding="utf-8")
        paths["taxonomy_snapshot"] = snapshot

    try:
        from registry_bridge import sync_merge_run
        sync_merge_run(output_dir)
    except Exception:
        pass

    return {k: str(v) for k, v in paths.items()}


def create_merge_project(
    capture_dirs: list[Path],
    metadata_rows: list[dict[str, Any]],
    output_dir: Path,
    table_id: str,
    taxonomy_path: Path,
    reference_capture_run_id: Optional[str] = None,
    merge_lineage: Optional[dict[str, Any]] = None,
    member_display_map: Optional[dict[str, str]] = None,
    order_policy: Optional[str] = None,
    reference_report_year: Optional[str] = None,
) -> dict[str, str]:
    table_id = normalize_table_id(table_id)
    metadata_map = {str(r["capture_run_id"]): r for r in metadata_rows}

    frames = []
    manifest_sources = []
    for capture_dir in capture_dirs:
        inferred = infer_capture_metadata(capture_dir)
        user_meta = metadata_map.get(capture_dir.name, {})
        meta = {**inferred, **user_meta}
        frame = load_capture_long(capture_dir, meta, table_id)
        frames.append(frame)
        manifest_sources.append(meta)

    raw_long = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    taxonomy = load_taxonomy(taxonomy_path)
    queue = build_mapping_queue(raw_long, table_id, taxonomy)

    source_ids = [str(x.get("capture_run_id")) for x in manifest_sources]
    if reference_capture_run_id not in source_ids:
        reference_capture_run_id = source_ids[0] if source_ids else None

    bundle_expansion = dict(merge_lineage or {})
    if bundle_expansion:
        applied_exclusions = [
            dict(exclusion)
            for source in manifest_sources
            for exclusion in (source.get("merge_row_exclusions_applied") or [])
        ]
        expected_count = int(bundle_expansion.get("row_cell_exclusion_count") or 0)
        if len(applied_exclusions) != expected_count:
            raise ValueError(
                "MERGE_ROW_EXCLUSION_COUNT_MISMATCH:"
                f"expected={expected_count}:applied={len(applied_exclusions)}"
            )
        bundle_expansion["row_cell_exclusions_applied"] = applied_exclusions
        bundle_expansion["row_cell_exclusions_applied_count"] = len(applied_exclusions)

    manifest = {
        "version": "v6.7",
        "merge_schema_version": "6.8_SEMANTIC_AXIS_CROSS_CAPTURE_IDENTITY",
        "canonical_observation_schema_version": "6.8_SEMANTIC_AXIS_BLOCK_LINEAGE",
        "table_id": table_id,
        "sources": manifest_sources,
        "taxonomy_path": str(taxonomy_path),
        "order_policy": "REFERENCE_CAPTURE_PRESERVE_WITH_CONTEXTUAL_INSERTION",
        "identity_policy": "TABLE_FAMILY__MEMBER_TABLE__MEMBER_ROLE__CLASSIFICATION_AXIS__ROW_PATH__DIMENSIONS__PHYSICAL_BLOCK_LINEAGE_ONLY",
        "reference_capture_run_id": reference_capture_run_id,
    }
    if member_display_map:
        manifest["member_display_map"] = dict(member_display_map)
    if order_policy:
        manifest["order_policy"] = str(order_policy)
    if reference_report_year:
        manifest["reference_report_year"] = str(reference_report_year)
    if bundle_expansion:
        manifest["bundle_expansion"] = bundle_expansion
    return write_merge_outputs(
        output_dir=output_dir,
        manifest=manifest,
        raw_long=raw_long,
        mapping_queue=queue,
        taxonomy_path=taxonomy_path,
        member_display_map=member_display_map,
    )


def refresh_merge_project(
    output_dir: Path,
    mapping_queue: Optional[pd.DataFrame] = None,
    persist_taxonomy: bool = False,
    order_policy: Optional[str] = None,
    reference_report_year: Optional[str] = None,
    member_display_map: Optional[dict[str, str]] = None,
) -> dict[str, str]:
    output_dir = Path(output_dir)
    manifest = json.loads((output_dir / "merge_manifest.json").read_text(encoding="utf-8"))
    # A refresh is also the non-destructive migration path for old derived
    # merge artifacts. Source capture evidence and human mapping decisions are
    # retained; only generated outputs are recomputed under the current
    # observation contract.
    manifest["merge_schema_version"] = "6.8_SEMANTIC_AXIS_CROSS_CAPTURE_IDENTITY"
    manifest["canonical_observation_schema_version"] = "6.8_SEMANTIC_AXIS_BLOCK_LINEAGE"
    manifest["identity_policy"] = "TABLE_FAMILY__MEMBER_TABLE__MEMBER_ROLE__CLASSIFICATION_AXIS__ROW_PATH__DIMENSIONS__PHYSICAL_BLOCK_LINEAGE_ONLY"
    if order_policy:
        manifest["order_policy"] = str(order_policy)
    if reference_report_year is not None:
        manifest["reference_report_year"] = str(reference_report_year)
    if member_display_map is not None:
        manifest["member_display_map"] = dict(member_display_map)
    raw = pd.read_csv(output_dir / "merge_raw_long.csv")
    queue = (
        mapping_queue.copy()
        if mapping_queue is not None
        else pd.read_csv(output_dir / "merge_mapping_queue.csv")
    )
    taxonomy_path = Path(manifest.get("taxonomy_path", output_dir / "table_taxonomy.json"))

    if persist_taxonomy:
        persist_confirmed_taxonomy(taxonomy_path, manifest["table_id"], queue)

    return write_merge_outputs(
        output_dir=output_dir,
        manifest=manifest,
        raw_long=raw,
        mapping_queue=queue,
        taxonomy_path=taxonomy_path,
        member_display_map=(
            member_display_map
            if member_display_map is not None
            else manifest.get("member_display_map") or None
        ),
    )
